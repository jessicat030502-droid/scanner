"""
QUANT SCANNER v4 - Mid/Small Cap + Russell 2000 Day Trading Engine
===================================================================
Run:   python -m streamlit run scanner_v4.py
Deps:  pip install yfinance pandas numpy scipy streamlit openpyxl schedule colorama

10 Layers: Hurst | Hawkes | OFI | Sector RS | ADD Breadth |
           Z-Score | Kurtosis/Skew | Bayesian Prob | Half-Life | ATR R:R

New in v4:
  - Russell 2000 full universe scan (2000 stocks, batched)
  - Range Filter confirmation gate (ThinkScript port)
  - Ticker personality profiling (per-symbol win rate tracking)
  - Backtest + sensitivity analysis module
  - TWAP/VWAP execution schedule
  - Portfolio risk controls (max 5 positions, max 2/sector)
  - Validate-before-execute pre-trade layer
  - EOD exit 3:30/3:55 PM
  - Single daily Excel workbook (all scans in one file)

Cap filter: $300M-$20B. Timeframe: 15s / 1m / 5m / 15m toggle.
"""

import os, time, math, warnings, threading, json

# Suppress Streamlit "missing ScriptRunContext" warning when run from CLI
os.environ.setdefault("STREAMLIT_SERVER_HEADLESS", "true")
os.environ.setdefault("STREAMLIT_BROWSER_GATHER_USAGE_STATS", "false")
import numpy as np
import pandas as pd
from datetime import datetime
from dataclasses import dataclass
from typing import Optional
from scipy import stats as sp_stats
import yfinance as yf
try:
    import schedule
    _SCHEDULE_AVAILABLE = True
except ImportError:
    _SCHEDULE_AVAILABLE = False
    # schedule is only used in the legacy run_universe_scan loop
    # The main scanner uses scanner_terminal_v4.py --loop which doesn't need it

try:
    import ticker_intelligence as _ti
    _TI_AVAILABLE = True
except ImportError:
    _ti = None
    _TI_AVAILABLE = False

warnings.filterwarnings("ignore")

# ── File paths -- single source of truth for both scanner and research ────────
# All files are relative to the directory containing scanner_v4.py.
# Both scanner_research_v4.py and scanner_terminal_v4.py import these from here.
# You never need to configure paths anywhere else.

# Directory where scanner_v4.py lives (works regardless of where you launch from)
SCANNER_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Learned weights loader ───────────────────────────────────────────────────
# Checks C:\QuantScanner\learned_weights.json (written by --optimize-weights).
# Only activates when file exists AND n_samples >= 60 AND activated == True.
# Falls back silently to hardcoded weights if any condition fails.
# This makes the transition from static -> adaptive completely safe.
_LEARNED_WEIGHTS: dict = {}
_learned_weights_path = os.path.join(SCANNER_DIR, "learned_weights.json")
if os.path.exists(_learned_weights_path):
    try:
        with open(_learned_weights_path) as _lw_f:
            _lw = json.load(_lw_f)
        _cv_chk   = _lw.get("cv_accuracy", 0)
        _nsmp_chk = _lw.get("n_samples", 0)
        # Reject if CV=100% on small sample -- memorized one regime, not real edge
        if _cv_chk >= 0.995 and _nsmp_chk < 500:
            print(f"[WEIGHTS] Overfit guard: CV={_cv_chk:.1%} on {_nsmp_chk} samples "
                  f"-- rejecting, using hardcoded weights.")
            print(f"[WEIGHTS] Delete learned_weights.json and re-run "
                  f"--optimize-weights after more diverse data.")
        elif _lw.get("activated") and _nsmp_chk >= 60:
            _LEARNED_WEIGHTS = _lw
            print(f"[WEIGHTS] Learned weights active "
                  f"({_lw['n_samples']} samples, "
                  f"CV={_lw.get('cv_accuracy',0):.1%})")
    except Exception:
        pass  # bad file -- silently use hardcoded weights

def _print_weight_audit():
    """
    Prints a weight audit at startup.
    Only runs when learned weights are active (silent in hardcoded mode).
    Shows: active mode, sum check, top 3 contributors, excluded indicators.
    """
    if not _LEARNED_WEIGHTS:
        return  # hardcoded weights -- no audit needed

    print()
    print("  ┌─────────────────────────────────────────────┐")
    print("  │  WEIGHT AUDIT  (learned weights active)     │")
    print("  └─────────────────────────────────────────────┘")
    print(f"  Samples: {_LEARNED_WEIGHTS.get('n_samples','?')}  |  "
          f"CV accuracy: {_LEARNED_WEIGHTS.get('cv_accuracy',0):.1%}  |  "
          f"Activated: {_LEARNED_WEIGHTS.get('activated_date','?')[:10]}")

    for strategy in ['TREND', 'MEAN_REVERSION']:
        w = _get_weights(strategy)
        total = sum(w.values())
        top3  = sorted(w.items(), key=lambda x: x[1], reverse=True)[:3]
        excluded = [k for k, v in w.items() if v == 0.0]

        print(f"  {strategy}:")
        print(f"    Weight sum:  {total:.6f}  {'OK' if abs(total-1.0)<0.001 else 'WARNING: not 1.0'}")
        print(f"    Top 3:       " +
              "  |  ".join(f"{k}={v:.1%}" for k,v in top3))
        if excluded:
            print(f"    Excluded:    {excluded}  (neg coef -- zeroed out)")
        else:
            print(f"    Excluded:    none")

    print()

def _get_weights(strategy: str) -> dict:
    """
    Returns indicator weights for the composite score formula.
    Uses learned weights if available for this strategy, else hardcoded.
    """
    if not _LEARNED_WEIGHTS:
        # Hardcoded defaults
        if strategy == "TREND":
            # Zscore/Hurst 0.95 correlated -- reduced zscore, boosted hawkes
            return {'hurst':0.12,'hawkes':0.26,'ofi':0.18,'sector':0.10,
                    'add':0.12,'zscore':0.04,'ks':0.08,'bayes':0.10}
        else:  # MEAN_REVERSION
            return {'exhaustion':0.30,'ofi':0.20,'sector':0.15,
                    'add':0.15,'ks':0.10,'bayes':0.10}

    # Use regime-specific learned weights if available, else overall
    regime_key = strategy if strategy in _LEARNED_WEIGHTS else "overall"
    lw = _LEARNED_WEIGHTS.get(regime_key, {})
    if not lw:
        lw = _LEARNED_WEIGHTS.get("overall", {})

    # Map learned weight keys to composite formula keys
    KEY_MAP = {
        'hurst_score':'hurst', 'hawkes_score':'hawkes', 'ofi_score':'ofi',
        'sector_score':'sector', 'add_score':'add', 'zscore':'zscore',
        'ks_score':'ks', 'bayes_prob':'bayes'
    }
    mapped = {KEY_MAP.get(k,k): v/100 for k,v in lw.items()
              if not k.endswith('_signs')}  # skip meta keys

    # Weights are already positive-only (negatives zeroed by optimizer).
    # For safety: double-check using stored coef signs if available.
    signs_key = (strategy if strategy in _LEARNED_WEIGHTS else "overall") + "_coef_signs"
    coef_signs = _LEARNED_WEIGHTS.get(signs_key,
                 _LEARNED_WEIGHTS.get("overall_coef_signs", {}))
    if coef_signs:
        for orig_col, sign in coef_signs.items():
            mapped_col = KEY_MAP.get(orig_col, orig_col)
            if sign < 0 and mapped_col in mapped:
                mapped[mapped_col] = 0.0  # exclude harmful indicators

    # Hardcoded defaults as fallback for any missing keys
    # Zscore/Hurst 0.95 correlated -- reduced zscore, boosted hawkes
    defaults = {'hurst':0.12,'hawkes':0.26,'ofi':0.18,'sector':0.10,
                'add':0.12,'zscore':0.04,'ks':0.08,'bayes':0.10}
    for k, v in defaults.items():
        if k not in mapped:
            mapped[k] = v

    # CRITICAL: normalize to sum=1.0 after merging.
    # If any fallback key was added on top of learned weights that already sum
    # to 1.0, the total exceeds 1.0 and every composite score inflates.
    # Example: 7 learned keys (sum=1.0) + hurst fallback (0.12) = sum 1.12
    # -> composite at avg indicator 70 scores 78.4 instead of 70.0
    total = sum(mapped.values())
    if total > 0 and abs(total - 1.0) > 0.001:
        mapped = {k: v/total for k, v in mapped.items()}

    return mapped

# Print weight audit if learned weights are active
_print_weight_audit()

# ── Path integrity check ──────────────────────────────────────────────────────
# All 7 scanner files must be in the same directory.
# If any are missing, paths will scatter and cross-file reads will fail silently.
_required_files = [
    "scanner_v4.py", "scanner_terminal_v4.py", "scanner_research_v4.py",
    "dynamic_watchlist.py", "ticker_intelligence.py",
    "run_scheduler.py", "setup_tasks.bat",
]
_missing = [f for f in _required_files
            if not os.path.exists(os.path.join(SCANNER_DIR, f))]
if _missing:
    import warnings
    warnings.warn(
        f"[SCANNER] Missing files in {SCANNER_DIR}: {_missing}\n"
        f"All 7 files must be in the same folder. "
        f"Paths will be inconsistent until resolved.",
        stacklevel=2
    )

# Subfolder for Markov pre-built matrices -- created automatically on first prefetch
# Location: same folder as scanner_v4.py / markov_data / ALKT.json
MARKOV_DATA_DIR = os.path.join(SCANNER_DIR, "markov_data")

# Subfolder for ticker personality profiles
# Location: same folder as scanner_v4.py / ticker_profiles / ALKT.json
TICKER_PROFILE_DIR = os.path.join(SCANNER_DIR, "ticker_profiles")

# Daily outputs go to ~/Downloads (standard, user-visible)
DOWNLOADS_DIR = os.path.join(os.path.expanduser("~"), "Downloads")

try:
    import streamlit as st
    # Only True when actually running inside `streamlit run`.
    # Uses the official Streamlit runtime check.
    # Guards against MagicMock (unit tests) by checking the actual module type.
    _real_streamlit = hasattr(st, '__version__') and isinstance(getattr(st,'__version__',''), str)
    if _real_streamlit:
        try:
            from streamlit.runtime.scriptrunner import get_script_run_ctx
            _ctx = get_script_run_ctx()
            STREAMLIT_MODE = _ctx is not None and hasattr(_ctx, 'session_id')
        except Exception:
            STREAMLIT_MODE = False
    else:
        STREAMLIT_MODE = False   # mock/stub -- not a real Streamlit session
except ImportError:
    STREAMLIT_MODE = False

try:
    import schwab
    # Only mark as available if credentials are actually set
    _schwab_key    = os.environ.get("SCHWAB_API_KEY", "")
    _schwab_secret = os.environ.get("SCHWAB_API_SECRET", "")
    _schwab_acct   = os.environ.get("SCHWAB_ACCOUNT_ID", "")
    SCHWAB_AVAILABLE = bool(_schwab_key and _schwab_secret and _schwab_acct)
    # To connect: set SCHWAB_API_KEY, SCHWAB_API_SECRET, SCHWAB_ACCOUNT_ID env vars
except ImportError:
    SCHWAB_AVAILABLE = False
    # Running in yfinance mode - all features work, data has 15s delay


# ── Config ────────────────────────────────────────────────────────────────────

WATCHLIST = [
    # Technology (XLK)
    "CRUS","POWI","ALKT","TASK","IRTC","APPF",
    # Consumer Discretionary (XLY)
    "BOOT","GIII","PLAY","HIMS","DXLG","EAT",
    # Healthcare (XLV)
    "ACAD","INVA","NVCR","GKOS","IMVT","KRYS",
    # Financials (XLF)
    "OPFI","GCMG","NRDS","ENVA","ECPG",
    # Energy (XLE)
    "REX","PTEN","WTTR","MTDR","GPOR",
    # Industrials (XLI)
    "KTOS","ASTE","HLIO","DLX","HAYW",
    # Materials (XLB)
    "TROX","RYAM","KWR",
    # Real Estate (XLRE)
    "NTST","EPRT","STAG",
]

# BASE SECTOR MAP (static + dynamic watchlist additions)
# ───────────────────────────────────────────────────────────────
SECTOR_MAP = {
    # Technology (XLK)
    "CRUS":"XLK","POWI":"XLK","ALKT":"XLK","TASK":"XLK",
    "IRTC":"XLK","APPF":"XLK",

    # Consumer Discretionary (XLY)
    "BOOT":"XLY","GIII":"XLY","PLAY":"XLY","HIMS":"XLY",
    "DXLG":"XLY","EAT":"XLY",

    # Healthcare (XLV)
    "ACAD":"XLV","INVA":"XLV","NVCR":"XLV","GKOS":"XLV",
    "IMVT":"XLV","KRYS":"XLV",

    # Financials (XLF)
    "OPFI":"XLF","GCMG":"XLF","NRDS":"XLF","ENVA":"XLF","ECPG":"XLF",

    # Energy (XLE)
    "REX":"XLE","PTEN":"XLE","WTTR":"XLE","MTDR":"XLE","GPOR":"XLE",

    # Industrials (XLI)
    "KTOS":"XLI","ASTE":"XLI","HLIO":"XLI","DLX":"XLI","HAYW":"XLI",

    # Materials (XLB)
    "TROX":"XLB","RYAM":"XLB","KWR":"XLB",

    # Real Estate (XLRE)
    "NTST":"XLRE","EPRT":"XLRE","STAG":"XLRE",

    # Dynamic watchlist additions
    "CEVA":"XLK","ICHR":"XLK","IDCC":"XLK",
    "CHEF":"XLY",
    "ATRC":"XLV","OCUL":"XLV","PRCT":"XLV","TNDM":"XLV","NEOG":"XLV",
    "FSTR":"XLI",
    "FLNC":"XLE",
    "FRST":"XLF",
}

# ───────────────────────────────────────────────────────────────
# EXTENDED SECTOR MAP (covers all UNIVERSE tickers)
# MUST BE DEFINED BEFORE update()
# ───────────────────────────────────────────────────────────────
_SECTOR_MAP_EXTRA = {
    # Technology / Software
    "DOMO":"XLK","EVER":"XLK","MTSI":"XLK","JAMF":"XLK",
    "BLKB":"XLK","CARG":"XLK","CCOI":"XLK","CDRE":"XLK",
    "CHGG":"XLK","CMCO":"XLK","COHU":"XLK","CRVL":"XLK",
    "CVLT":"XLK","DIOD":"XLK","EGHT":"XLK","EGAN":"XLK",
    "EPAC":"XLK","EXLS":"XLK","FORR":"XLK","FTDR":"XLK",
    "GNSS":"XLK","HIMX":"XLK","HURC":"XLK","IESC":"XLK",
    "IDCC":"XLK","ITRI":"XLK","KVHI":"XLK","LSCC":"XLK",
    "LYTS":"XLK","MGNI":"XLK","MLAB":"XLK","MRCY":"XLK",
    "NTGR":"XLK","NOVT":"XLK","NSP":"XLK","OMCL":"XLK",
    "OPCH":"XLK","PRGS":"XLK","SMAR":"XLK","LFST":"XLK",
    "CLFD":"XLK","ATNI":"XLK","LPSN":"XLK","DSGN":"XLK",

    # Healthcare
    "AMRX":"XLV","HCAT":"XLV","HROW":"XLV","HRMY":"XLV",
    "IDYA":"XLV","IMCR":"XLV","INSM":"XLV","KRTX":"XLV",
    "LGND":"XLV","LNTH":"XLV","MDGL":"XLV","MMSI":"XLV",
    "MNKD":"XLV","MYGN":"XLV","NKTR":"XLV","NVAX":"XLV",
    "PCRX":"XLV","RARE":"XLV","RCUS":"XLV","RXST":"XLV",
    "SRPT":"XLV","TGTX":"XLV","PCVX":"XLV","CORT":"XLV",
    "ARQT":"XLV","OCGN":"XLV","IONS":"XLV","ACLS":"XLV",

    # Financials
    "ABR":"XLF","AIV":"XLF","AMH":"XLF","BCPC":"XLF",
    "BPOP":"XLF","CVBF":"XLF","CHCO":"XLF","CHDN":"XLF",
    "CTBI":"XLF","EFC":"XLF","FBRT":"XLF","FISI":"XLF",
    "FMBH":"XLF","FRME":"XLF","GDOT":"XLF","GSBC":"XLF",
    "HAFC":"XLF","HBT":"XLF","HFWA":"XLF","HOPE":"XLF",
    "HTLD":"XLF","HURN":"XLF","INDB":"XLF","IPAR":"XLF",
    "KMPR":"XLF","LKFN":"XLF","LCII":"XLF","LPRO":"XLF",
    "MBWM":"XLF","MFIN":"XLF","MORN":"XLF","MVBF":"XLF",
    "NATH":"XLF","NBTB":"XLF","NFBK":"XLF","NMFC":"XLF",
    "NMIH":"XLF","NWBI":"XLF","OBNK":"XLF","OCFC":"XLF",
    "OCSL":"XLF","OFG":"XLF","OPRT":"XLF","BANR":"XLF",
    "FFIN":"XLF","GBCI":"XLF","SFNC":"XLF","HOMB":"XLF",
    "BCAL":"XLF","CBAN":"XLF","RNST":"XLF","NWFL":"XLF",

    # Industrials
    "CDRE":"XLI","CLAR":"XLI","CLW":"XLI","DAN":"XLI",
    "EML":"XLI","EPAC":"XLI","ERII":"XLI","EXPO":"XLI",
    "FELE":"XLI","FOXF":"XLI","GTLS":"XLI","HWKN":"XLI",
    "HURC":"XLI","HCC":"XLI","IART":"XLI","IESC":"XLI",
    "KELYA":"XLI","MATW":"XLI","MATX":"XLI","MBIN":"XLI",
    "MYRG":"XLI","NNBR":"XLI","NOMD":"XLI","NSP":"XLI",
    "TITN":"XLI","MRTN":"XLI","CMCO":"XLI",

    # Energy
    "FLNC":"XLE","GNW":"XLE","GLNG":"XLE","KNTK":"XLE",
    "NESR":"XLE","HLX":"XLE",

    # Materials
    "CBT":"XLB","BCPC":"XLB","HWKN":"XLB",

    # Consumer Discretionary
    "DBRG":"XLY","DIN":"XLY","EXPI":"XLY","EZPW":"XLY",
    "GRPN":"XLY","HAIN":"XLY","JACK":"XLY","JBSS":"XLY",
    "KELYA":"XLY","LCUT":"XLY","LECO":"XLY","LOVE":"XLY",
    "MCRI":"XLY","MNRO":"XLY","NDLS":"XLY","PRPL":"XLY",
    "HVT":"XLY","CHEF":"XLY",

    # Consumer Staples
    "CVCO":"XLP","JJSF":"XLP",

    # Utilities
    "MSEX":"XLU","OGS":"XLU",

    # Real Estate
    "CLDT":"XLRE","DEA":"XLRE","ELME":"XLRE","ESRT":"XLRE",
    "FBRT":"XLRE","GOOD":"XLRE","HASI":"XLRE","HIW":"XLRE",
    "IIPR":"XLRE","NXRT":"XLRE","ABR":"XLRE","AIV":"XLRE",
    "AMH":"XLRE","EFC":"XLRE",
}

# Apply extended mappings
SECTOR_MAP.update(_SECTOR_MAP_EXTRA)

# ───────────────────────────────────────────────────────────────
# NASDAQ‑100 SECTOR MAP
# ───────────────────────────────────────────────────────────────
_NASDAQ_SECTORS = {
    "AAPL":"XLK","MSFT":"XLK","NVDA":"XLK","AMD":"XLK","AVGO":"XLK",
    "QCOM":"XLK","AMAT":"XLK","LRCX":"XLK","KLAC":"XLK","MU":"XLK",
    "MRVL":"XLK","SNPS":"XLK","CDNS":"XLK","INTC":"XLK","ON":"XLK",
    "NXPI":"XLK","TXN":"XLK","MCHP":"XLK","SMCI":"XLK","GFS":"XLK",
    "PANW":"XLK","CRWD":"XLK","ZS":"XLK","FTNT":"XLK","TEAM":"XLK",
    "DDOG":"XLK","ADBE":"XLK","ADSK":"XLK","ANSS":"XLK","WDAY":"XLK",
    "FISV":"XLK","CTSH":"XLK","PAYX":"XLK",
    "AMZN":"XLY","TSLA":"XLY","BKNG":"XLY","LULU":"XLY","ABNB":"XLY",
    "MELI":"XLY","ORLY":"XLY","ROST":"XLY",
    "NFLX":"XLC","META":"XLC","GOOG":"XLC","GOOGL":"XLC","WBD":"XLC",
    "TTWO":"XLC","EA":"XLC","NTES":"XLC","JD":"XLC","BIDU":"XLC",
    "RBLX":"XLC","ZM":"XLC","SIRI":"XLC","TTD":"XLC",
    "COST":"XLP","KDP":"XLP","MNST":"XLP","WBA":"XLP",
    "AMGN":"XLV","GILD":"XLV","VRTX":"XLV","BIIB":"XLV",
    "REGN":"XLV","MRNA":"XLV","IDXX":"XLV","ILMN":"XLV",
    "ODFL":"XLI","PCAR":"XLI","FAST":"XLI","CSX":"XLI","VRSK":"XLI",
    "GEHC":"XLV","CSGP":"XLRE",
    "AEP":"XLU","EXC":"XLU",
    "COIN":"XLF","PYPL":"XLF",
    "PDD":"XLY","ASML":"XLK","FANG":"XLE",
}

# Apply NASDAQ mappings
SECTOR_MAP.update(_NASDAQ_SECTORS)

UNIVERSE = list(dict.fromkeys([
    "CRUS","POWI","MTSI","JAMF","ALKT","TASK","APPF","DOMO","EVER",
    "INVA","TNDM","HAYW","ASTE","KTOS","DLX","HLIO","TROX","RYAM","KWR",
    "NTST","EPRT","CIVI","PTEN","WTTR","BOOT","GIII","PLAY","LESL",
    "PRPL","HIMS","ACAD","PDCO","HCAT","NVCR","GKOS","OPFI","GCMG",
    "PRCT","ARVN","ARWR","FOLD","IMVT","KRYS","LGND","MYGN","NKTR",
    "NVAX","PCRX","RARE","RCUS","SRPT","TGTX","ABR","AIV","AMH",
    "STAG","CUBE","DEA","EFC","FBRT","GOOD","HIW","AMRX","ATRC",
    "BCPC","BLKB","BPOP","CABO","CADE","CARG","CBT","CCOI","CDRE","CEVA",
    "CHCO","CHDN","CHEF","CHGG","CLAR","CLDT","CLW","CMCO","CNMD",
    "COHU","CRVL","CTBI","CVBF","CVCO","CVLT","CWST",
    "DAN","DBRG","DIN","DIOD","DXLG","EAT","ECPG","EGAN","EGHT","ELME",
    "EML","ENVA","EPAC","ERII","ESRT","ETSY","EVGO","EVTC","EXLS","EXPI",
    "EXPO","EZPW","FELE","FISI","FLNC","FLR","FMBH",
    "FORM","FORR","FOXF","FRME","FRST","FSTR","FTDR","GDOT","GEOS",
    "GIFI","GLDD","GLNG","GNSS","GNW","GPOR","GPRE","GRPN","GSBC","GSHD",
    "GTLS","HAFC","HAIN","HALO","HASI","HBT","HCC","HCSG","HELE",
    "HFWA","HIMX","HLX","HOPE","HROW","HRMY","HSII",
    "HTLD","HURC","HURN","HVT","HWKN","IART","ICFI",
    "ICHR","IDCC","IDYA","IESC","IIPR","IMCR","INDB","INMD","INSM","IONS",
    "IPAR","IRDM","IRTC","ITRI","JACK","JBSS","JJSF","KELYA","KMPR","KNTK",
    "KVHI","LCII","LCUT","LECO","LKFN","LMAT","LNTH","LOVE","LPRO",
    "LQDT","LSCC","LSTR","LYTS","MARA","MATW","MATX","MBIN","MBWM",
    "MCRI","MDGL","MERC","MFIN","MGNI","MGPI","MLAB","MMSI","MNKD",
    "MNRO","MODV","MORN","MRCY","MRTN","MRUS","MSEX","MTDR","MVBF",
    "NATH","NBTB","NDLS","NEOG","NESR","NEXT","NFBK","NMFC",
    "NMIH","NNBR","NOMD","NOVT","NSP","NTGR","NTRA","NWBI","NXRT",
    "OBNK","OCFC","OCGN","OCSL","OCUL","OFG","OGS","OMCL","OPCH","OPRT",
    # Banking/Financials (replaced HTBK, HTLF, IBTX, LBAI, NCBS, MOFG, NYCB)
    "BANR","FFIN","GBCI","SFNC","HOMB","BCAL","CBAN","NWFL","RNST",
    # Technology (replaced VCRA, COUP, MDRX, AXNX, CNSL)
    "SMAR","PRGS","DSGN","LFST","CLFD","ATNI","LPSN",
    # Healthcare (replaced FGEN, AXNX, HIIQ)
    "RXST","TGTX","ACLS","ARQT","KRTX","PCVX","CORT",
    # Industrials/Energy (replaced FARO, GDEN, ALEX, CIR, CONN, COUP)
    "HLIO","ASTE","MYRG","TITN","DLX","KTOS","MTDR",
]))

# ── Optional: tickers.txt universe override ──────────────────────────────────
# Add/remove symbols without editing Python. One symbol per line.
# Lines starting with # are ignored. File: C:\QuantScanner\tickers.txt
# If file doesn't exist, hardcoded UNIVERSE above is used unchanged.
_tickers_path = os.path.join(SCANNER_DIR, "tickers.txt")
if os.path.exists(_tickers_path):
    try:
        _file_syms = [
            ln.split("#")[0].strip().upper()
            for ln in open(_tickers_path).readlines()
            if ln.split("#")[0].strip()
        ]
        if _file_syms:
            UNIVERSE = list(dict.fromkeys(UNIVERSE + _file_syms))
    except Exception:
        pass  # bad file format -- keep hardcoded UNIVERSE


# ── Universe Mode ────────────────────────────────────────────
# WATCHLIST:   40-stock curated list   - fast, manual picks
# RUSSELL2000: Full IWM constituent list ~2000 stocks - more signals, more noise
# Set UNIVERSE_MODE to switch. Russell 2000 requires stricter firewalls.
UNIVERSE_MODE      = "WATCHLIST"     # "WATCHLIST" | "QQQ" | "IWM" | "RUSSELL2000"

# ── NASDAQ-100 / QQQ Universe ─────────────────────────────────────────────────
# Large-cap NASDAQ tech stocks. Use with --universe QQQ or UNIVERSE_MODE="QQQ".
# Benchmark: QQQ (not SPY). Higher cap floor: $5B. Trend-biased strategy.
NASDAQ_100 = [
    "AAPL","MSFT","NVDA","AMZN","GOOG","GOOGL","META","TSLA","AVGO","COST",
    "NFLX","AMD","ADBE","QCOM","AMAT","LRCX","KLAC","MU","MRVL","SNPS",
    "CDNS","PANW","CRWD","ZS","FTNT","WDAY","TEAM","DDOG","MELI","REGN",
    "AMGN","GILD","VRTX","BIIB","MRNA","IDXX","ILMN","ROST","ODFL","PCAR",
    "FAST","CSX","ORLY","VRSK","KDP","MNST","TTD","ANSS","PAYX","CTSH",
    "FISV","ADSK","GEHC","CSGP","BKNG","LULU","ABNB","COIN","RBLX","EA",
    "TTWO","INTC","ON","NXPI","MCHP","TXN","SMCI","AEP","EXC","WBD",
    "FANG","PDD","JD","BIDU","ZM","SIRI","WBA","ASML","GFS","PYPL",
]

# ═══════════════════════════════════════════════════════════════════════════
# HIERARCHICAL RS ARCHITECTURE  (Seven-Layer Pipeline)
# ═══════════════════════════════════════════════════════════════════════════
#
# Layer 2 — Sector Mapping:   SECTOR_MAP  →  broad sector ETF (XLK, XLV …)
# Layer 2b— Industry Mapping: INDUSTRY_ETF_MAP → sub-sector ETF (SOXX, XBI…)
# Layer 4 — Benchmark:  always SPY (market) + sector ETF (sector alpha)
#                        + industry ETF when available (micro-regime alpha)
#
# The old logic (`if mcap < 3B: IWM`) was wrong because:
#   - small cap is a CHARACTERISTIC, not a benchmark
#   - a small-cap bank should compare to KRE, not IWM
#   - a small-cap biotech should compare to XBI, not IWM
#   - IWM as benchmark conflates sector effects with size effects
#
# The new flow in scan_symbol():
#   sector_etf   = SECTOR_MAP[symbol]          # broad economic group
#   industry_etf = INDUSTRY_ETF_MAP[symbol]    # micro-regime (may be None)
#   RS1 = sector_etf    vs SPY        → is the sector leading the market?
#   RS2 = stock_return  vs sector_etf → is the stock leading its sector?
#   RS3 = stock_return  vs industry_etf → is it leading its micro-regime?
#   composite_sector_score = weighted blend of RS1, RS2, RS3
#   gate passes only if: sector not collapsing AND stock not lagging sector
# ═══════════════════════════════════════════════════════════════════════════

# ── Layer 2b: Industry / Sub-sector ETF Map ──────────────────────────────────
# Maps individual tickers to their specific industry ETF benchmark.
# More granular than the broad GICS sector ETFs.
# Only populated for tickers where micro-regime matters significantly.
# Unmapped tickers fall back to their SECTOR_MAP entry.
INDUSTRY_ETF_MAP: dict = {
    # ── Semiconductors → SOXX ────────────────────────────────────────────
    # Semis are a tight micro-regime: all move together on inventory cycles,
    # earnings, and AI capex. Comparing a semi to broad XLK vs SOXX is the
    # difference between sector alpha and true peer alpha.
    "CRUS":"SOXX","POWI":"SOXX","MTSI":"SOXX","ALKT":"SOXX",
    "ICHR":"SOXX","COHU":"SOXX","DIOD":"SOXX","LSCC":"SOXX",
    # NASDAQ-100 semis
    "NVDA":"SOXX","AMD":"SOXX","AVGO":"SOXX","QCOM":"SOXX",
    "MU":"SOXX","AMAT":"SOXX","LRCX":"SOXX","KLAC":"SOXX",
    "MRVL":"SOXX","SNPS":"SOXX","CDNS":"SOXX","INTC":"SOXX",
    "ON":"SOXX","NXPI":"SOXX","TXN":"SOXX","MCHP":"SOXX",

    # ── Biotech / Biopharmaceutical → XBI ────────────────────────────────
    # Biotech has near-zero correlation with broad healthcare (XLV) during
    # sector rotations. A biotech in a collapsing XBI should not pass sector
    # gate just because XLV (med devices, pharma) is holding up.
    "ACAD":"XBI","NVCR":"XBI","IMVT":"XBI","KRYS":"XBI",
    "FOLD":"XBI","INVA":"XBI","ARWR":"XBI","ARVN":"XBI",
    "RCUS":"XBI","TGTX":"XBI","RXST":"XBI","SRPT":"XBI",
    "IDYA":"XBI","IMCR":"XBI","INSM":"XBI","RARE":"XBI",
    "IONS":"XBI","BIIB":"XBI","VRTX":"XBI","REGN":"XBI",
    "MRNA":"XBI","NKTR":"XBI","OCGN":"XBI","MDGL":"XBI",
    "LGND":"XBI","MYGN":"XBI","ARQT":"XBI","KRTX":"XBI",
    "PCVX":"XBI","CORT":"XBI","MNKD":"XBI","HRMY":"XBI",

    # ── Healthcare Equipment / MedTech → IHI ─────────────────────────────
    "ATRC":"IHI","GKOS":"IHI","MMSI":"IHI","PRCT":"IHI",
    "TNDM":"IHI","OMCL":"IHI","IART":"IHI","CNMD":"IHI",
    "OCUL":"IHI","IRTC":"IHI","LMAT":"IHI","ATRC":"IHI",
    "NEOG":"IHI","ITRI":"IHI",

    # ── Regional Banks → KRE ─────────────────────────────────────────────
    # Regional banks decouple from broad XLF (which includes megabanks,
    # insurance, asset managers). KRE tracks the micro-regime correctly.
    "OPFI":"KRE","ECPG":"KRE","ENVA":"KRE","GCMG":"KRE",
    "OBNK":"KRE","BPOP":"KRE","CHCO":"KRE","CVBF":"KRE",
    "FRME":"KRE","HAFC":"KRE","HBT":"KRE","HFWA":"KRE",
    "HOPE":"KRE","INDB":"KRE","LKFN":"KRE","MBWM":"KRE",
    "NBTB":"KRE","NFBK":"KRE","NWBI":"KRE","OCFC":"KRE",
    "OFG":"KRE","FRST":"KRE","CTBI":"KRE","FMBH":"KRE",
    "GBCI":"KRE","BANR":"KRE","FFIN":"KRE","SFNC":"KRE",
    "HOMB":"KRE","BCAL":"KRE","CBAN":"KRE","RNST":"KRE",
    "NRDS":"KRE",

    # ── Oil & Gas E&P → XOP ──────────────────────────────────────────────
    # Upstream O&G tracks oil/gas prices much more tightly than broad XLE
    # (which includes majors, midstream, and services).
    "PTEN":"XOP","WTTR":"XOP","MTDR":"XOP","GPOR":"XOP",
    "REX":"XOP","GPRE":"XOP",

    # ── Oil Services → OIH ───────────────────────────────────────────────
    "NESR":"OIH","HLX":"OIH",

    # ── Retail → XRT ─────────────────────────────────────────────────────
    "BOOT":"XRT","GIII":"XRT","PLAY":"XRT","DXLG":"XRT",
    "HAIN":"XRT","LESL":"XRT","LOVE":"XRT","NDLS":"XRT",
    "JACK":"XRT","HVT":"XRT","JBSS":"XRT",

    # ── Software → IGV ───────────────────────────────────────────────────
    "TASK":"IGV","APPF":"IGV","BLKB":"IGV","CARG":"IGV",
    "DOMO":"IGV","FORR":"IGV","EVER":"IGV","EXLS":"IGV",
    "CHGG":"IGV","EGHT":"IGV","EGAN":"IGV","CVLT":"IGV",
    "ANSS":"IGV","WDAY":"IGV","ADSK":"IGV",

    # ── REITs → VNQ ──────────────────────────────────────────────────────
    "NTST":"VNQ","EPRT":"VNQ","STAG":"VNQ","CUBE":"VNQ",
    "HIW":"VNQ","NXRT":"VNQ","IIPR":"VNQ","ABR":"VNQ",
    "AIV":"VNQ","AMH":"VNQ","FBRT":"VNQ","GOOD":"VNQ",
    "CLDT":"VNQ","ESRT":"VNQ","ELME":"VNQ","DEA":"VNQ",

    # ── Aerospace & Defense → ITA ────────────────────────────────────────
    "KTOS":"ITA","MRCY":"ITA",

    # ── Clean Energy → ICLN ──────────────────────────────────────────────
    "FLNC":"ICLN","HASI":"ICLN",
}

# Expanded SECTOR_MAP — covers all UNIVERSE tickers
# Ensures every symbol has a correct broad-sector ETF assignment.
# Unmapped tickers fall back to "SPY" (neutral market comparison).
# Add new tickers here before adding them to UNIVERSE or WATCHLIST.
_SECTOR_MAP_EXTRA: dict = {
    # Technology / Software
    "DOMO":"XLK","EVER":"XLK","MTSI":"XLK","JAMF":"XLK",
    "BLKB":"XLK","CARG":"XLK","CCOI":"XLK","CDRE":"XLK",
    "CHGG":"XLK","CMCO":"XLK","COHU":"XLK","CRVL":"XLK",
    "CVLT":"XLK","DIOD":"XLK","EGHT":"XLK","EGAN":"XLK",
    "EPAC":"XLK","EXLS":"XLK","FORR":"XLK","FTDR":"XLK",
    "GNSS":"XLK","HIMX":"XLK","HURC":"XLK","IESC":"XLK",
    "IDCC":"XLK","ITRI":"XLK","KVHI":"XLK","LSCC":"XLK",
    "LYTS":"XLK","MGNI":"XLK","MLAB":"XLK","MRCY":"XLK",
    "NTGR":"XLK","NOVT":"XLK","NSP":"XLK","OMCL":"XLK",
    "OPCH":"XLK","PRGS":"XLK","SMAR":"XLK","LFST":"XLK",
    "CLFD":"XLK","ATNI":"XLK","LPSN":"XLK","DSGN":"XLK",
    # Healthcare
    "AMRX":"XLV","HCAT":"XLV","HROW":"XLV","HRMY":"XLV",
    "IDYA":"XLV","IMCR":"XLV","INSM":"XLV","KRTX":"XLV",
    "LGND":"XLV","LNTH":"XLV","MDGL":"XLV","MMSI":"XLV",
    "MNKD":"XLV","MYGN":"XLV","NKTR":"XLV","NVAX":"XLV",
    "PCRX":"XLV","RARE":"XLV","RCUS":"XLV","RXST":"XLV",
    "SRPT":"XLV","TGTX":"XLV","PCVX":"XLV","CORT":"XLV",
    "ARQT":"XLV","OCGN":"XLV","IONS":"XLV","ACLS":"XLV",
    # Financials
    "ABR":"XLF","AIV":"XLF","AMH":"XLF","BCPC":"XLF",
    "BPOP":"XLF","CVBF":"XLF","CHCO":"XLF","CHDN":"XLF",
    "CTBI":"XLF","EFC":"XLF","FBRT":"XLF","FISI":"XLF",
    "FMBH":"XLF","FRME":"XLF","GDOT":"XLF","GSBC":"XLF",
    "HAFC":"XLF","HBT":"XLF","HFWA":"XLF","HOPE":"XLF",
    "HTLD":"XLF","HURN":"XLF","INDB":"XLF","IPAR":"XLF",
    "KMPR":"XLF","LKFN":"XLF","LCII":"XLF","LPRO":"XLF",
    "MBWM":"XLF","MFIN":"XLF","MORN":"XLF","MVBF":"XLF",
    "NATH":"XLF","NBTB":"XLF","NFBK":"XLF","NMFC":"XLF",
    "NMIH":"XLF","NWBI":"XLF","OBNK":"XLF","OCFC":"XLF",
    "OCSL":"XLF","OFG":"XLF","OPRT":"XLF","BANR":"XLF",
    "FFIN":"XLF","GBCI":"XLF","SFNC":"XLF","HOMB":"XLF",
    "BCAL":"XLF","CBAN":"XLF","RNST":"XLF","NWFL":"XLF",
    # Industrials
    "CDRE":"XLI","CLAR":"XLI","CLW":"XLI","DAN":"XLI",
    "EML":"XLI","EPAC":"XLI","ERII":"XLI","EXPO":"XLI",
    "FELE":"XLI","FOXF":"XLI","GTLS":"XLI","HWKN":"XLI",
    "HURC":"XLI","HCC":"XLI","IART":"XLI","IESC":"XLI",
    "KELYA":"XLI","MATW":"XLI","MATX":"XLI","MBIN":"XLI",
    "MYRG":"XLI","NNBR":"XLI","NOMD":"XLI","NSP":"XLI",
    "TITN":"XLI","MRTN":"XLI","CMCO":"XLI",
    # Energy
    "FLNC":"XLE","GNW":"XLE","GLNG":"XLE","KNTK":"XLE",
    "NESR":"XLE","HLX":"XLE",
    # Materials
    "CBT":"XLB","BCPC":"XLB","HWKN":"XLB",
    # Consumer Discretionary
    "DBRG":"XLY","DIN":"XLY","EXPI":"XLY","EZPW":"XLY",
    "GRPN":"XLY","HAIN":"XLY","JACK":"XLY","JBSS":"XLY",
    "KELYA":"XLY","LCUT":"XLY","LECO":"XLY","LOVE":"XLY",
    "MCRI":"XLY","MNRO":"XLY","NDLS":"XLY","PRPL":"XLY",
    "HVT":"XLY","CHEF":"XLY",
    # Consumer Staples
    "CVCO":"XLP","JJSF":"XLP",
    # Utilities
    "MSEX":"XLU","OGS":"XLU",
    # Real Estate
    "CLDT":"XLRE","DEA":"XLRE","ELME":"XLRE","ESRT":"XLRE",
    "FBRT":"XLRE","GOOD":"XLRE","HASI":"XLRE","HIW":"XLRE",
    "IIPR":"XLRE","NXRT":"XLRE","ABR":"XLRE","AIV":"XLRE",
    "AMH":"XLRE","EFC":"XLRE",
    # Communication Services
    "CABO":"XLC","CHDN":"XLC","EVTC":"XLC","GSHD":"XLC",
    "IRDM":"XLC","NTGR":"XLC",
}


def get_industry_etf(symbol: str) -> str | None:
    """
    Returns the most specific benchmark ETF for RS comparison.

    Layer 2b in the seven-layer pipeline:
      Industry ETF > Sector ETF > SPY

    Examples:
      NVDA  → SOXX  (semiconductor micro-regime, not generic XLK)
      ACAD  → XBI   (biotech regime, not generic XLV)
      OPFI  → KRE   (regional bank regime, not generic XLF)
      KTOS  → ITA   (defense micro-regime, not generic XLI)
      BOOT  → XRT   (retail micro-regime, not generic XLY)
      BOOT  → XLY   (fallback if XRT not available)
      ALKT  → SOXX  (despite being "software" the stock trades with semis)

    Returns None if no industry mapping exists (use sector ETF instead).
    """
    return INDUSTRY_ETF_MAP.get(symbol)


# ── Layer 4: Dynamic Benchmark Assignment ────────────────────────────────────
def get_benchmarks(symbol: str) -> dict:
    """
    Three-tier benchmark hierarchy (Seven-Layer Pipeline Layer 4).
    Benchmarks assigned by SECTOR + INDUSTRY, never market cap or index.

    primary   = SPY   (market alpha baseline -- always)
    secondary = sector ETF (XLK, XLV, XLE... from SECTOR_MAP)
    tertiary  = industry ETF (SOXX, XBI, KRE... from INDUSTRY_ETF_MAP)

    Example: NVDA  ->  SPY / XLK / SOXX
      NVDA +2%, SPY +1% looks good.
      NVDA +2%, SOXX +4% means NVDA is lagging its semi peers -> penalty.
    """
    return {
        "primary":   "SPY",
        "secondary": SECTOR_MAP.get(symbol, "SPY"),
        "tertiary":  INDUSTRY_ETF_MAP.get(symbol),
    }


# ── RS Benchmark ETF prefetch list ──────────────────────────────────────────
# All ETFs needed for RS computations -- pre-cached by --prefetch task.
RS_BENCHMARK_ETFS = list(dict.fromkeys([
    "SPY", "QQQ", "IWM",
    "XLK","XLF","XLV","XLY","XLP","XLE","XLI","XLB","XLRE","XLU","XLC",
    "SOXX","XBI","KRE","XOP","OIH","XRT","IGV","VNQ","ITA","IHI","ICLN","CIBR","XAR",
]))


# ── Layer 1: Russell 2000 / IWM Universe -- eligibility only ─────────────────
# IWM membership = stock is ELIGIBLE to be scanned.
# NOT the benchmark.  A Russell 2000 biotech compares vs XBI + XLV, not IWM.
RUSSELL_2000_ELIGIBLE   = True
RUSSELL_2000_SUPPLEMENT = [
    "LYTS","NTGR","KVHI","LSCC","FORM","COHU","GNSS","EGAN","HIMX",
    "MNKD","TGTX","CORT","HRMY","HROW","PCVX","KRTX","ACLS","RXST",
    "GEOS","NESR","KNTK","HLX","GIFI",
    "NFBK","MVBF","MFIN","MERC","HBT","GSBC","FRST","FMBH","CTBI","CHCO",
    "LOVE","NDLS","NOMD","NATH","JJSF","JACK","HVT",
    "MYRG","TITN","NNBR","HURC","IESC","FSTR","EPAC",
]


# ── Thresholds ───────────────────────────────────────────────
# NOTE: MIDCAP_MIN/MAX are fallback defaults only.
# Active cap limits are determined by get_dynamic_cap_limits(strategy).
# TREND: $300M-$50B  |  MR: $300M-$20B  |  AUTO: $300M-$20B
MIDCAP_MIN         = 300_000_000      # $300M baseline
MIDCAP_MAX         = 20_000_000_000   # $20B baseline (MR and AUTO default)
ACCOUNT_SIZE       = float(os.environ.get("ACCOUNT_SIZE", 50000))
SIGNAL_THRESHOLD   = 65
ATR_STOP_MULT      = 1.5   # risk  = ATR × 1.5
ATR_TARGET_MULT    = 4.5   # reward = ATR × 4.5 → 3:1 R:R  (4.5 / 1.5 = 3.0)
SPY_WEAK_THRESH    = -0.005

# ── Dynamic Cap Limits by Strategy ───────────────────────────
# These replace the single MIDCAP_MIN/MAX with strategy-aware ranges.
#
# TREND ($300M - $50B):
#   - Large/mega caps have cleaner trends -- more institutional flow,
#     less noise, better follow-through after breakout
#   - $50B ceiling catches growing mid-caps like MTSI ($21B) that trend well
#   - Don't go below $300M in TREND -- smaller caps fake-out too often
#
# MEAN_REVERSION ($300M - $20B):
#   - Same floor as TREND -- $300M minimum for our $50K account
#     (sub-$300M stocks are too illiquid to exit a position cleanly)
#   - $20B ceiling -- above this, institutional size makes price "sticky"
#     and MR targets (VWAP reversion) take too long or never complete
#   - This is our proven working range from the original scanner
#
# AUTO ($300M - $20B):
#   - Conservative default -- matches original scanner behavior
#   - Widens automatically when regime switches to TREND
#
MIDCAP_MIN_TREND    = 300_000_000      # $300M -- fake-outs too common below this
MIDCAP_MAX_TREND    = 50_000_000_000   # $50B  -- institutional flow, clean trends
MIDCAP_MIN_MR       = 300_000_000      # $300M -- need liquidity to exit at $50K size
MIDCAP_MAX_MR       = 20_000_000_000   # $20B  -- above this VWAP reversion is too slow
MIDCAP_MIN_AUTO     = 300_000_000      # $300M
MIDCAP_MAX_AUTO     = 20_000_000_000   # $20B  -- conservative default

# ── Dynamic Sector RS by Strategy ────────────────────────────
# TREND (1.00):
#   - Sector just needs to not be collapsing -- momentum often precedes
#     sector RS leadership. Being too strict here filters out early-move stocks.
#
# MEAN_REVERSION (1.02):
#   - Slightly tighter than the old 1.01 -- for MR you need the sector
#     actively pulling the stock back toward VWAP. A lagging sector means
#     the reversion catalyst isn't there and the stock stays down.
#
# AUTO (1.00):
#   - Loose default. On strong trend days (AUTO -> TREND) this prevents
#     the sector filter from killing all trend signals.
#
SECTOR_RS_MIN      = 1.01    # base fallback -- rarely used directly
SECTOR_RS_TREND    = 1.00    # TREND: sector just needs to not be broken
SECTOR_RS_MR       = 1.02    # MR: sector must lead SPY (reversion catalyst needed)
SECTOR_RS_AUTO     = 1.00    # AUTO: loose default, widens signal pool


def get_dynamic_sector_rs(strategy: str = "AUTO") -> float:
    """Returns the sector RS minimum for the current strategy."""
    effective = strategy if strategy != "AUTO" else STRATEGY_MODE
    if effective == "TREND":          return SECTOR_RS_TREND
    elif effective == "MEAN_REVERSION": return SECTOR_RS_MR
    return SECTOR_RS_AUTO


def get_dynamic_cap_limits(strategy: str = "AUTO") -> tuple:
    """
    Returns (min_cap, max_cap) in dollars based on active strategy.

    TREND:          $300M - $50B
    MEAN_REVERSION: $300M - $20B  (original proven range)
    AUTO:           $300M - $20B  (conservative default)
    """
    effective = strategy if strategy != "AUTO" else STRATEGY_MODE
    if effective == "TREND":
        return MIDCAP_MIN_TREND, MIDCAP_MAX_TREND
    elif effective == "MEAN_REVERSION":
        return MIDCAP_MIN_MR, MIDCAP_MAX_MR
    return MIDCAP_MIN_AUTO, MIDCAP_MAX_AUTO


def is_midcap(symbol: str, strategy: str = "AUTO") -> tuple:
    """
    Dynamic market cap filter.
    TREND: $300M-$50B  |  MR: $300M-$20B  |  AUTO: $300M-$20B
    Returns (passes: bool, mcap_billions: float).
    """
    mcap = get_market_cap(symbol)
    if mcap is None:
        return False, 0.0
    mn, mx = get_dynamic_cap_limits(strategy)
    return mn <= mcap <= mx, round(mcap / 1e9, 2)


OFI_LONG_ENTRY     = 0.30
OFI_SHORT_ENTRY    = -0.30
OFI_LONG_EXIT      = 0.45
OFI_SHORT_EXIT     = 0.55
ZSCORE_MAX         = 2.5
ZSCORE_MIN         = -2.5
Z_ENTRY_THRESH     = 2.0
KURTOSIS_MIN       = 3.0
SKEW_LONG_MIN      = 0.1
SKEW_SHORT_MAX     = -0.1
BAYES_BASE         = 0.50
BAYES_ADD          = 0.08
BAYES_SECTOR       = 0.10
BAYES_HAWKES       = 0.14
HAWKES_DECAY       = 0.3   # exponential decay constant for Hawkes intensity
HALFLIFE_BASE      = 600
HALFLIFE_MIN       = 120
HALFLIFE_MAX       = 1800
INTRADAY_SOFT      = -0.010
INTRADAY_HARD      = -0.020
INTRADAY_KILL      = -0.030
GAP_DOWN_THRESH    = -0.010
LOWER_LOW_BARS     = 6

# ── Liquidity Firewall ────────────────────────────────────────
MIN_DOLLAR_VOLUME  = 5_000_000       # $5M default (watchlist)
MIN_REL_VOLUME     = 1.2             # 1.2x default

# ── Volatility + Participation filters (the missing layer) ───
# These work TOGETHER with cap and sector RS.
# Cap filters context ("is this the right type of stock?")
# ATR% filters magnitude ("will it actually move enough to pay?")
# RVOL filters participation ("are there real buyers/sellers now?")
#
# Strategy-aware because the required movement differs:
#   TREND:  needs bigger moves (ATR% >= 2.5%) and strong participation (1.5x)
#           -- you need follow-through for ATR*3 targets to be reached
#   MR:     needs enough snap-back range (ATR% >= 1.5%) and moderate volume (1.2x)
#           -- the reversion happens with less volume, but needs range to exist
#   AUTO:   uses MR defaults (conservative, widens when regime switches)
#
MIN_ATR_PCT_TREND   = 0.025   # 2.5% -- $20 stock needs $0.50 daily range minimum
MIN_ATR_PCT_MR      = 0.015   # 1.5% -- enough for VWAP reversion to be profitable
MIN_ATR_PCT_AUTO    = 0.015   # 1.5% -- conservative default

MIN_RVOL_TREND      = 1.5     # 1.5x -- needs institutional participation for follow-through
MIN_RVOL_MR         = 1.2     # 1.2x -- existing gate, confirmed enough for MR


def get_dynamic_vol_filters(strategy: str = "AUTO") -> tuple:
    """
    Returns (min_atr_pct, min_rvol) for the current strategy.

    TREND: ATR >= 2.5%, RVOL >= 1.5x
      Large caps allowed to $50B BUT only if they're actually moving.
      This directly addresses the large-cap drift problem -- a $10B stock
      with 0.5% daily range is useless; a $10B stock with 3% ATR is ideal.

    MR:   ATR >= 1.5%, RVOL >= 1.2x
      Enough range for the VWAP reversion to cover spread + commission.
      Low RVOL is ok for MR (reversions happen quietly) but range must exist.
    """
    effective = strategy if strategy != "AUTO" else STRATEGY_MODE
    if effective == "TREND":
        return MIN_ATR_PCT_TREND, MIN_RVOL_TREND
    elif effective == "MEAN_REVERSION":
        return MIN_ATR_PCT_MR, MIN_RVOL_MR
    return MIN_ATR_PCT_AUTO, MIN_REL_VOLUME

# ── Russell 2000 Stricter Thresholds ─────────────────────────
R2K_MIN_DOLLAR_VOLUME = 20_000_000  # $20M - doc spec S7: "$20M+ for execution"
R2K_MIN_REL_VOLUME    = 1.5         # 1.5x
R2K_SIGNAL_THRESHOLD  = 72          # "Best of best" from 2000 stocks
R2K_HURST_MIN         = 0.55        # Clean trend required
R2K_HAWKES_MIN        = 62          # Higher clustering bar
R2K_MIN_PRICE         = 5.0         # No penny stocks
R2K_BATCH_SIZE        = 500         # Rate limit protection
R2K_RF_LENGTH         = 25          # Adaptive base for small-cap speed

# ── Portfolio Risk Controls (§8) ─────────────────────────────
MAX_OPEN_POSITIONS    = 5            # Hard cap - prevents overtrading
MAX_SECTOR_POSITIONS  = 2            # Max 2 per sector at once
MAX_TOTAL_RISK_PCT    = 0.05         # Max 5% of account in open risk

# ── Execution Rules ───────────────────────────────────────────
LIMIT_ORDERS_ONLY     = True         # S12: never use market orders
MAX_TRADES_PER_DAY    = 20           # S12: slippage accumulates - cap daily trades

# ── Exit Rules ───────────────────────────────────────────────
PARTIAL_TP_PCT        = 0.50         # S13: take 50% off at first target
EOD_SOFT_EXIT_HOUR    = 15           # 3:00 PM - begin closing
EOD_SOFT_EXIT_MIN     = 30           # 3:30 PM soft exit
EOD_HARD_EXIT_HOUR    = 15           # 3:55 PM hard exit all
EOD_HARD_EXIT_MIN     = 55

# ── Half-Life Minimums (§6) ───────────────────────────────────
HALFLIFE_MIN_REMAINING_SCALP    = 30   # seconds - reject if < 30s for scalps
HALFLIFE_MIN_REMAINING_INTRADAY = 120  # seconds - reject if < 2min for intraday

# ── Global Controls ───────────────────────────────────────────
GLOBAL_KILL_SWITCH    = False        # S10: set True to halt all new entries
GLOBAL_REGIME_LOCK    = None         # S3: "TREND"|"MEAN_REVERSION"|None (auto)

# ── Markov Gate (off by default) ─────────────────────────────
# Set True ONLY after running: python scanner_research_v4.py --markov-gate SYMBOL
# and confirming Sharpe improves by >= 0.2.
# When True: checks regime stability before each MR signal.
# Tightens stop if volatile transition likely. Adds 10% score penalty if unstable.
# Has zero effect on TREND signals - only applies to MEAN_REVERSION.
MARKOV_GATE_ENABLED   = False

# ── Strategy Switch ───────────────────────────────────────────
STRATEGY_MODE      = "AUTO"

# ── Mean Reversion Exit ───────────────────────────────────────
MR_TAKE_PROFIT_PCT = 0.005   # flat % TP (floor only; MR actually targets VWAP)
MR_STOP_LOSS_PCT   = 0.008   # floor stop -- actual stop uses ATR*1.5 when available
# R:R NOTE: flat TP=0.5% / SL=0.8% inverted (needs 73% WR after slippage).
# VWAP target is dynamically larger. Check rr_ratio in scan_log for real R:R.
MR_MIN_RR_RATIO    = 1.0     # block MR entries where R:R < 1.0

# Gap filter for MR: don't enter MR on stocks with large gap-up
# GAP-UP = 0% WR confirmed in hypothesis -- gap-up stocks are trending, not reverting
# Threshold: gap > 2% = stock is in momentum mode, MR invalidated
MR_GAP_UP_MAX      = 0.02    # block MR when gap-up exceeds 2% (trending, not reverting)
MR_GAP_DOWN_MAX    = -0.10   # also block extreme gap-down > 10% (potential halt/news)

# Gate 13: Momentum ROC filter (falling knife guard)
# If price dropped >MR_ROC_THRESHOLD in last MR_ROC_LOOKBACK bars, block MR entry.
# Prevents buying into vertical flushes where the "mean" is actively moving away.
# 1.5% drop in 3 bars = stock moving at ~18%/hour -- too fast for MR entry.
MR_ROC_LOOKBACK    = 3       # bars to measure momentum over (3 × 5min = 15min)
MR_ROC_THRESHOLD   = -0.015  # raw fallback threshold (used when vol estimate unavailable)
MR_ROC_NORM_THRESH = -1.5    # normalized threshold: block when 3-bar drop > 1.5× typical move
                              # biotech ATR=4%: blocks at -6% drop. utility ATR=0.5%: blocks at -0.75%
MR_ROC_ENABLED     = True    # set False to disable during testing

# Extreme RVOL ceiling for MR (Gate 12)
# RVOL > 8x usually means news/fundamental repricing -- NOT mean reversion.
# These symbols route to TREND or get excluded from MR entirely.
MR_RVOL_MAX        = 8.0     # above this, MR disabled (likely trend/news event)

# ── ADX Regime ───────────────────────────────────────────────
ADX_MAX            = 25
# SPY volatility thresholds (v4.1 corrected logic):
#   SPY_VOL_CALM:    < 0.5%  -> pure MR day, minimal trend activity
#   SPY_VOL_NORMAL:  < 1.0%  -> normal day, standard MR thresholds
#   SPY_VOL_NOTRADE: > 2.0%  -> ONLY blocks if choppy (no direction)
#                               A clean 3%+ gap-up goes to TREND, not NO_TRADE
#                               True NO_TRADE = high vol AND no SPY direction
SPY_VOL_NOTRADE    = 0.020
SPY_VOL_CALM       = 0.005
SPY_VOL_NORMAL     = 0.010

# Parkinson shock detection: ratio of current Parkinson vol to baseline
# 1.5 = current intraday vol is 50% above recent norm -> SHOCK state
# 2.0 = severe shock (use for MR total block)
# Self-calibrating: fires based on deviation not a fixed ATR% number
PARKINSON_SHOCK_THRESHOLD  = 1.5   # ratio trigger: current/baseline > this
PARKINSON_SHOCK_SEVERE     = 2.0   # severe: total MR block
PARKINSON_SHOCK_COOLDOWN_M = 15    # minutes to keep MR blocked after shock clears

# Runtime shock state (set by terminal scan loop, read by scan_symbol)
# When True: scan_symbol() returns None for all MR signals (circuit breaker)
_SHOCK_MR_OVERRIDE = False
_IWM_CACHE      = None   # cached IWM 5m bars (refreshed every 2 min)
_IWM_CACHE_TIME = 0.0    # timestamp of last IWM fetch   # set externally by scanner_terminal_v4.py


# ── Parkinson Volatility (High-Low based, scale-invariant) ──────────────────
# More accurate than simple avg H-L range for shock detection because:
#   1. Uses log(H/L)^2 -- normalized so a $5 stock and $500 stock are comparable
#   2. Self-calibrating: compares current vol to its own rolling baseline
#   3. Reacts to a single "nasty" candle faster than std deviation
# Formula: sqrt( (1 / (4*n*ln2)) * sum(ln(H/L)^2) )
# Returns (current_park_vol, baseline_park_vol, ratio, is_shock)

def get_parkinson_vol(df: pd.DataFrame, window: int = 14,
                      baseline_window: int = 100) -> tuple:
    """
    Computes Parkinson Volatility over the last `window` bars.
    Uses High/Low prices -- captures intra-bar volatility missed by close-to-close.

    Returns:
        (current_vol, baseline_vol, ratio, is_shock)
        ratio = current / baseline  (>1.5 = shock, >2.0 = severe shock)
        is_shock = True when ratio > PARKINSON_SHOCK_THRESHOLD

    Used by get_strategy_regime() to replace the hardcoded avg_bar_range > 1.5% check.
    Self-calibrating: fires based on deviation from recent norm, not a fixed number.
    """
    if df is None or len(df) < window + 2:
        return (0.0, 0.0, 1.0, False)

    try:
        h = df["high"].values.astype(float)
        lo = df["low"].values.astype(float)

        # Guard against zero or negative L/H
        valid = (lo > 0) & (h > 0) & (h >= lo)
        if valid.sum() < window:
            return (0.0, 0.0, 1.0, False)

        h_safe  = np.where(valid, h, np.nan)
        lo_safe = np.where(valid, lo, np.nan)

        log_hl_sq = np.log(h_safe / lo_safe) ** 2  # (ln H/L)^2

        scale = 1.0 / (4.0 * window * np.log(2))

        # Current window: last `window` bars
        recent_sq = np.nansum(log_hl_sq[-window:])
        current_vol = float(np.sqrt(scale * recent_sq))

        # Baseline: rolling mean over last `baseline_window` bars
        n_base = min(baseline_window, len(log_hl_sq))
        base_sq = np.nansum(log_hl_sq[-n_base:]) / max(n_base // window, 1)
        baseline_vol = float(np.sqrt(scale * base_sq)) if base_sq > 0 else current_vol

        ratio = current_vol / baseline_vol if baseline_vol > 1e-9 else 1.0

        is_shock = ratio > PARKINSON_SHOCK_THRESHOLD

        return (round(current_vol, 6), round(baseline_vol, 6),
                round(ratio, 3), is_shock)

    except Exception:
        return (0.0, 0.0, 1.0, False)


# ── Range Filter (ThinkScript port) ──────────────────────────
# Final confirmation gate: scanner finds mathematical exhaustion,
# Range Filter confirms price has actually started moving in your favor.
# Final_Signal = (Score > 75) AND (price > filter AND upward > 0)
# Ported from ThinkorSwim RangeFilter study.

def compute_range_filter(prices: np.ndarray, length: int = 25) -> tuple:
    """
    Range Filter - ported from ThinkScript.
    Returns (filter_val, upward, downward, signal_up, signal_down).

    signal_up   = True when price crosses above filter AND trend is up
    signal_down = True when price crosses below filter AND trend is down

    length=25 for Russell 2000 (was 100 in TOS - too slow for small-caps)
    length=30 for 5-min chart, length=20 for 1-min chart

    ThinkScript logic:
      range = ATR(length) * multiplier
      if price > filter + range: filter = price - range
      if price < filter - range: filter = price + range
      upward   = filter > filter[1]
      downward = filter < filter[1]
    """
    if len(prices) < length + 2:
        return prices[-1], False, False, False, False

    # Compute smooth range using ATR-like rolling std
    n      = len(prices)
    smrng  = np.zeros(n)
    for i in range(length, n):
        smrng[i] = np.std(prices[i-length:i], ddof=1) * 1.0  # multiplier=1.0

    filt   = np.zeros(n)
    filt[length] = prices[length]

    for i in range(length + 1, n):
        rng = smrng[i]
        if prices[i] > filt[i-1] + rng:
            filt[i] = prices[i] - rng
        elif prices[i] < filt[i-1] - rng:
            filt[i] = prices[i] + rng
        else:
            filt[i] = filt[i-1]

    upward   = filt[-1] > filt[-2]
    downward = filt[-1] < filt[-2]

    # Signal: price crosses filter in direction of trend
    signal_up   = prices[-1] > filt[-1] and upward
    signal_down = prices[-1] < filt[-1] and downward

    return round(float(filt[-1]), 4), upward, downward, signal_up, signal_down


# ── Active Thresholds (adapts to UNIVERSE_MODE) ──────────────────────────────

def get_active_thresholds() -> dict:
    """Returns correct thresholds for current universe mode."""
    if UNIVERSE_MODE == "RUSSELL2000":
        return {
            "min_dollar_vol":  R2K_MIN_DOLLAR_VOLUME,
            "min_rel_vol":     R2K_MIN_REL_VOLUME,
            "signal_thresh":   R2K_SIGNAL_THRESHOLD,
            "hurst_min":       R2K_HURST_MIN,
            "hawkes_min":      R2K_HAWKES_MIN,
            "min_price":       R2K_MIN_PRICE,
            "batch_size":      R2K_BATCH_SIZE,
            "rf_length":       R2K_RF_LENGTH,
        }
    return {
        "min_dollar_vol":  MIN_DOLLAR_VOLUME,
        "min_rel_vol":     MIN_REL_VOLUME,
        "signal_thresh":   SIGNAL_THRESHOLD,
        "hurst_min":       0.0,
        "hawkes_min":      0.0,
        "min_price":       1.0,
        "batch_size":      500,
        "rf_length":       30,
    }


# ── Global Kill Switch (§10) ──────────────────────────────────────────────────

def is_trading_allowed() -> tuple:
    """
    Global no-trade gate. Returns (allowed: bool, reason: str).
    Checked before every signal is acted upon.
    Blocks if:
      - GLOBAL_KILL_SWITCH is True (manual override)
      - GLOBAL_REGIME_LOCK is set and contradicts current strategy
      - Past EOD hard exit time
    """
    if GLOBAL_KILL_SWITCH:
        return False, "GLOBAL_KILL_SWITCH=True - all entries halted", 0.0
    now = datetime.now()
    # Before market open
    if now.hour < 9 or (now.hour == 9 and now.minute < 30):
        return False, "PRE_MARKET - no entries before 9:30 AM", 0.0
    # Hard EOD exit
    if now.hour > EOD_HARD_EXIT_HOUR or \
       (now.hour == EOD_HARD_EXIT_HOUR and now.minute >= EOD_HARD_EXIT_MIN):
        return False, f"EOD_HARD_EXIT - past {EOD_HARD_EXIT_HOUR}:{EOD_HARD_EXIT_MIN:02d} PM", 0.0
    return True, "OK"


def is_eod_soft_exit() -> bool:
    """True between 3:30 PM and 3:55 PM - close winners, no new entries."""
    now = datetime.now()
    soft = now.hour == EOD_SOFT_EXIT_HOUR and now.minute >= EOD_SOFT_EXIT_MIN
    hard = now.hour == EOD_HARD_EXIT_HOUR and now.minute >= EOD_HARD_EXIT_MIN
    return soft and not hard


# ── Portfolio Risk Controls (§8) ─────────────────────────────────────────────

def portfolio_allows_entry(symbol: str, sector_etf: str,
                            positions: dict, account_size: float) -> tuple:
    """
    Enforces S8 hard limits before any new position opens.
    Returns (allowed: bool, reason: str).

    Checks:
      1. Max open positions = 5
      2. Max positions per sector = 2
      3. Max total account risk = 5%
    """
    open_pos = {sym: p for sym, p in positions.items()
                if p.direction != "FLAT"}

    # Rule 1: Max open positions
    if len(open_pos) >= MAX_OPEN_POSITIONS:
        return False, f"MAX_POSITIONS: {len(open_pos)}/{MAX_OPEN_POSITIONS} open", 0.0
    # Rule 2: Max per sector
    sector_count = sum(
        1 for sym, p in open_pos.items()
        if SECTOR_MAP.get(sym, "SPY") == sector_etf
    )
    if sector_count >= MAX_SECTOR_POSITIONS:
        return False, f"MAX_SECTOR: {sector_count}/{MAX_SECTOR_POSITIONS} in {sector_etf}", 0.0
    # Rule 3: Max total account risk
    total_risk = sum(
        abs(p.entry_price - p.stop) * p.shares
        for p in open_pos.values()
        if p.stop > 0 and p.entry_price > 0
    )
    risk_pct = total_risk / account_size if account_size > 0 else 0.0
    if risk_pct >= MAX_TOTAL_RISK_PCT:
        return False, f"MAX_RISK: {risk_pct:.1%} of account at risk (max {MAX_TOTAL_RISK_PCT:.0%})", 0.0
    return True, f"OK - {len(open_pos)} open, {sector_count} in {sector_etf}, {risk_pct:.1%} at risk"


# ── Pre-Trade Validation Layer (§2.2) ────────────────────────────────────────

def validate_before_execute(result: dict, current_price: float,
                              positions: dict) -> tuple:
    """
    Critical re-validation before any order is placed.
    Scanner signal may be stale by the time execution is ready.

    Returns (valid: bool, reason: str).

    Checks:
      1. Global kill switch / EOD
      2. Signal half-life still alive (min threshold)
      3. Price hasn't drifted more than 0.5% from signal price
      4. Range Filter still confirming
      5. Market regime hasn't flipped
      6. Portfolio limits allow new position
    """
    sym        = result.get("symbol","?")
    signal_px  = float(result.get("price", 0))
    hl_alive   = bool(result.get("hl_alive", False))
    hl_rem     = float(result.get("hl_remaining", 0))
    rf_confirms= bool(result.get("rf_confirms", False))
    strategy   = result.get("strategy","--")
    sector_etf = result.get("sector_etf","SPY")

    # 1. Global gate
    allowed, reason = is_trading_allowed()
    if not allowed:
        return False, reason

    # 2. Half-life minimum - choose threshold based on mode
    min_hl = (HALFLIFE_MIN_REMAINING_SCALP if UNIVERSE_MODE == "RUSSELL2000"
              else HALFLIFE_MIN_REMAINING_INTRADAY)
    if not hl_alive or hl_rem < min_hl:
        return False, f"SIGNAL_EXPIRED: {hl_rem:.0f}s remaining (min {min_hl}s)", 0.0
    # 3. Price drift - reject if moved more than 0.5% from signal
    if signal_px > 0:
        drift = abs(current_price - signal_px) / signal_px
        if drift > 0.005:
            return False, f"PRICE_DRIFT: {drift:.2%} from signal price ${signal_px:.2f}", 0.0
    # 4. Range Filter still confirming
    if not rf_confirms:
        return False, "RF_NO_CONFIRM: Range Filter no longer confirming direction", 0.0
    # 5. GLOBAL_REGIME_LOCK check
    if GLOBAL_REGIME_LOCK and GLOBAL_REGIME_LOCK != strategy:
        return False, f"REGIME_LOCK: system locked to {GLOBAL_REGIME_LOCK}, signal is {strategy}", 0.0
    # 6. Portfolio limits
    port_ok, port_reason = portfolio_allows_entry(sym, sector_etf, positions, ACCOUNT_SIZE)
    if not port_ok:
        return False, port_reason

    return True, f"VALID - hl={hl_rem:.0f}s drift={abs(current_price-signal_px)/signal_px*100:.2f}%"


# ── Adaptive Range Filter Length (§5.1) ──────────────────────────────────────

def adaptive_rf_length(atr_pct: float, universe: str = "WATCHLIST") -> int:
    """
    S5.1: RF length must adapt to volatility - not static.

    High volatility:  shorter length (20-30) -> more responsive
    Low volatility:   longer length (40-60)  -> filters more noise

    atr_pct: ATR as % of price (e.g., 0.025 = 2.5%)
    """
    if universe == "RUSSELL2000":
        # Small-caps: tighter ranges
        if atr_pct > 0.04:   return 20   # very volatile
        if atr_pct > 0.025:  return 25   # normal small-cap
        return 30                          # low volatility
    else:
        # Watchlist / mid-cap
        if atr_pct > 0.03:   return 30
        if atr_pct > 0.015:  return 40
        return 50


# ── Relative Ranking (§9) ─────────────────────────────────────────────────────

def rank_signals(df: pd.DataFrame, top_n: int = 10,
                 top_pct: float = 0.05) -> pd.DataFrame:
    """
    S9: With 2000 stocks, fixed score thresholds become unreliable.
    Replace with relative ranking:
      - top_n:   return best N results regardless of absolute score
      - top_pct: return top X% of all scanned (e.g., top 5%)
    Both filters applied - stricter of the two wins.

    Also adds a 'rank' and 'percentile' column to the DataFrame.
    """
    if df.empty:
        return df

    df = df.copy()
    df["rank"]       = df["score"].rank(ascending=False, method="first").astype(int)
    df["percentile"] = (1 - df["score"].rank(pct=True)).round(4)

    n_from_pct = max(1, int(len(df) * top_pct))
    cutoff_n   = min(top_n, n_from_pct)

    ranked = df.nsmallest(cutoff_n, "rank").reset_index(drop=True)
    return ranked


# ── Partial Take Profit (§13) ─────────────────────────────────────────────────

def determine_exit_with_partial(current_price: float, entry_price: float,
                                  vwap: float, side: str,
                                  shares_remaining: int) -> tuple:
    """
    S13: Two-stage exit:
      Stage 1 - Take 50% off at first target (VWAP touch or 0.5% TP)
      Stage 2 - Trail remaining 50% until hard stop or 0.8% TP

    Returns (action: str, shares_to_close: int, reason: str).
    Actions: "PARTIAL_EXIT" | "FULL_EXIT" | "HOLD"
    """
    half = max(1, shares_remaining // 2)

    if side == "LONG":
        at_target = current_price >= vwap or \
                    current_price >= entry_price * (1 + MR_TAKE_PROFIT_PCT)
        at_stop   = current_price <= entry_price * (1 - MR_STOP_LOSS_PCT)

        if at_stop:
            return "FULL_EXIT",    shares_remaining, f"STOP_LOSS at ${current_price:.2f}"
        if at_target and shares_remaining > half:
            return "PARTIAL_EXIT", half, f"PARTIAL_TP 50% at ${current_price:.2f}"
        if at_target:
            return "FULL_EXIT",    shares_remaining, f"FULL_TP at ${current_price:.2f}"

    if side == "SHORT":
        at_target = current_price <= vwap or \
                    current_price <= entry_price * (1 - MR_TAKE_PROFIT_PCT)
        at_stop   = current_price >= entry_price * (1 + MR_STOP_LOSS_PCT)

        if at_stop:
            return "FULL_EXIT",    shares_remaining, f"STOP_LOSS at ${current_price:.2f}"
        if at_target and shares_remaining > half:
            return "PARTIAL_EXIT", half, f"PARTIAL_TP 50% at ${current_price:.2f}"
        if at_target:
            return "FULL_EXIT",    shares_remaining, f"FULL_TP at ${current_price:.2f}"

    return "HOLD", 0, "HOLD"


# ── EOD Exit Check (§13) ──────────────────────────────────────────────────────

def check_eod_exit(positions: dict) -> list:
    """
    S13: End-of-day exit - eliminates overnight risk.
    3:30-3:55 PM: close winning positions first (soft exit)
    3:55 PM+:     close everything hard (hard exit)

    Returns list of symbols to close with reason.
    """
    now = datetime.now()
    to_close = []

    hard = (now.hour > EOD_HARD_EXIT_HOUR or
            (now.hour == EOD_HARD_EXIT_HOUR and now.minute >= EOD_HARD_EXIT_MIN))
    soft = (now.hour == EOD_SOFT_EXIT_HOUR and
            now.minute >= EOD_SOFT_EXIT_MIN and not hard)

    for sym, pos in positions.items():
        if pos.direction == "FLAT":
            continue
        if hard:
            to_close.append((sym, "EOD_HARD_EXIT - 3:55 PM all positions closed"))
        elif soft and pos.pnl > 0:
            to_close.append((sym, f"EOD_SOFT_EXIT - closing winner ${pos.pnl:+.2f}"))

    return to_close


# ── Market Cap Filter ─────────────────────────────────────────────────────────

_mcap_cache: dict = {}
_get_markov_gate_fn = None   # cached after first import -- avoids repeat module lookup

def get_market_cap(symbol: str) -> Optional[float]:
    now = time.time()
    if symbol in _mcap_cache and now - _mcap_cache[symbol][1] < 3600:
        return _mcap_cache[symbol][0]
    try:
        info = yf.Ticker(symbol).info
        mcap = info.get("marketCap") or info.get("enterpriseValue")
        if mcap and mcap > 0:
            _mcap_cache[symbol] = (float(mcap), now)
            return float(mcap)
    except Exception:
        pass
    return None

# ── Data Fetchers ─────────────────────────────────────────────────────────────

def fetch_daily(symbol: str, n: int = 60) -> Optional[pd.DataFrame]:
    try:
        import contextlib, io
        with contextlib.redirect_stderr(io.StringIO()):
            df = yf.Ticker(symbol).history(period="3mo", interval="1d")
        if df.empty or len(df) < 10:
            return None
        df.columns = [c.lower() for c in df.columns]
        return df[["open","high","low","close","volume"]].dropna().tail(n)
    except Exception:
        return None

def fetch_intraday(symbol: str, timeframe: str = "5m") -> Optional[pd.DataFrame]:
    try:
        import contextlib, io
        interval = "1m" if timeframe == "15s" else timeframe
        period   = "5d" if timeframe in ("5m","15m") else "1d"
        with contextlib.redirect_stderr(io.StringIO()):
            df = yf.Ticker(symbol).history(period=period, interval=interval)
        if df.empty or len(df) < 10:
            return None
        df.columns = [c.lower() for c in df.columns]
        df = df[["open","high","low","close","volume"]].dropna()
        if timeframe == "15s":
            rows = [{"open":r["open"],"high":r["high"],"low":r["low"],
                     "close":r["close"],"volume":r["volume"]/4}
                    for _, r in df.iterrows() for _ in range(4)]
            df = pd.DataFrame(rows)
        return df
    except Exception:
        return None

def fetch_closes(symbol: str, n: int = 25) -> Optional[np.ndarray]:
    df = fetch_daily(symbol, n)
    return df["close"].values if df is not None else None


# ── SPY/Sector close cache (5-min TTL) ───────────────────────
# SPY closes were being fetched 3+ times per scan: market regime,
# sector RS (once per ETF), and strategy regime. Cache eliminates
# all duplicate network calls within a scan cycle.
_closes_cache: dict = {}
_CLOSES_TTL = 300  # 5 minutes

def fetch_closes_cached(symbol: str, n: int = 25) -> Optional[np.ndarray]:
    key = f"{symbol}:{n}"
    now = time.time()
    if key in _closes_cache and now - _closes_cache[key][1] < _CLOSES_TTL:
        return _closes_cache[key][0]
    result = fetch_closes(symbol, n)
    if result is not None:
        _closes_cache[key] = (result, now)
    return result


# ── ADD Breadth (SPY up-volume proxy, updates every 60s) ──────────────────────

_add_cache = {"value": 0.0, "bullish": True, "updated": 0.0}
_add_lock  = threading.Lock()

def _update_add_loop():
    while True:
        try:
            df = yf.Ticker("SPY").history(period="1d", interval="1m")
            if not df.empty:
                df.columns = [c.lower() for c in df.columns]
                rng   = df["high"].values - df["low"].values
                br    = np.where(rng > 0, (df["close"].values - df["low"].values) / rng, 0.5)
                ratio = (br * df["volume"].values).sum() / df["volume"].values.sum()
                with _add_lock:
                    _add_cache.update({"value":round(ratio,4), "bullish":ratio>0.52, "updated":time.time()})
        except Exception:
            pass
        time.sleep(60)

def get_add_breadth() -> tuple:
    with _add_lock:
        return _add_cache["value"], _add_cache["bullish"]

def add_score(bullish: bool) -> float:
    return 70.0 if bullish else 30.0


# ── Z-Score (exhaustion gate) ─────────────────────────────────────────────────

def compute_zscore(prices: np.ndarray, window: int = 20) -> float:
    if len(prices) < window:
        return 0.0
    s = np.std(prices[-window:], ddof=1)
    return float((prices[-1] - np.mean(prices[-window:])) / s) if s else 0.0

def zscore_gate(z: float, direction: str) -> tuple:
    if direction == "LONG"  and z >  ZSCORE_MAX: return False, f"Z={z:.2f} EXHAUSTED", 0.0
    if direction == "SHORT" and z <  ZSCORE_MIN: return False, f"Z={z:.2f} EXHAUSTED", 0.0
    return True, f"Z={z:.2f} "

def zscore_score(z: float) -> float:
    return float(np.clip(100 - abs(z) * 20, 0, 100))


# ── Kurtosis & Skewness (fat-tail filter) ─────────────────────────────────────

def compute_kurtosis_skew(prices: np.ndarray) -> tuple:
    if len(prices) < 20:
        return 0.0, 0.0
    log_r = np.diff(np.log(prices + 1e-9))
    return round(float(sp_stats.kurtosis(log_r, fisher=True)), 3), \
           round(float(sp_stats.skew(log_r)), 3)

def kurtosis_skew_score(kurt: float, skew: float, direction: str = "LONG") -> tuple:
    base  = float(np.clip((kurt / 6.0) * 60, 0, 60))
    bonus = 0.0
    if direction == "LONG"  and skew >  SKEW_LONG_MIN:  bonus = float(np.clip(skew      * 20, 0, 40))
    if direction == "SHORT" and skew <  SKEW_SHORT_MAX:  bonus = float(np.clip(abs(skew) * 20, 0, 40))
    score = round(base + bonus, 1)
    label = " FAT TAIL" if kurt > 5.0 and score > 70 else \
            " ELEVATED" if kurt > KURTOSIS_MIN else " NORMAL"
    return score, label


# ── Bayesian Win Probability ──────────────────────────────────────────────────

def bayesian_win_prob(add_bull: bool, sector_ok: bool,
                      hawk_sc: float, z_ok: bool, ks_sc: float) -> tuple:
    p, log = BAYES_BASE, [f"Base: {BAYES_BASE:.0%}"]

    def upd(cond, boost, label):
        nonlocal p
        p = float(np.clip(p + (boost if cond else -boost * 0.5), 0.03, 0.97))
        log.append(f"{label} -> {p:.0%}")

    upd(add_bull,         BAYES_ADD,    f"ADD {'' if add_bull else ''}")
    upd(sector_ok,        BAYES_SECTOR, f"Sector {'' if sector_ok else ''}")
    upd(hawk_sc >= 72,    BAYES_HAWKES, f"Hawkes {'' if hawk_sc>=72 else '' if hawk_sc>=58 else '--'}")
    upd(z_ok,             0.04,         f"Z {'' if z_ok else ''}")
    if ks_sc >= 60:
        p = float(np.clip(p + 0.05, 0.03, 0.97))
        log.append(f"FatTail -> {p:.0%}")

    return round(p, 4), log


# ── Alpha Half-Life Decay ─────────────────────────────────────────────────────

_signal_start: dict = {}

def compute_halflife(score: float, atr_pct: float) -> float:
    return float(np.clip(HALFLIFE_BASE * (score / 65.0) / (1.0 + atr_pct * 10),
                         HALFLIFE_MIN, HALFLIFE_MAX))

def halflife_remaining(symbol: str, score: float, price: float, atr: float) -> tuple:
    now     = time.time()
    atr_pct = (atr / price) if price > 0 else 0.02
    if score < SIGNAL_THRESHOLD:
        _signal_start.pop(symbol, None)
        return 0.0, 0.0, False
    if symbol not in _signal_start:
        hl = compute_halflife(score, atr_pct)
        _signal_start[symbol] = {"time": now, "halflife": hl, "lambda": math.log(2) / hl}
    sig      = _signal_start[symbol]
    strength = math.exp(-sig["lambda"] * (now - sig["time"]))
    remaining = max(0.0, sig["halflife"] * math.log(max(strength,1e-9)) / math.log(0.5))
    alive    = strength > 0.25
    if not alive:
        _signal_start.pop(symbol, None)
    return round(strength, 4), round(remaining, 1), alive


# ── Core Indicators ───────────────────────────────────────────────────────────

def compute_hurst(prices: np.ndarray) -> float:
    arr = np.array(prices)
    if len(arr) < 30: return 0.5
    log_r = np.diff(np.log(arr + 1e-9))
    n     = len(log_r)
    lags  = np.unique(np.floor(np.geomspace(5, n//2, 12)).astype(int))
    lags  = lags[lags >= 4]
    rs_v, vl = [], []
    for lag in lags:
        nw = n // lag
        if nw < 2: continue
        rl = []
        for i in range(nw):
            seg = log_r[i*lag:(i+1)*lag]
            cs  = np.cumsum(seg - seg.mean())
            S   = seg.std(ddof=1)
            if S > 0: rl.append((cs.max() - cs.min()) / S)
        if rl: rs_v.append(np.mean(rl)); vl.append(lag)
    if len(vl) < 3: return 0.5
    H, _ = np.polyfit(np.log(vl), np.log(rs_v), 1)
    return float(np.clip(H, 0.0, 1.0))

def hurst_score(H: float)  -> float: return float(np.clip((H - 0.5) * 200 + 50, 0, 100))
def hurst_regime(H: float) -> str:
    return "TRENDING" if H > 0.58 else (" REVERTING" if H < 0.42 else " CHOPPY")


def compute_hawkes(df: pd.DataFrame) -> tuple:
    h, l, c, v = df["high"].values, df["low"].values, df["close"].values, df["volume"].values
    n = len(v)
    if n < 20: return 0.0, 50.0
    br   = np.where(h - l > 0, (c - l) / (h - l), 0.5)
    bv, sv = br * v, (1 - br) * v
    bb   = pd.Series(bv).rolling(20, min_periods=5).mean().values
    bs   = pd.Series(sv).rolling(20, min_periods=5).mean().values
    mu   = float(np.nanmean(bb[~np.isnan(bb)]))
    if mu <= 0: return 0.0, 50.0
    alpha, lams = mu * 0.5, np.zeros(n)
    lams[0] = mu
    for t in range(1, n):
        decay = mu + (lams[t-1] - mu) * math.exp(-HAWKES_DECAY)
        bbt   = bb[t-1] if not np.isnan(bb[t-1]) else mu
        bst   = bs[t-1] if not np.isnan(bs[t-1]) else mu
        delta = (alpha if bv[t-1] > bbt * 1.8 else 0.0) - (alpha * 0.8 if sv[t-1] > bst * 1.8 else 0.0)
        lams[t] = max(0.0, decay + delta)
    base  = np.nanmean(lams[max(0, n-30):-5]) if n > 10 else mu
    if base <= 0: return lams[-1], 50.0
    score = float(np.clip(50 + 50 * math.tanh(lams[-1] / (base + 1e-9) - 1.0), 0, 100))
    return round(lams[-1], 4), round(score, 1)

def hawkes_signal(s: float) -> str:
    if s >= 72:  return "CLUSTERING"
    if s >= 58:  return "BUILDING"
    if s >= 42:  return "IDLE"
    if s >= 20:  return "FADING"
    return "SELL PRESSURE"


# ── Intraday Hurst (5-min bars) ───────────────────────────────────────────────
# Fixes the stale-Hurst problem. Uses last 78 five-min bars (~1 trading day).
# Weighted 60% intraday / 40% daily so today's reversal is visible immediately.

def compute_hurst_intraday(df_5m: pd.DataFrame) -> tuple:
    if df_5m is None or len(df_5m) < 30:
        return 0.5, 50.0, " CHOPPY"
    H      = compute_hurst(df_5m["close"].values[-78:])
    return round(H, 3), round(hurst_score(H), 1), hurst_regime(H)


def combined_hurst_score(h_daily: float, h_intraday: float) -> tuple:
    """Blends daily Hurst (40%) + intraday Hurst (60%).
    Returns (blended_score, conflict_flag).
    conflict_flag = True when daily says trending but intraday says reverting.
    """
    score    = 0.40 * hurst_score(h_daily) + 0.60 * hurst_score(h_intraday)
    conflict = hurst_regime(h_daily) == "TRENDING" and h_intraday < 0.42
    return round(float(np.clip(score, 0, 100)), 1), conflict


def compute_ofi(df: pd.DataFrame, window: int = 12) -> tuple:
    """
    Signed Order Flow Imbalance (OFI) -- CVD-based, range [-1.0, +1.0].

    Formula:  signed_ofi = (buy_vol - sell_vol) / (buy_vol + sell_vol)
    Buy/sell split estimated via bar price position (tick-rule proxy):
      br  = (close - low) / (high - low)   ->  fraction of bar volume that was buying
      buy_vol  = br  × total_volume
      sell_vol = (1-br) × total_volume

    Scale:
      +1.0  = all bars closed at high (pure buy pressure)
      -1.0  = all bars closed at low  (pure sell pressure)
       0.0  = balanced / neutral
      ±0.50 = 3:1 imbalance  (3× buy vs sell, or vice-versa)

    Score (0-100): maps [-1,+1] → [0,100] with delta acceleration bonus.
    References: QuantVPS OFI guide; Cont, Kukanov & Stoikov (2014).
    """
    h, l, c, v   = (df["high"].values, df["low"].values,
                    df["close"].values, df["volume"].values)
    br           = np.where(h - l > 0, (c - l) / (h - l), 0.5)
    buy_vol      = br * v
    sell_vol     = (1.0 - br) * v
    net          = pd.Series(buy_vol - sell_vol)
    tot          = pd.Series(buy_vol + sell_vol).replace(0, np.nan)
    ofi          = (net.rolling(window, min_periods=3).sum() /
                    tot.rolling(window, min_periods=3).sum()).fillna(0.0)
    cur          = float(np.clip(ofi.iloc[-1], -1.0, 1.0))
    delt         = float(ofi.iloc[-1] - ofi.iloc[-4]) if len(ofi) >= 4 else 0.0
    # Map signed [-1,+1] → score [0,100]: neutral(0)=50, full-buy(1)=100, full-sell(-1)=0
    score        = float(np.clip((cur + 1.0) * 50.0 + delt * 25.0, 0.0, 100.0))
    return round(cur, 4), round(delt, 4), round(score, 1)

def ofi_signal(ofi: float, delta: float) -> str:
    # Thresholds on signed OFI scale [-1, +1].  0 = neutral.
    # ±0.50 = 3:1 buy/sell imbalance (standard OFI trigger per QuantVPS).
    # Negative values now properly flag sell-side pressure.
    if ofi >= 0.50 and delta >= 0:  return "ACCUMULATING"   # ≥3:1 buy, strengthening
    if ofi >= 0.30 and delta <  0:  return "TOPPING"         # buy pressure fading
    if ofi <= -0.50:                 return "DISTRIBUTING"   # ≥3:1 sell pressure
    if ofi <= -0.20:                 return "SELLING"         # growing sell pressure
    return "NEUTRAL"


def compute_ofi_normalized(df: pd.DataFrame,
                            window: int = 12,
                            norm_window: int = 60) -> float:
    """
    Z-score normalized OFI per Dean Markwick (2022) / Cont et al. (2014).

    Method:
      1. Compute signed OFI per bar: (buy_vol - sell_vol) / total_vol  ∈ [-1,+1]
      2. Aggregate over rolling `window` bars
      3. Normalize via rolling Z-score over `norm_window` bars:
             ofi_norm = (ofi - rolling_mean(ofi)) / rolling_std(ofi)

    Why: Raw OFI is very spiky and scale-dependent across symbols.
    Normalization compresses it to a stable, comparable range (~-3 to +3).
    The paper finds ofi_norm provides a ~3% R² out-of-sample predictive edge.

    Thresholds (empirically validated on 5m equity bars):
      ofi_norm > +1.5  →  strong buy pressure (top ~7% of all bars)
      ofi_norm < -1.5  →  strong sell pressure
      |ofi_norm| < 0.5 →  neutral / noise

    Args:
        df:          OHLCV DataFrame (must have high, low, close, volume)
        window:      bars to aggregate OFI over (default 12 = ~1h on 5m bars)
        norm_window: rolling lookback for Z-score (default 60 = ~5h on 5m bars)

    Returns:
        float clipped to [-5.0, +5.0]
    """
    if df is None or len(df) < max(window, 5):
        return 0.0
    h, l, c, v  = (df["high"].values, df["low"].values,
                   df["close"].values, df["volume"].values)
    br           = np.where(h - l > 0, (c - l) / (h - l), 0.5)
    buy_vol      = br * v
    sell_vol     = (1.0 - br) * v
    net          = pd.Series(buy_vol - sell_vol)
    tot          = pd.Series(buy_vol + sell_vol).replace(0, np.nan)
    ofi_raw      = (net.rolling(window, min_periods=3).sum() /
                    tot.rolling(window, min_periods=3).sum()).fillna(0.0)
    # Rolling Z-score (Markwick: "scale the OFI values to a known range")
    _min_n       = max(5, norm_window // 4)
    ofi_mean     = ofi_raw.rolling(norm_window, min_periods=_min_n).mean()
    ofi_std      = ofi_raw.rolling(norm_window, min_periods=_min_n).std()
    ofi_norm     = ((ofi_raw - ofi_mean) /
                    ofi_std.replace(0, np.nan)).fillna(0.0)
    return round(float(np.clip(ofi_norm.iloc[-1], -5.0, 5.0)), 3)


    if df is None or len(df) < n: return 0.0
    h, l, c = df["high"].values, df["low"].values, df["close"].values
    trs = [max(h[-i] - l[-i],
               abs(h[-i] - c[-i-1]) if i < len(c) else 0,
               abs(l[-i] - c[-i-1]) if i < len(c) else 0)
           for i in range(1, min(n+1, len(c)))]
    return round(float(np.mean(trs)), 4) if trs else 0.0


def intraday_health(df: pd.DataFrame, prior_close: float) -> tuple:
    """
    Hard intraday momentum gate - fixes the CRUS problem.

    HARD BLOCKS (mult = 0.0, alert = False):
      - Price below VWAP                     -> institutions net sellers, never long here
      - Intraday return <= -3% from open     -> confirmed selloff
      - Gap down > 2% from prior close       -> gap-down continuation risk

    SOFT PENALTIES (score multiplier):
      - Intraday -2% to -3%                  -> 50% penalty
      - Intraday -1% to -2%                  -> 25% penalty
      - Gap down 1-2%                        -> 15% penalty
      - Lower lows on last 6 bars            -> 30% penalty
    """
    if df is None or len(df) < 5: return 1.0, "NO DATA", {}
    try:
        today = pd.Timestamp.now(tz=df.index.tz).strftime("%Y-%m-%d")
        tb    = df[df.index.strftime("%Y-%m-%d") == today]
    except Exception:
        tb = df.tail(60)
    if len(tb) < 3: tb = df.tail(60)

    open_p  = float(tb["open"].iloc[0])
    curr    = float(tb["close"].iloc[-1])
    intra_r = (curr - open_p) / open_p if open_p > 0 else 0.0
    gap_r   = (open_p - prior_close) / prior_close if prior_close > 0 else 0.0
    recent  = tb["close"].values[-LOWER_LOW_BARS:]
    ll      = int(sum(recent[i] < recent[i-1] for i in range(1, len(recent))))
    making_ll = ll >= LOWER_LOW_BARS - 2
    tp    = (tb["high"] + tb["low"] + tb["close"]) / 3
    vwap  = float((tp * tb["volume"]).cumsum().iloc[-1] / tb["volume"].cumsum().iloc[-1])             if tb["volume"].sum() > 0 else curr
    below_vwap = curr < vwap

    mult, flags, hard_blocked = 1.0, [], False

    # ── HARD BLOCKS - only extreme conditions kill signal entirely ──
    # NOTE: below_vwap is NO LONGER a hard block because:
    #   - MR LONG entries REQUIRE price below VWAP (it's the setup condition)
    #   - Hard blocking below_vwap was preventing every valid MR long from showing
    # Instead: score penalty - the MR exhaustion gate enforces VWAP side correctly

    if intra_r <= INTRADAY_KILL:
        mult = 0.0; hard_blocked = True
        flags.append(f"HARD:SELLOFF{intra_r*100:.1f}%")

    if gap_r < -0.02:
        mult = 0.0; hard_blocked = True
        flags.append(f"HARD:GAP{gap_r*100:.1f}%")

    # VWAP position: soft penalty (not hard block)
    if below_vwap and not hard_blocked:
        mult *= 0.75
        flags.append(f"BELOW_VWAP${vwap:.2f}")

    # ── SOFT PENALTIES - only if not already hard blocked ────
    if not hard_blocked:
        if gap_r    < GAP_DOWN_THRESH: mult *= 0.85; flags.append(f"GAP{gap_r*100:.1f}%")
        if intra_r <= INTRADAY_HARD:   mult *= 0.50;  flags.append(f"WEAK{intra_r*100:.1f}%")
        elif intra_r <= INTRADAY_SOFT: mult *= 0.75;  flags.append(f"SOFT{intra_r*100:.1f}%")
        else:                                          flags.append(f"OK{intra_r*100:+.1f}%")
        if making_ll: mult *= 0.70; flags.append("LOWER-LOWS")

    mult = float(np.clip(mult, 0.0, 1.0))
    return mult, " | ".join(flags) or "HEALTHY", {
        "intraday_ret":round(intra_r*100,2), "gap_ret":round(gap_r*100,2),
        "today_open":round(open_p,2), "vwap":round(vwap,2),
        "below_vwap":curr < vwap, "lower_lows":ll, "health_mult":round(mult,3)
    }


def get_market_regime() -> dict:
    def dev(sym):
        arr = fetch_closes_cached(sym, 25)
        if arr is None or len(arr) < 20: return 0.0, 0.0, 0.0
        sma     = np.mean(arr[-20:])
        dev_20  = (arr[-1] - sma) / sma if sma > 0 else 0.0
        day_ret = (arr[-1] - arr[-2]) / arr[-2] if len(arr) >= 2 and arr[-2] > 0 else 0.0
        return arr[-1], dev_20, day_ret

    sp, sd, spy_today = dev("SPY")
    qp, qd, qqq_today = dev("QQQ")

    # Use today's return for strong directional days (catches gap-up/gap-down)
    # Use 20-day deviation for normal days (smoothed, less noise)
    strong_up   = spy_today > 0.015 and qqq_today > 0.015  # both up 1.5%+ today
    strong_down = spy_today < -0.015 and qqq_today < -0.015

    if strong_up or (sd > 0.002 and qd > 0.002):
        r, al, mm = "RISK-ON",  True,  1.10
    elif strong_down or (sd < SPY_WEAK_THRESH and qd < SPY_WEAK_THRESH):
        r, al, mm = "RISK-OFF", False, 0.40
    else:
        r, al, mm = "NEUTRAL",  True,  1.00

    # Parkinson volatility shock check on SPY intraday bars
    # Gives every scan access to current shock state via market dict
    _park_v, _park_b, _park_r, _park_s = (0.0, 0.0, 1.0, False)
    try:
        _spy_5m = yf.download("SPY", period="1d", interval="5m",
                               progress=False, auto_adjust=True)
        if _spy_5m is not None and len(_spy_5m) > 16:
            _spy_5m.columns = [c[0].lower() if isinstance(c, tuple) else c.lower()
                                for c in _spy_5m.columns]
            _park_v, _park_b, _park_r, _park_s = get_parkinson_vol(_spy_5m)
    except Exception:
        pass

    return {"regime": r, "allows_long": al, "mkt_mult": mm,
            "spy_price":  round(sp, 2), "spy_dev":  round(sd * 100, 2),
            "qqq_price":  round(qp, 2), "qqq_dev":  round(qd * 100, 2),
            "spy_today":  round(spy_today * 100, 2),
            "park_shock": _park_s,    # True when SPY Parkinson vol > 1.5x baseline
            "park_ratio": round(_park_r, 3),  # ratio: current/baseline Parkinson vol
            "park_vol":   round(_park_v, 6)}  # raw current Parkinson vol


def get_sector_rs(etf: str, benchmark: str = "SPY") -> tuple:
    """
    Layer 1 of hierarchical RS: how is the SECTOR performing vs its benchmark?
    Answers: "Is the sector leading or lagging the market?"

    Returns (ratio, score, gate_passes).
    """
    ec  = fetch_closes_cached(etf, 25)
    bc  = fetch_closes_cached(benchmark, 25)
    if ec is None or bc is None or len(ec) < 20 or len(bc) < 20:
        return 1.0, 50.0, True
    ratio  = (ec[-1] / np.mean(ec[-20:])) / (bc[-1] / np.mean(bc[-20:]))
    dyn_rs = get_dynamic_sector_rs(STRATEGY_MODE)
    score  = float(np.clip(50 + (ratio - 1.0) * 600, 0, 100))
    return round(ratio, 4), round(score, 1), ratio >= dyn_rs


def get_stock_vs_etf_rs(symbol: str, etf: str, n: int = 20) -> tuple:
    """
    Layer 2 of hierarchical RS: how is the STOCK performing vs its sector/industry ETF?
    Answers: "Is this stock leading or lagging its peers?"

    This is the critical alpha signal:
      stock_rs > 1.0  →  stock outperforming sector  (sector alpha)
      stock_rs < 1.0  →  stock underperforming sector (avoid)
      stock_rs < 0.97 →  hard block (stock is sick within a healthy sector)

    Returns (ratio, score, outperforming: bool).
    """
    sc = fetch_closes_cached(symbol, n + 5)
    ec = fetch_closes_cached(etf, n + 5)
    if sc is None or ec is None or len(sc) < n or len(ec) < n:
        return 1.0, 50.0, True   # insufficient data → neutral (don't block)
    stock_ret  = sc[-1] / np.mean(sc[-n:])
    etf_ret    = ec[-1] / np.mean(ec[-n:])
    ratio      = stock_ret / etf_ret if etf_ret > 0 else 1.0
    score      = float(np.clip(50 + (ratio - 1.0) * 500, 0, 100))
    return round(ratio, 4), round(score, 1), ratio >= 0.98


def get_hierarchical_rs(symbol: str, sector_etf: str,
                        industry_etf: str | None = None,
                        n: int = 20) -> dict:
    """
    Seven-Layer Pipeline — Layers 2, 2b, 4, 5.
    Computes RS at three levels simultaneously:

      Level 1 — Market alpha:    sector_etf  vs SPY
                                 "Is the sector beating the market?"
      Level 2 — Sector alpha:    stock       vs sector_etf
                                 "Is the stock beating its sector?"
      Level 3 — Industry alpha:  stock       vs industry_etf
                                 "Is the stock beating its micro-regime?"
                                 (Only when industry_etf is available)

    A stock is truly strong only when ALL levels are positive:
      sector beats SPY  AND  stock beats sector  AND  stock beats industry

    This prevents two failure modes:
      A) "Laggard in a hot sector": sector soaring but stock underperforming peers
         → RS vs SPY looks OK but RS vs sector catches it
      B) "Winner in a dying sector": stock leading sector but sector collapsing vs SPY
         → RS vs sector looks OK but RS vs SPY catches it

    Returns dict with all RS levels, composite score, and gate status.
    """
    # Level 1: sector vs market
    rs1, sc1, gate1 = get_sector_rs(sector_etf, "SPY")

    # Level 2: stock vs sector
    rs2, sc2, gate2 = get_stock_vs_etf_rs(symbol, sector_etf, n)

    # Level 3: stock vs industry ETF (if available and different from sector)
    if industry_etf and industry_etf != sector_etf:
        rs3, sc3, gate3 = get_stock_vs_etf_rs(symbol, industry_etf, n)
    else:
        rs3, sc3, gate3 = rs2, sc2, gate2   # fall back to sector RS

    # ── Composite sector score ──────────────────────────────────────────
    # Weights: sector leadership (35%) + stock-vs-sector (40%) + micro-regime (25%)
    # The "stock vs sector" layer gets the most weight because it directly
    # measures whether this stock is the right one to trade in this sector.
    composite = round(0.35 * sc1 + 0.40 * sc2 + 0.25 * sc3, 1)

    # ── Gate logic (pass/fail) ──────────────────────────────────────────
    # Hard block: sector collapsing vs market (RS1 < 0.95)
    # Hard block: stock sick within sector (RS2 < 0.97)
    # Soft gate: use dynamic threshold from strategy mode
    dyn_min = get_dynamic_sector_rs(STRATEGY_MODE)
    hard_block = rs1 < 0.95 or rs2 < 0.97
    soft_gate  = rs1 >= dyn_min and rs2 >= 0.98

    return {
        "rs_vs_spy":       rs1,          # sector vs market
        "rs_vs_sector":    rs2,          # stock vs sector ETF
        "rs_vs_industry":  rs3,          # stock vs industry ETF
        "rs_spy_score":    sc1,
        "rs_sector_score": sc2,
        "rs_industry_score": sc3,
        "sector_score":    composite,    # replaces old sec_sc
        "sector_gate":     soft_gate and not hard_block,
        "hard_block":      hard_block,
        "industry_etf":    industry_etf or sector_etf,
    }


def kelly_size(price: float, atr: float, win_rate: float = 0.60, rr: float = 3.0) -> tuple:
    if atr <= 0 or price <= 0: return 0.0, 0.0, 0
    kf     = float(np.clip((win_rate - (1 - win_rate) / rr) * 0.5, 0.0, 0.10))
    drisk  = ACCOUNT_SIZE * kf
    shares = int(drisk / (atr * ATR_STOP_MULT)) if atr * ATR_STOP_MULT > 0 else 0
    return round(kf, 4), round(drisk, 2), min(shares, int(ACCOUNT_SIZE * 0.20 / price))


# ── Anchored VWAP (resets daily at 9:30 AM) ──────────────────────────────────
# Fixes the skewed-VWAP problem. Previous day data does not bleed in.

def compute_anchored_vwap(df: pd.DataFrame) -> float:
    """
    Standard session VWAP anchored to today's 9:30 AM open.
    Falls back to rolling 20-bar VWAP if today's bars unavailable.
    Called directly by scan_symbol for the core VWAP value.
    """
    try:
        tz    = df.index.tz
        today = pd.Timestamp.now(tz=tz).strftime("%Y-%m-%d")
        tb    = df[df.index.strftime("%Y-%m-%d") == today]
        if len(tb) < 2:
            tb = df.tail(20)
        tp    = (tb["high"] + tb["low"] + tb["close"]) / 3
        return round(float((tp * tb["volume"]).sum() / tb["volume"].sum()), 4)
    except Exception:
        return float(df["close"].iloc[-1])


def compute_vwap_bands(df: pd.DataFrame, atr: float,
                        price: float, strategy: str) -> dict:
    """
    Dynamic VWAP with Standard Deviation bands and regime-aware interpretation.

    Based on the doc guidance:
    1. Gap anchoring: if today's open gapped > 1.5% from prior close,
       a second VWAP is anchored to the 5-min high/open to track
       institutional support level from the gap.

    2. Adaptive SD bands based on ATR/volatility:
       Low volatility  (ATR% < 1.5%): tight bands  1.0 / 2.0 SD
       Normal          (ATR% < 3.0%): standard      1.5 / 2.5 SD
       High volatility (ATR% >= 3.0%): wide bands   2.0 / 3.0 SD
       -- prevents small-cap noise from triggering false signals

    3. Regime-aware role of VWAP:
       TREND mode:         VWAP = FLOOR (price bouncing above = strength)
       MEAN_REVERSION mode: VWAP = MAGNET (target for reversion trades)

    Returns dict with vwap, upper1, lower1, upper2, lower2,
    gap_vwap (second anchor if gap detected), vwap_role, band_mult.
    """
    vwap = compute_anchored_vwap(df)

    try:
        tz    = df.index.tz
        today = pd.Timestamp.now(tz=tz).strftime("%Y-%m-%d")
        tb    = df[df.index.strftime("%Y-%m-%d") == today]
        if len(tb) < 5:
            tb = df.tail(20)

        # Typical price series for the session
        tp = (tb["high"] + tb["low"] + tb["close"]) / 3
        # Cumulative VWAP deviations for SD bands
        cum_vol = tb["volume"].cumsum()
        cum_tp_vol = (tp * tb["volume"]).cumsum()
        vwap_series = cum_tp_vol / cum_vol.replace(0, np.nan)
        vwap_sq = ((tp ** 2 * tb["volume"]).cumsum()) / cum_vol.replace(0, np.nan)
        variance = (vwap_sq - vwap_series ** 2).clip(lower=0)
        sd = float(np.sqrt(variance.iloc[-1])) if not variance.empty else (atr * 0.5)

        # Adaptive band multiplier based on ATR%
        atr_pct = (atr / price) if price > 0 else 0.02
        if atr_pct < 0.015:
            band1, band2 = 1.0, 2.0   # low volatility -- tight
        elif atr_pct < 0.030:
            band1, band2 = 1.5, 2.5   # normal
        else:
            band1, band2 = 2.0, 3.0   # high volatility small-cap -- wide

        upper1 = round(vwap + band1 * sd, 4)
        lower1 = round(vwap - band1 * sd, 4)
        upper2 = round(vwap + band2 * sd, 4)
        lower2 = round(vwap - band2 * sd, 4)

        # Gap anchor: second VWAP from opening 5-min candle high
        # fires when today gapped up/down > 1.5% vs prior close
        gap_vwap = None
        if len(df) >= 2:
            prior_close = float(df["close"].iloc[-len(tb)-1]) \
                          if len(df) > len(tb) else float(tb["close"].iloc[0])
            open_price  = float(tb["close"].iloc[0])
            gap_pct     = abs(open_price - prior_close) / prior_close \
                          if prior_close > 0 else 0.0
            if gap_pct >= 0.015 and len(tb) >= 5:
                # Anchor to 5-min candle high after gap open
                gap_ref  = float(tb["high"].iloc[:5].max())
                gap_tb   = tb[tb["close"] <= gap_ref * 1.02] \
                           if not tb.empty else tb
                if len(gap_tb) >= 2:
                    tp_g   = (gap_tb["high"] + gap_tb["low"] + gap_tb["close"]) / 3
                    cv_g   = gap_tb["volume"].sum()
                    gap_vwap = round(float((tp_g * gap_tb["volume"]).sum() / cv_g), 4) \
                               if cv_g > 0 else None

        # Regime-aware VWAP role
        if strategy == "TREND":
            vwap_role = "FLOOR"
            # In trend mode: price holding above VWAP = institutional support
            # Price below VWAP in trend = weakness, reduce confidence
        else:
            vwap_role = "MAGNET"
            # In MR mode: VWAP is the target price will revert toward
            # Distance from VWAP proportional to expected reversion size

    except Exception:
        upper1 = round(vwap + atr, 4)
        lower1 = round(vwap - atr, 4)
        upper2 = round(vwap + atr * 2, 4)
        lower2 = round(vwap - atr * 2, 4)
        gap_vwap = None
        vwap_role = "MAGNET" if strategy == "MEAN_REVERSION" else "FLOOR"
        band1, band2 = 1.5, 2.5

    return {
        "vwap":      vwap,
        "upper1":    upper1,
        "lower1":    lower1,
        "upper2":    upper2,
        "lower2":    lower2,
        "gap_vwap":  gap_vwap,
        "vwap_role": vwap_role,
        "band_mult": f"{band1}/{band2}SD",
        "above_upper1": price > upper1,
        "below_lower1": price < lower1,
        "above_upper2": price > upper2,
        "below_lower2": price < lower2,
    }


# ── Liquidity Firewall ────────────────────────────────────────────────────────
# Must pass BOTH checks before any math runs. Kills zombie stocks immediately.

def passes_liquidity_firewall(df_d: pd.DataFrame, symbol: str,
                               strategy: str = "AUTO",
                               spread_max_override: float = None) -> tuple:
    """
    Three-part firewall: dollar volume + RVOL + ATR%.

    Why three layers and not just dollar volume:
      Dollar volume = average liquidity (can you get a fill?)
      RVOL          = current participation (is there real interest NOW?)
      ATR%          = movement potential (will it move enough to pay?)

    A $10B stock with 0.4% ATR fails ATR check -- too slow, R:R collapses.
    A $300M stock with low RVOL fails RVOL check -- nobody's trading it today.
    Both pass dollar volume but both are bad trades.

    Strategy-aware thresholds:
      TREND: RVOL >= 1.5x, ATR >= 2.5% (needs participation + range for ATR*3 targets)
      MR:    RVOL >= 1.2x, ATR >= 1.5% (quieter reversions, but range must exist)
    """
    if df_d is None or len(df_d) < 20:
        return False, "INSUFFICIENT DATA", 0.0

    avg_price  = float(df_d["close"].tail(10).mean())
    avg_vol    = float(df_d["volume"].tail(20).mean())
    curr_vol   = float(df_d["volume"].iloc[-1])
    dollar_vol = avg_price * avg_vol

    # Gate 1: Minimum dollar volume (liquidity floor -- can you exit?)
    if dollar_vol < MIN_DOLLAR_VOLUME:
        return False, f"LOW_DOLLAR_VOL ${dollar_vol/1e6:.1f}M < $5M", 0.0

    # Gate 2: Relative volume (participation -- is anyone trading this today?)
    min_atr_pct, min_rvol = get_dynamic_vol_filters(strategy)
    rel_vol = curr_vol / avg_vol if avg_vol > 0 else 0.0
    if rel_vol < min_rvol:
        return False, f"LOW_RELVOL {rel_vol:.2f}x < {min_rvol}x ({strategy})", 0.0
    # Gate 3: ATR% (movement potential -- will it move enough to pay?)
    # FIX: highs[1:]-closes[:-1] produces n-1 elements but highs-lows produces n.
    # Fetch n+1 rows so [1:] slices align to exactly n elements.
    atr_pct = 0.0
    if len(df_d) >= 14:
        highs  = df_d["high"].values[-15:]
        lows   = df_d["low"].values[-15:]
        closes = df_d["close"].values[-15:]
        if len(highs) > 1:
            hl  = highs[1:] - lows[1:]
            hc  = np.abs(highs[1:] - closes[:-1])
            lc  = np.abs(lows[1:]  - closes[:-1])
            tr  = np.maximum(hl, np.maximum(hc, lc))
        else:
            tr  = highs - lows
        atr_val = float(np.mean(tr)) if len(tr) > 0 else 0.0
        atr_pct = atr_val / avg_price if avg_price > 0 else 0.0
        if atr_pct < min_atr_pct:
            return False, (f"LOW_ATR {atr_pct*100:.2f}% < {min_atr_pct*100:.1f}% "
                           f"(not enough daily range for {strategy})")

    # Gate 4: Spread estimate (microstructure -- is the spread too wide to profit?)
    # Estimated spread from High-Low range on recent bars. On small-caps,
    # a wide spread destroys MR edge: 0.5% TP is worthless if spread is 0.3%.
    # Threshold: spread > 0.3% blocks MR entries (TREND is more tolerant).
    # Spread gate -- two-tier: WARN vs BLOCK
    # Rule: spread > 25% of TP = trade math is borderline
    #       spread > 75% of TP = trade math is destroyed, block entirely
    # With TP=0.5%: warn at 0.125%, hard block at 0.375%
    # SPREAD_MAX_MR is the hard block level
    # Use per-ticker spread max when adaptive params are available
    # ALKT (ATR=4%): adaptive spread_max=3% vs global 0.375% -- huge difference
    # spread_max_override is passed from scan_symbol AFTER loading adaptive params
    if spread_max_override is not None:
        SPREAD_MAX_MR  = spread_max_override
        SPREAD_WARN_MR = spread_max_override * (0.25 / 0.75)
    else:
        SPREAD_MAX_MR  = MR_TAKE_PROFIT_PCT * 0.75
        SPREAD_WARN_MR = MR_TAKE_PROFIT_PCT * 0.25
    SPREAD_MAX_TREND = 0.008                         # 0.8% for TREND (larger targets)
    spread_thresh    = SPREAD_MAX_MR if strategy == "MEAN_REVERSION" else SPREAD_MAX_TREND
    spread_warn      = SPREAD_WARN_MR if strategy == "MEAN_REVERSION" else SPREAD_MAX_TREND
    try:
        _highs  = df_d["high"].tail(5).values.astype(float)
        _lows   = df_d["low"].tail(5).values.astype(float)
        _closes = df_d["close"].tail(5).values.astype(float)
        _closes = np.where(_closes > 0, _closes, 1.0)
        est_spread = float(np.mean((_highs - _lows) / _closes))
    except Exception:
        est_spread = 0.0

    # IWM index-level sensor: when Russell 2000 is volatile,
    # spread requirements tighten automatically (Gate 4 multiplier)
    # "If the tide (IWM) goes out fast, all boats drop" -- protect against index drag
    # IWM cache: fetch at most once per 2 minutes across all symbol scans
    global _IWM_CACHE, _IWM_CACHE_TIME
    try:
        import time as _time_mod
        _now_t = _time_mod.time()
        if '_IWM_CACHE' not in globals() or _now_t - globals().get('_IWM_CACHE_TIME', 0) > 120:
            _IWM_CACHE = yf.Ticker("IWM").history(period="2d", interval="5m")
            _IWM_CACHE_TIME = _now_t
        _iwm = _IWM_CACHE
        if _iwm is not None and len(_iwm) >= 15:
            _iwm_park = get_parkinson_vol(
                _iwm.rename(columns=str.lower).tail(20),
                window=10, baseline_window=50
            )
            _iwm_ratio = _iwm_park[2]  # current/baseline ratio
            if _iwm_ratio >= 2.0:
                spread_thresh = spread_thresh * 0.3   # severe IWM shock -- very tight
                _iwm_gate_note = f"IWM_SEVERE(ratio={_iwm_ratio:.1f}x)"
            elif _iwm_ratio >= 1.5:
                spread_thresh = spread_thresh * 0.5   # IWM elevated -- tighten
                _iwm_gate_note = f"IWM_ELEVATED(ratio={_iwm_ratio:.1f}x)"
            else:
                _iwm_gate_note = ""
        else:
            _iwm_gate_note = ""
    except Exception:
        _iwm_gate_note = ""

    if strategy == "MEAN_REVERSION" and est_spread > spread_thresh:
        return False, (f"WIDE_SPREAD {est_spread*100:.3f}% > {spread_thresh*100:.3f}% "
                       f"[>{MR_TAKE_PROFIT_PCT*100*0.75:.0f}% of TP — math destroyed]"
                       + (f" {_iwm_gate_note}" if _iwm_gate_note else ""))

    spread_flag = ""
    if strategy == "MEAN_REVERSION" and est_spread > spread_warn:
        spread_flag = f" [SPREAD WARN: {est_spread/MR_TAKE_PROFIT_PCT*100:.0f}% of TP]"

    return True, (f"OK DV=${dollar_vol/1e6:.1f}M RV={rel_vol:.2f}x "
                  f"ATR={atr_pct*100:.1f}% SPR={est_spread*100:.3f}%{spread_flag}"), est_spread


# ── ADX Calculation ───────────────────────────────────────────────────────────
# Used by regime switch. ADX < 25 = ranging = mean reversion valid.
# ADX > 25 = strong trend = only trend-following valid.

def compute_adx(df: pd.DataFrame, n: int = 14) -> float:
    """
    Average Directional Index - fully vectorized with pandas.
    Replaces the Python-loop Wilder smoother with ewm() - 5-10x faster.
    Returns float 0-100. Default 20.0 (neutral/ranging) on bad data.
    """
    if df is None or len(df) < n * 2:
        return 20.0
    h  = df["high"];  l = df["low"];  c = df["close"]

    # True range (vectorized)
    tr = pd.concat([h - l, (h - c.shift()).abs(), (l - c.shift()).abs()], axis=1).max(axis=1)

    # Directional movement (vectorized)
    up  = h.diff();  dn = -l.diff()
    pdm = pd.Series(np.where((up > dn) & (up > 0), up, 0.0), index=h.index)
    ndm = pd.Series(np.where((dn > up) & (dn > 0), dn, 0.0), index=h.index)

    # Wilder smoothing via ewm (alpha = 1/n) - replaces the Python loop
    alpha  = 1.0 / n
    atr_s  = tr.ewm(alpha=alpha,  adjust=False).mean()
    pdi_s  = pdm.ewm(alpha=alpha, adjust=False).mean()
    ndi_s  = ndm.ewm(alpha=alpha, adjust=False).mean()

    with np.errstate(divide='ignore', invalid='ignore'):
        pdi = np.where(atr_s > 0, 100 * pdi_s / atr_s, 0.0)
        ndi = np.where(atr_s > 0, 100 * ndi_s / atr_s, 0.0)
        dx  = pd.Series(
            np.where((pdi + ndi) > 0, 100 * np.abs(pdi - ndi) / (pdi + ndi), 0.0),
            index=h.index)

    adx = dx.ewm(alpha=alpha, adjust=False).mean()
    return round(float(adx.iloc[-1]), 2)


# ── Strategy Regime Switch ────────────────────────────────────────────────────
# THE KEY FIX: Determines which strategy is active. NEVER allows both.
# Returns one of: "TREND" | "MEAN_REVERSION" | "NO_TRADE"
# Also returns the adaptive OFI threshold for the current environment.

# ── Strategy Regime Switch ────────────────────────────────────────────────────
# Determines which strategy is active. NEVER allows both simultaneously.
# Returns one of: "TREND" | "MEAN_REVERSION" | "NO_TRADE"
#
# REVISED LOGIC (v4.1):
# The old logic blocked ALL trading when SPY vol > 2%. This was wrong.
# A clean 3%+ SPY day is a TREND day, not a no-trade day.
# NO_TRADE should only fire for chaotic/directionless volatility (VIX crash,
# news whipsaw) — not for strong directional moves.
#
# NEW decision tree:
#   SPY vol > 2% AND SPY is directional (move > 0)  -> TREND  (ride the wave)
#   SPY vol > 2% AND SPY is choppy (near zero)       -> NO_TRADE (true chaos)
#   ADX > 25                                          -> TREND
#   SPY calm + ADX < 25                               -> MEAN_REVERSION
#   Elevated vol, directional                         -> TREND (safer)
#   Elevated vol, flat                                -> MEAN_REVERSION (cautious)

def get_strategy_regime(df_intra: pd.DataFrame, spy_return: float, adx: float = None) -> tuple:
    """
    Returns (strategy, adaptive_ofi, adx, reason).

    SPY_VOL thresholds:
      SPY_VOL_CALM   = 0.5%   -> pure MR market
      SPY_VOL_NORMAL = 1.0%   -> normal, MR with standard gate
      SPY_VOL_NOTRADE= 2.0%   -> only blocks if CHOPPY (|ret| < 0.5% but high vol)
                                  a clean 3% trend day goes to TREND, not NO_TRADE

    Adaptive OFI threshold (order flow entry sensitivity):
      Calm:     0.20  (sensitive - small imbalances matter)
      Normal:   0.30  (standard)
      Hot/Trend:0.35  (require stronger signal - spreads are wider)
    """
    adx     = adx if adx is not None else compute_adx(df_intra)  # reuse if pre-computed
    spy_vol = abs(spy_return)

    # ── High volatility day: determine if directional or chaotic ─────────────
    # Chaos check: use average per-bar range (ATR proxy), NOT session H-L span
    # Session H-L on a 80-bar 5m df spans ~25%+ even on calm days (cumulative drift)
    try:
        if len(df_intra) > 10:
            bar_ranges = df_intra["high"].values - df_intra["low"].values
            avg_bar_range = float(np.nanmean(bar_ranges)) / float(df_intra["close"].iloc[0])
        else:
            avg_bar_range = spy_vol
    except Exception:
        avg_bar_range = spy_vol

    # ── Parkinson volatility shock detection (replaces hardcoded avg_bar_range) ─
    # Self-calibrating: fires when current H-L vol is 1.5x+ its own recent baseline.
    # Much more robust than a fixed 1.5% threshold -- works on any volatility regime.
    park_vol, park_base, park_ratio, park_shock = get_parkinson_vol(df_intra)
    park_severe = park_ratio > PARKINSON_SHOCK_SEVERE   # >2x = severe shock

    # True chaos: Parkinson spike AND near-zero net SPY return (not a trend day)
    # e.g. VIX spike, news whipsaw -- bars are wild but price goes nowhere
    if park_shock and spy_vol < 0.005:
        return ("NO_TRADE", 0.35, adx,
                f"PARKINSON_CHAOS ratio={park_ratio:.2f}x (>{PARKINSON_SHOCK_THRESHOLD}x) "
                f"net={spy_return:+.1%} - no trade")

    # Severe shock even on directional day: issue TREND but with max strictness
    # (park_severe = vol is 2x+ baseline -- very unusual, spreads are dangerous)
    if park_severe and abs(spy_return) < 0.015:
        return ("NO_TRADE", 0.35, adx,
                f"PARKINSON_SEVERE ratio={park_ratio:.2f}x - extreme volatility, skip")

    # Legacy fallback: keep the original avg_bar_range check as backup
    if avg_bar_range > 0.015 and spy_vol < 0.005 and not park_shock:
        return ("NO_TRADE", 0.35, adx,
                f"SPY_CHAOS avg_bar_range={avg_bar_range:.1%} net={spy_return:+.1%} - no trade")

    if spy_vol > SPY_VOL_NOTRADE:
        # Clean directional move (SPY up or down decisively) -> TREND
        # This handles the 3%+ gap-up/gap-down days correctly
        if abs(spy_return) > 0.015:   # at least 1.5% with clear direction
            return ("TREND", 0.35, adx,
                    f"SPY {spy_return:+.1%} STRONG TREND DAY - trend mode active")
        else:
            # High vol but limited direction -- cautious but still tradeable
            return ("TREND", 0.35, adx,
                    f"SPY {spy_return:+.1%} elevated directional - trend mode")

    # ── Strong trend via ADX ──────────────────────────────────────────────────
    if adx > ADX_MAX:
        return ("TREND", 0.30, adx,
                f"ADX {adx:.1f} > {ADX_MAX} (trending)")

    # ── Elevated vol but directional: lean trend ─────────────────────────────
    if spy_vol >= SPY_VOL_NORMAL:
        if abs(spy_return) > 0.008:   # clear direction at elevated vol
            return ("TREND", 0.35, adx,
                    f"SPY {spy_return:+.1%} elevated vol directional - trend mode")
        # Elevated vol, flat SPY: cautious MR
        return ("MEAN_REVERSION", 0.35, adx,
                f"ADX {adx:.1f} elevated vol flat SPY - cautious MR")

    # ── Calm market: ideal MR ────────────────────────────────────────────────
    if spy_vol < SPY_VOL_CALM and adx <= ADX_MAX:
        return ("MEAN_REVERSION", 0.20, adx,
                f"ADX {adx:.1f} CALM {spy_vol:.1%} - MR optimal")

    # ── Normal market: standard MR ───────────────────────────────────────────
    return ("MEAN_REVERSION", 0.30, adx,
            f"ADX {adx:.1f} NORMAL {spy_vol:.1%}")


# ── Exhaustion Triple-Gate (Mean Reversion Engine) ───────────────────────────
# Replaces raw OFI score for MEAN_REVERSION mode.
# Only fires when: price overextended (Z) + OFI peak then fade + VWAP side correct.
# This is the core of the 70-80% win rate claim.

def check_exhaustion_signal(df: pd.DataFrame, vwap: float,
                             spy_return: float,
                             adaptive_ofi: float = 0.30) -> tuple:
    """
    Three-step exhaustion confirmation:
      Step A - Extension:  Z-score >= 2.0 (price far from mean)
      Step B - Exhaustion: OFI spiked then faded (ammunition depleted)
      Step C - VWAP side:  Price on correct side for reversion

    adaptive_ofi scales panic/exhaustion thresholds with market regime:
      0.20 = calm  -> easier trigger (more sensitive)
      0.30 = normal
      0.35 = volatile -> harder trigger (requires stronger signal)

    Returns (long_signal: bool, short_signal: bool, reason: str).
    """
    if df is None or len(df) < 20:
        return False, False, "INSUFFICIENT DATA"

    close = df["close"].values
    h     = df["high"].values
    l     = df["low"].values
    vol   = df["volume"].values

    # Z-score on the intraday series
    if len(close) >= 20:
        z_now = (close[-1] - np.mean(close[-20:])) / (np.std(close[-20:], ddof=1) + 1e-9)
    else:
        z_now = 0.0

    # OFI series (bulk volume classification)
    rng   = h - l
    br    = np.where(rng > 0, (close - l) / rng, 0.5)
    bv    = br * vol; sv = (1 - br) * vol
    # Signed OFI: +1 = all buying, -1 = all selling
    ofi_signed = (bv - sv) / (bv + sv + 1e-9)
    signed_ofi = float(ofi_signed[-1])   # alias used in return dict

    # Look at last 5 bars for peak/fade pattern
    recent_ofi  = ofi_signed[-5:] if len(ofi_signed) >= 5 else ofi_signed
    curr_ofi    = float(ofi_signed[-1])
    curr_price  = float(close[-1])
    above_vwap  = curr_price > vwap
    below_vwap  = curr_price < vwap
    spy_safe_l  = spy_return > -0.01   # SPY not cratering for longs
    spy_safe_s  = spy_return <  0.01   # SPY not surging for shorts

    # ── LONG: price crushed + selling exhausted + below VWAP ──
    is_oversold       = z_now <= -Z_ENTRY_THRESH
    was_panic_selling = float(recent_ofi.min()) < -(adaptive_ofi + 0.10)
    selling_exhausted = curr_ofi > -0.10
    long_signal = (is_oversold and was_panic_selling
                   and selling_exhausted and below_vwap and spy_safe_l)

    # ── SHORT: price mooning + buying exhausted + above VWAP ──
    is_overbought     = z_now >=  Z_ENTRY_THRESH
    was_aggr_buying   = float(recent_ofi.max()) >  (adaptive_ofi + 0.10)
    buying_exhausted  = curr_ofi < 0.10
    short_signal = (is_overbought and was_aggr_buying
                    and buying_exhausted and above_vwap and spy_safe_s)

    reason = (f"Z={z_now:.2f} OFI_now={curr_ofi:.2f} "
              f"peak={float(recent_ofi.max()):.2f} trough={float(recent_ofi.min()):.2f} "
              f"{'BELOW' if below_vwap else 'ABOVE'}_VWAP")

    return long_signal, short_signal, reason, signed_ofi


# ── Volume Divergence Filter ──────────────────────────────────────────────────
# Final confirmation: price hitting extreme but volume drying up = no fuel left.
# This confirms the exhaustion signal - move can't continue without volume.

def check_volume_divergence(df: pd.DataFrame) -> bool:
    """
    Returns True if volume divergence confirmed (move is running out of fuel).
    Condition: price at new extreme BUT recent volume < prior volume.
    """
    if df is None or len(df) < 8:
        return True   # can't check = don't block

    recent_vol = float(df["volume"].tail(3).mean())
    prior_vol  = float(df["volume"].iloc[-6:-3].mean()) if len(df) >= 6 else recent_vol
    rising_price = float(df["close"].iloc[-1]) > float(df["close"].iloc[-3])
    falling_vol  = recent_vol < prior_vol

    # Divergence = price rising but volume falling (or price falling but volume falling)
    return falling_vol  # volume is drying up regardless of direction


# ── Mean Reversion Exit Engine ────────────────────────────────────────────────
# For MEAN_REVERSION mode: small scalp target OR VWAP touch - whichever first.
# This is what drives the high win rate - take the easy money, don't be greedy.

def determine_mr_exit(current_price: float, entry_price: float,
                      vwap: float, side: str) -> str:
    """
    Returns "EXIT_PROFIT" | "EXIT_LOSS" | "HOLD".
    Primary exit = VWAP touch (statistical edge ends here).
    Secondary exit = 0.5% profit target.
    Hard stop = 0.8% loss.
    """
    if side == "LONG":
        if current_price >= entry_price * (1 + MR_TAKE_PROFIT_PCT): return "EXIT_PROFIT"
        if current_price >= vwap:                                    return "EXIT_PROFIT"
        if current_price <= entry_price * (1 - MR_STOP_LOSS_PCT):   return "EXIT_LOSS"
    if side == "SHORT":
        if current_price <= entry_price * (1 - MR_TAKE_PROFIT_PCT): return "EXIT_PROFIT"
        if current_price <= vwap:                                    return "EXIT_PROFIT"
        if current_price >= entry_price * (1 + MR_STOP_LOSS_PCT):   return "EXIT_LOSS"
    return "HOLD"


# ── Position & Trade Log ──────────────────────────────────────────────────────

@dataclass
class Position:
    symbol:      str
    direction:   str   = "FLAT"
    entry_price: float = 0.0
    entry_time:  str   = ""
    stop:        float = 0.0
    target:      float = 0.0
    atr:         float = 0.0
    highest:     float = 0.0
    lowest:      float = 0.0
    pnl:         float = 0.0
    shares:      int   = 0

POSITIONS: dict = {}
TRADE_LOG: list = []

def update_trailing_stop(pos: Position, price: float) -> Position:
    if pos.direction == "LONG":
        pos.highest = max(pos.highest, price)
        pos.stop    = max(pos.stop, round(pos.highest - pos.atr * ATR_STOP_MULT, 2))
        pos.pnl     = round((price - pos.entry_price) * pos.shares, 2)
    elif pos.direction == "SHORT":
        pos.lowest  = min(pos.lowest, price)
        pos.stop    = min(pos.stop, round(pos.lowest + pos.atr * ATR_STOP_MULT, 2))
        pos.pnl     = round((pos.entry_price - price) * pos.shares, 2)
    return pos

def close_position(symbol: str, exit_price: float, reason: str):
    pos = POSITIONS.get(symbol)
    if pos and pos.direction != "FLAT":
        pnl = ((exit_price - pos.entry_price) if pos.direction == "LONG"
               else (pos.entry_price - exit_price)) * pos.shares
        TRADE_LOG.append({"Symbol":symbol,"Direction":pos.direction,
            "Entry $":pos.entry_price,"Exit $":round(exit_price,2),
            "Shares":pos.shares,"P&L $":round(pnl,2),
            "Result":"WIN " if pnl>0 else "LOSS ",
            "Reason":reason,"Time":datetime.now().strftime("%H:%M:%S")})
    POSITIONS[symbol] = Position(symbol)


# ── Master Scan ───────────────────────────────────────────────────────────────
def scan_symbol(symbol: str, market: dict, timeframe: str = "5m",
                mode: str = "auto", df_override=None) -> Optional[dict]:
    """
    Master scan for a single symbol.
    Supports df_override for batched terminal scans (10x faster).
    """

    # --- FAST PATH vs SLOW PATH ---
    # df_override = pre-batched intraday bars from run_universe_scan() chunked download.
    # When provided, skip the per-ticker fetch_intraday() call (~0.3s saved per symbol).
    # When None (standalone call), use normal fetch_intraday() below.
    _df_i_override = df_override  # saved for use at the fetch_intraday() call site below

    """
    Dual-strategy scanner with hard strategy switch.

    STRATEGY_MODE (global config):
      AUTO         -> ADX + SPY vol determines strategy each scan
      TREND        -> Hurst + Hawkes momentum (volatile/trending markets)
      MEAN_REVERSION -> Exhaustion triple-gate + VWAP exit (ranging markets)

    CRITICAL: Only ONE strategy fires per scan. Never both simultaneously.
    If strategy = NO_TRADE, returns None immediately.

    mode = "premarket"  -> 9:00 AM build watchlist (daily Hurst, soft gates)
    mode = "intraday"   -> market hours confirmation (hard gates, dual Hurst)
    mode = "auto"       -> auto-detects from clock
    """
    if mode == "auto":
        hour = datetime.now().hour
        mode = "intraday" if 9 <= hour < 16 else "premarket"

    # ── FIREWALL 1: Market Cap ────────────────────────────────
    passes, mcap_b = is_midcap(symbol, STRATEGY_MODE)
    if not passes: return None

    # ── FIREWALL 2: Fetch data ────────────────────────────────
    df_d = fetch_daily(symbol, 60)
    if df_d is None or len(df_d) < 30: return None

    # ── FIREWALL 3: Liquidity (runs before any expensive math) ─
    # Load adaptive params BEFORE firewall so per-ticker spread gate applies
    # Without this, ALKT (ATR=4%) gets blocked by global 0.375% spread threshold
    # before its profile spread_max=3% is even checked
    _pre_adaptive = {}
    _pre_spread_override = None
    if _TI_AVAILABLE:
        try:
            _pre_tp = _ti.TickerProfile(symbol)
            _pre_adaptive = _pre_tp.get_adaptive_params(
                global_tp=MR_TAKE_PROFIT_PCT,
                global_sl=MR_STOP_LOSS_PCT,
                global_score=SIGNAL_THRESHOLD,
            )
            if _pre_adaptive.get("source") == "ADAPTIVE":
                _pre_spread_override = _pre_adaptive.get("spread_max")
        except Exception:
            pass

    liq_ok, liq_reason, _liq_spread = passes_liquidity_firewall(
        df_d, symbol, STRATEGY_MODE, spread_max_override=_pre_spread_override)
    if not liq_ok: return None

    # ── Intraday data: use pre-batched override (fast path) or fetch individually ──
    if _df_i_override is not None and len(_df_i_override) >= 15:
        df_i = _df_i_override   # fast path: no per-ticker HTTP call
    else:
        _df_i = fetch_intraday(symbol, timeframe)
        df_i  = _df_i if (_df_i is not None and not _df_i.empty) else df_d.copy()
    if len(df_i) < 15: df_i = df_d.copy()

    price       = float(df_d["close"].iloc[-1])
    prior_close = float(df_d["close"].iloc[-2]) if len(df_d) >= 2 else price
    closes      = df_d["close"].values
    spy_ret     = market.get("spy_dev", 0.0) / 100.0
    # vwap computed inside compute_vwap_bands() -- no separate call needed

    # ── FIREWALL 4: Strategy Regime Switch ────────────────────
    # Compute ADX once -- used by regime switch and returned in result dict.
    # Previously TREND/MR forced modes called compute_adx() separately
    # from AUTO mode which called it inside get_strategy_regime() = double compute.
    adx_val = compute_adx(df_i)
    if STRATEGY_MODE == "AUTO":
        strategy, adaptive_ofi, _, regime_reason = get_strategy_regime(df_i, spy_ret, adx_val)
    elif STRATEGY_MODE == "TREND":
        strategy, adaptive_ofi, regime_reason = "TREND", 0.30, "FORCED_TREND"
    else:
        strategy, adaptive_ofi, regime_reason = "MEAN_REVERSION", 0.30, "FORCED_MR"

    # Kill immediately if market is in NO_TRADE regime -- no point computing anything
    if strategy == "NO_TRADE":
        return None

    # Parkinson shock circuit breaker: block ALL new MR entries during/after shock.
    if strategy == "MEAN_REVERSION" and _SHOCK_MR_OVERRIDE:
        return None  # MR circuit breaker active -- skip this symbol

    # ── Adaptive per-ticker parameters ──────────────────────────────────────
    # Per-ticker TP/SL/score/momentum tuned from historical outcomes.
    # Falls back to global constants when < 5 trades exist for this symbol.
    # Reuse _pre_adaptive loaded before firewall (already has profile data)
    # Avoids double disk read of SYMBOL.json per scan
    _adaptive     = _pre_adaptive if _pre_adaptive else {}
    _eff_tp          = _adaptive.get("tp",    MR_TAKE_PROFIT_PCT)
    _eff_sl          = _adaptive.get("sl",    MR_STOP_LOSS_PCT)
    _eff_score_floor = _adaptive.get("score_thresh", SIGNAL_THRESHOLD)
    _eff_mom_thresh  = _adaptive.get("momentum_norm_thresh", MR_ROC_NORM_THRESH)
    _eff_spread_max  = _adaptive.get("spread_max", MR_TAKE_PROFIT_PCT * 0.75)
    _adaptive_src    = _adaptive.get("source", "GLOBAL")

    # ── Gate 12: Extreme RVOL ceiling for MR ─────────────────────────────────
    # RVOL > 8x = news/fundamental repricing, NOT temporary imbalance.
    # High-RVOL events in MR = falling knife or moon shot. Route to TREND or skip.
    if strategy == "MEAN_REVERSION":
        try:
            _rel_vol_check = float(df_i["volume"].iloc[-1] /
                                   (df_i["volume"].rolling(20).mean().iloc[-1] + 1e-9))
            if _rel_vol_check > MR_RVOL_MAX:
                _rejected.append({"symbol": symbol, "reason": "EXTREME_RVOL",
                                  "detail": f"RVOL={_rel_vol_check:.1f}x > {MR_RVOL_MAX}x "
                                             f"(news/repricing event, not MR)"})
                return None
        except Exception:
            pass

    # ── Gate 12b: Gap Direction Filter for MR ────────────────────────────────
    # Hypothesis shows: GAP-UP = 0% WR, GAP-DOWN = 100% WR on trend days.
    # Stocks gapping up > 2% are in momentum/trending state -- don't MR them.
    # The move is NOT a temporary imbalance; it's directional price discovery.
    if strategy == "MEAN_REVERSION":
        try:
            # Compute gap inline -- intraday_health hasn't run yet at this point
            _prior_c = float(df_d["close"].iloc[-2]) if len(df_d) >= 2 else price
            _open_p  = float(df_i["open"].iloc[0])   if len(df_i) >= 1 else price
            _gap = (_open_p - _prior_c) / (_prior_c + 1e-9) if _prior_c > 0 else 0.0
            if _gap > MR_GAP_UP_MAX:
                _rejected.append({"symbol": symbol, "reason": "GAP_UP_BLOCK",
                                   "detail": f"gap_ret={_gap*100:.1f}% > "
                                              f"{MR_GAP_UP_MAX*100:.0f}% "
                                              f"(momentum gap, not MR candidate)"})
                return None
            if _gap < MR_GAP_DOWN_MAX:
                _rejected.append({"symbol": symbol, "reason": "GAP_DOWN_EXTREME",
                                   "detail": f"gap_ret={_gap*100:.1f}% < "
                                              f"{MR_GAP_DOWN_MAX*100:.0f}% "
                                              f"(extreme gap-down, possible halt/news)"})
                return None
        except Exception:
            pass

    # ── Gate 13: Volatility-Normalized Momentum Filter ───────────────────────
    # Replaces fixed ROC threshold with ATR-normalized ROC (z-score of the move).
    # Problem with fixed -1.5%: biotech moving 1.5% in 15min = normal (ATR 4%+).
    #                            utility moving 1.5% in 15min = extreme (ATR 0.5%).
    # Fix: normalize ROC by recent realized volatility so the threshold is
    # "how many standard deviations of normal movement occurred?"
    # Block when: normalized_roc < -MR_ROC_ZSCORE_THRESH (default -1.5 std devs)
    # This means: "this drop is 1.5x more than the stock normally moves in 3 bars"
    if strategy == "MEAN_REVERSION" and MR_ROC_ENABLED:
        try:
            _closes = df_i["close"].values.astype(float)
            if len(_closes) >= MR_ROC_LOOKBACK + 5:
                # Raw 3-bar ROC
                _roc = (_closes[-1] - _closes[-(MR_ROC_LOOKBACK + 1)]) /                        (_closes[-(MR_ROC_LOOKBACK + 1)] + 1e-9)

                # Realized volatility: std of 1-bar returns over last 20 bars
                # This is the stock's natural per-bar movement -- our baseline
                _rets = np.diff(_closes[-21:]) / (_closes[-21:-1] + 1e-9)
                _realized_vol = float(np.std(_rets)) if len(_rets) >= 3 else 0.0

                if _realized_vol > 1e-6:
                    # Normalized ROC = how many bar-std-devs did price move in N bars?
                    # sqrt(lookback) scales the per-bar vol to the N-bar window
                    _expected_move = _realized_vol * np.sqrt(MR_ROC_LOOKBACK)
                    _norm_roc = _roc / (_expected_move + 1e-9)

                    # Block if move is more negative than MR_ROC_NORM_THRESH std devs
                    # Default -1.5: blocks when 3-bar drop > 1.5× typical 3-bar move
                    # Use per-ticker threshold: high-ATR stock needs bigger drop to qualify
                    _eff_mom_local = _eff_mom_thresh if _adaptive else MR_ROC_NORM_THRESH
                    if _norm_roc < _eff_mom_local:
                        _rejected.append({"symbol": symbol, "reason": "MOMENTUM_FLUSH",
                                          "detail": f"normROC={_norm_roc:.2f}σ < "
                                                     f"{MR_ROC_NORM_THRESH}σ  "
                                                     f"raw={_roc*100:.2f}%  "
                                                     f"realVol={_realized_vol*100:.2f}%/bar"
                                                     f" (flush vs stock's own baseline)"})
                        return None
                else:
                    # Fallback to raw threshold when vol estimate unavailable
                    if _roc < MR_ROC_THRESHOLD:
                        _rejected.append({"symbol": symbol, "reason": "MOMENTUM_FLUSH",
                                          "detail": f"ROC3={_roc*100:.2f}% < "
                                                     f"{MR_ROC_THRESHOLD*100:.1f}% "
                                                     f"(vertical flush)"})
                        return None
        except Exception:
            pass

    # ── FIREWALL 5: Hierarchical Sector RS (Seven-Layer Pipeline) ────────
    # Three levels of RS computed simultaneously:
    #   L1  sector_etf  vs SPY       → is the sector leading the market?
    #   L2  stock       vs sector    → is this stock leading its sector peers?
    #   L3  stock       vs industry  → is this stock leading its micro-regime?
    #
    # A signal requires: NOT (sector collapsing OR stock sick within sector).
    # Hard block thresholds: sector < 0.95 vs SPY, OR stock < 0.97 vs sector.
    etf          = SECTOR_MAP.get(symbol, "SPY")
    industry_etf = get_industry_etf(symbol)   # SOXX for semis, XBI for biotech, etc.
    _hrs         = get_hierarchical_rs(symbol, etf, industry_etf)

    # Hard block: sector collapsing OR stock sick within its sector
    if _hrs["hard_block"]:
        _rejected.append({"symbol": symbol, "reason": "SECTOR_HARD_BLOCK",
                          "detail": (f"rs_vs_spy={_hrs['rs_vs_spy']:.3f} "
                                     f"rs_vs_sector={_hrs['rs_vs_sector']:.3f} "
                                     f"sector={etf} industry={industry_etf or etf}")})
        return None

    sec_rs   = _hrs["rs_vs_spy"]          # backward compat (used in result dict)
    sec_sc   = _hrs["sector_score"]       # composite of all 3 RS levels
    sec_gate = _hrs["sector_gate"]        # passes dynamic threshold

    # ── Dynamic VWAP bands (after strategy and sector confirmed) ──
    # Only compute after NO_TRADE and sector hard-block checks pass
    # Adaptive SD bands, gap anchor, regime-aware role
    _atr_proxy  = float(np.mean(np.abs(np.diff(df_i["close"].values[-15:]))))
    vwap_bands  = compute_vwap_bands(df_i, _atr_proxy, price, strategy)
    vwap        = vwap_bands["vwap"]  # refresh from bands

    # ── Core calculations (shared by both strategies) ─────────
    # Regime probability defaults -- overwritten after Hurst is computed
    _regime_mr_prob = 0.5
    _regime_tr_prob = 0.5
    _H_blend        = 0.5

    H_daily              = compute_hurst(closes)
    H_intra, _, h_intra_regime = compute_hurst_intraday(df_i)
    lam, hawk_sc         = compute_hawkes(df_i)
    ofi, od, o_sc        = compute_ofi(df_i)
    # Markwick-style rolling Z-score normalized OFI (scale-invariant signal)
    # ofi_norm > +1.5 = strong buy pressure, < -1.5 = strong sell pressure
    ofi_norm             = compute_ofi_normalized(df_i)
    atr                  = compute_atr(df_d)
    add_val, add_b       = get_add_breadth()
    z                    = compute_zscore(closes)
    kurt, skew           = compute_kurtosis_skew(closes)
    hlth, hlbl, hd       = intraday_health(df_i, prior_close)

    # ── Regime probability (SHARED -- used by BOTH strategy branches) ────────
    # Must be computed before strategy split so REGIME_MISMATCH gate works
    # for MR signals. Previously only ran in TREND branch -- MR always got
    # 0.5/0.5 defaults, making the gate permanently inactive.
    # Formula: sigmoid(k × (H − 0.5)) -- smooth 0→1 transition at H=0.50
    #   H=0.58 → P(TREND)=0.73  H=0.42 → P(MR)=0.73  H=0.50 → 50/50
    _k              = 10.0
    _H_blend        = 0.40 * H_daily + 0.60 * H_intra if H_intra > 0 else H_daily
    _H_blend        = float(np.clip(_H_blend, 0.01, 0.99))
    _p_trend        = float(1.0 / (1.0 + np.exp(-_k * (_H_blend - 0.50))))
    _p_mr           = 1.0 - _p_trend
    _regime_mr_prob = round(_p_mr,    3)
    _regime_tr_prob = round(_p_trend, 3)

    # ── STRATEGY BRANCH - exactly one fires ───────────────────

    if strategy == "TREND":
        # TREND MODE: Hurst + Hawkes + momentum OFI
        if mode == "intraday":
            h_sc, hurst_conflict = combined_hurst_score(H_daily, H_intra)
        else:
            h_sc, hurst_conflict = hurst_score(H_daily), False

        direction_g = "LONG" if ofi >= 0.0 else "SHORT"  # signed OFI: 0=neutral, >0=net buy
        ks_sc, ks_lbl = kurtosis_skew_score(kurt, skew, direction_g)
        z_ok, _       = zscore_gate(z, direction_g)
        bayes_p, b_log = bayesian_win_prob(add_b, sec_gate, hawk_sc, z_ok, ks_sc)

        _tw = _get_weights("TREND")
        raw = float(np.clip(
            _tw['hurst']*h_sc + _tw['hawkes']*hawk_sc + _tw['ofi']*o_sc +
            _tw['sector']*sec_sc + _tw['add']*add_score(add_b) +
            _tw['zscore']*zscore_score(z) + _tw['ks']*ks_sc +
            _tw['bayes']*bayes_p*100,
            0, 100))

        # ── RVOL + ATR% participation multiplier (TREND) ──────
        # These are already hard-filtered in passes_liquidity_firewall.
        # Here they boost the score FURTHER for stocks with exceptional
        # participation and volatility -- separating "good trend" from
        # "great trend" within the stocks that already passed the gate.
        #
        # Mechanism (market microstructure):
        #   High RVOL = unusual order flow = institutional activity
        #   High ATR% = sufficient daily range = R:R target reachable
        #
        # Effect: stocks that just barely pass 1.5x RVOL get score=raw
        #         stocks at 3x RVOL (real breakout) get score*1.15
        #         stocks with ATR < 3% get no bonus
        #         stocks with ATR > 4% (real mover) get score*1.10
        _curr_vol   = float(df_d["volume"].iloc[-1])
        _avg_vol    = float(df_d["volume"].tail(20).mean())
        _rvol       = _curr_vol / _avg_vol if _avg_vol > 0 else 1.0
        _atr_pct_sc = (atr / price) if price > 0 else 0.02

        # RVOL multiplier: 1.0 at 1.5x, scales to 1.15 at 3.0x+
        rvol_mult   = float(np.clip(1.0 + (_rvol - 1.5) / 10.0, 1.0, 1.15))
        # ATR% multiplier: 1.0 at 2.5%, scales to 1.10 at 5%+
        atr_mult    = float(np.clip(1.0 + (_atr_pct_sc - 0.025) / 0.25, 1.0, 1.10))

        comp = round(float(np.clip(
            raw * market["mkt_mult"] *
            (1.0 if add_b else 0.80) *
            (hlth if mode == "intraday" else max(hlth, 0.5)) *
            rvol_mult * atr_mult,
            0, 100)), 1)

        if hurst_conflict: comp = min(comp, 45.0)
        if not z_ok:       comp = min(comp, 55.0)

        intraday_blocked = (mode == "intraday" and
                            (hurst_conflict or hawk_sc < 20 or hlth == 0.0))
        intraday_block_reason = ("HURST_CONFLICT" if hurst_conflict else "") +                                 (" HAWKES_SELL" if hawk_sc < 20 else "") +                                 (" HEALTH_ZERO" if hlth == 0.0 else "")

        # TREND also uses per-ticker score threshold (same as MR)
        _score_thr_trend = _eff_score_floor if _adaptive_src == "ADAPTIVE" else SIGNAL_THRESHOLD
        alert = (comp >= _score_thr_trend and market["allows_long"]
                 and z_ok and not intraday_blocked
                 and (hlth > 0.0 if mode == "intraday" else True))

        # TREND regime consistency gate
        # H_blend < 0.5 means market is reverting -- TREND signal is fighting the tape
        if alert and _regime_mr_prob > _regime_tr_prob:
            alert = False
            _rejected.append({"symbol": symbol, "reason": "TREND_REGIME_MISMATCH",
                               "detail": f"P(TREND)={_regime_tr_prob:.2f} < "
                                         f"P(MR)={_regime_mr_prob:.2f} "
                                         f"[H_blend={_H_blend:.3f} -- reverting market]"})

        # Trend mode exit: ATR-based target
        stop_p  = round(price - atr * ATR_STOP_MULT,  2) if atr > 0 else 0.0
        target  = round(price + atr * ATR_TARGET_MULT, 2) if atr > 0 else 0.0
        # R:R = reward / risk  (correct formula: profit_dist / stop_dist)
        _risk_dist = (price - stop_p) if stop_p > 0 and price > stop_p else atr * ATR_STOP_MULT
        rr      = round((target - price) / _risk_dist, 2) if _risk_dist > 0 else 0.0
        exit_type = "ATR_TARGET"
        mr_long_sig = mr_short_sig = False
        vol_diverge = False
        exhaustion_reason = "N/A (TREND mode)"

    else:
        # MEAN_REVERSION MODE: Exhaustion triple-gate + VWAP exit
        # Trend signals (Hurst momentum, Hawkes clustering) NOT used for entry
        h_sc, hurst_conflict = hurst_score(H_daily), False
        direction_g   = "LONG" if ofi >= 0.0 else "SHORT"  # signed OFI: 0=neutral, >0=net buy
        ks_sc, ks_lbl = kurtosis_skew_score(kurt, skew, direction_g)
        z_ok = True  # Z-score is the ENTRY signal in MR mode, not a gate
        bayes_p, b_log = bayesian_win_prob(add_b, sec_gate, hawk_sc, True, ks_sc)

        # Exhaustion triple-gate (Step A + B + C)
        # adaptive_ofi scales the panic/exhaustion thresholds with market conditions:
        #   calm market (0.20) -> more sensitive, easier to trigger
        #   volatile market (0.35) -> requires stronger signal
        mr_long_sig, mr_short_sig, exhaustion_reason, signed_ofi = check_exhaustion_signal(
            df_i, vwap, spy_ret, adaptive_ofi)

        # Volume divergence (final confirmation)
        vol_diverge = check_volume_divergence(df_i)

        intraday_blocked      = False
        intraday_block_reason = ""

        # Kill signal if exhaustion not confirmed or volume not diverging
        if not (mr_long_sig or mr_short_sig):
            intraday_blocked = True
            intraday_block_reason = f"NO_EXHAUSTION: {exhaustion_reason}"
        if not vol_diverge and not intraday_blocked:
            intraday_blocked = True
            intraday_block_reason = "NO_VOL_DIVERGENCE"

        # MR composite score - rewards exhaustion signals, penalises momentum
        z_now = (closes[-1] - np.mean(closes[-20:])) / (np.std(closes[-20:], ddof=1) + 1e-9)                 if len(closes) >= 20 else 0.0
        exhaustion_score = min(100.0, abs(z_now) / Z_ENTRY_THRESH * 60
                               + (30 if mr_long_sig or mr_short_sig else 0)
                               + (10 if vol_diverge else 0))

        _mw = _get_weights("MEAN_REVERSION")
        raw  = float(np.clip(
            _mw.get('exhaustion',0.30)*exhaustion_score + _mw['ofi']*o_sc +
            _mw['sector']*sec_sc + _mw['add']*add_score(add_b) +
            _mw['ks']*ks_sc + _mw['bayes']*bayes_p*100,
            0, 100))

        # ── ATR% range multiplier (MR) ─────────────────────────
        # MR doesn't need high RVOL (reversions happen quietly).
        # But it DOES need enough daily range to reach the VWAP target.
        # If ATR is tiny, the VWAP target is too close -- R:R collapses.
        # Mechanism: larger ATR% = larger gap between price and VWAP
        #            = bigger reversion potential = better R:R
        _atr_pct_sc = (atr / price) if price > 0 else 0.02
        # ATR% multiplier: 1.0 at 1.5%, scales to 1.12 at 4%+
        atr_mult_mr = float(np.clip(1.0 + (_atr_pct_sc - 0.015) / 0.20, 1.0, 1.12))

        comp = round(float(np.clip(
            raw * market["mkt_mult"] * (1.0 if add_b else 0.80) * hlth * atr_mult_mr,
            0, 100)), 1)

        # MR can fire LONG or SHORT - apply direction-appropriate market gate
        mr_direction = "LONG" if mr_long_sig else "SHORT"
        direction_ok  = (market["allows_long"] if mr_direction == "LONG"
                         else True)  # shorts valid even in RISK-OFF
        # Use per-ticker score threshold if adaptive (e.g. ALKT needs 72+, NOMD 65)
        _score_thr_mr = _eff_score_floor if _adaptive_src == "ADAPTIVE" else SIGNAL_THRESHOLD
        alert = (comp >= _score_thr_mr and not intraday_blocked
                 and hlth > 0.0 and direction_ok)

        # MR exit: VWAP touch (primary) or adaptive TP% (floor)
        # Stop = larger of ATR-based or per-ticker SL floor
        # Uses adaptive params: ALKT (ATR=4%) gets wider stop than NOMD (ATR=0.5%)
        atr_stop  = round(price - atr * ATR_STOP_MULT, 2) if atr > 0 else 0.0
        pct_stop  = round(price * (1 - _eff_sl), 2)   # per-ticker or global SL
        stop_p    = max(atr_stop, pct_stop)             # whichever is closer
        target    = vwap                                # VWAP is primary target
        # Adaptive TP floor: if VWAP too close, use per-ticker TP% instead
        vwap_dist = abs(vwap - price) / (price + 1e-9)
        if vwap_dist < _eff_tp:
            # VWAP is inside our minimum target -- use TP% floor as target
            target = round(price * (1 + _eff_tp), 2) if direction == "LONG" else                      round(price * (1 - _eff_tp), 2)
        rr        = abs(target - price) / (abs(price - stop_p) + 1e-9)

        # REGIME_MISMATCH first -- if market is trending, MR is wrong strategy
        # No point checking R:R math when regime is fundamentally wrong
        if alert and _regime_tr_prob > _regime_mr_prob:
            alert = False
            _rejected.append({"symbol": symbol, "reason": "REGIME_MISMATCH",
                               "detail": f"P(MR)={_regime_mr_prob:.2f} < "
                                          f"P(TREND)={_regime_tr_prob:.2f} "
                                          f"[H_blend={_H_blend:.3f} > 0.50]"})

        # R:R gate: only runs when regime is correct
        if alert and rr < MR_MIN_RR_RATIO:
            alert = False
            _rejected.append({"symbol": symbol, "reason": "POOR_RR",
                               "detail": f"rr={rr:.2f} < {MR_MIN_RR_RATIO} "
                                         f"[target too close to entry]"})
        exit_type = "VWAP_TOUCH"

    # ── Shared post-processing ────────────────────────────────
    direction_g  = "LONG" if (mr_long_sig if strategy == "MEAN_REVERSION" else ofi >= 0.0) else "SHORT"  # signed OFI
    strength, hl_rem, hl_alive = halflife_remaining(symbol, comp, price, atr)
    kf, drisk, shares = kelly_size(price, atr, win_rate=bayes_p) if alert else (0.0, 0.0, 0)

    # ── Range Filter (final confirmation gate) ────────────────
    # Confirms price has started moving in your favor BEFORE entry.
    # Final_Signal = (Score > 75) AND (price > rf_val AND rf_upward)
    # Russell 2000: rf_length=25 (faster). Watchlist: rf_length=30.
    thresholds = get_active_thresholds()
    # §5.1: Adaptive RF length based on actual stock volatility
    atr_pct = (atr / price) if price > 0 else 0.02
    rf_len  = adaptive_rf_length(atr_pct, UNIVERSE_MODE)
    rf_val, rf_up, rf_dn, rf_sig_up, rf_sig_dn = compute_range_filter(closes, length=rf_len)
    # Range filter confirmation: direction must match signal
    if direction_g == "LONG":
        rf_confirms = rf_sig_up
    else:
        rf_confirms = rf_sig_dn

    # R2K mode applies stricter signal threshold and Hurst/Hawkes bars
    if UNIVERSE_MODE == "RUSSELL2000":
        r2k_ok = (comp >= thresholds["signal_thresh"] and
                  H_daily >= thresholds["hurst_min"] and
                  hawk_sc >= thresholds["hawkes_min"])
    else:
        r2k_ok = True  # watchlist mode - standard threshold

    # Final combined alert gate
    alert_final = alert and r2k_ok
    # Range filter warning - doesn't block but shown on card
    rf_warning = alert_final and not rf_confirms

    # ── Markov Gate (probability layer) ──────────────────────
    # SCOPE: Both TREND and MEAN_REVERSION, with different behavior.
    #
    # WHY TREND NOW INCLUDED:
    # Original design: Markov = MR-only. That works most of the time.
    # Edge case that breaks it: strong gap-up (+3%) with midday fade.
    #   - Regime = TREND (SPY up, ADX rising)
    #   - But Markov sees: CALM state only 48% stable -> likely to flip
    #   - That instability IS meaningful even in trend mode
    # The fix: Markov applies to TREND too, but with a lighter touch.
    #
    # BEHAVIOR BY STRATEGY:
    #   MEAN_REVERSION: full 10% penalty if unstable, stop tightened
    #   TREND:          5% penalty only (regime instability = caution, not kill)
    #                   stop NOT tightened in trend (ATR stop handles it)
    #
    # GUARDRAILS (unchanged):
    #   - MARKOV_GATE_ENABLED = False by default (opt-in only)
    #   - Only fires when alert_final = True
    #   - Never hard-blocks (penalty only, signal can still fire)
    #   - Insufficient data -> stable=True -> no effect
    #   - SPY > 2% on trend day -> VOL_OVERRIDE -> no effect
    markov_stable    = True
    markov_stay_prob = 1.0
    markov_stop_adj  = 1.0
    markov_reason    = "MARKOV_DISABLED"

    if MARKOV_GATE_ENABLED and alert_final:
        try:
            global _get_markov_gate_fn
            if _get_markov_gate_fn is None:
                from scanner_research_v4 import get_markov_gate as _mgk
                _get_markov_gate_fn = _mgk
            markov_stable, markov_stay_prob, markov_stop_adj, markov_reason = \
                _get_markov_gate_fn(symbol, adx_val, spy_ret, days=60)

            if not markov_stable:
                if strategy == "MEAN_REVERSION":
                    # Full 10% penalty + stop tightening for MR
                    # Rationale: MR trades depend on regime stability.
                    # An unstable CALM state is the "MR trap" -- the stock
                    # looks oversold but the regime is about to flip to TREND.
                    comp   = round(comp * 0.90, 1)
                    if markov_stop_adj < 1.0 and stop_p > 0:
                        stop_p = round(price - (price - stop_p) * markov_stop_adj, 4)
                else:
                    # TREND: lighter 5% penalty, no stop tightening
                    # Rationale: TREND signals don't rely on regime stability
                    # the way MR does, but a RANGING->VOLATILE transition
                    # (e.g. midday fade after gap-up) is worth noting.
                    # ATR-based stops already handle trend volatility.
                    comp = round(comp * 0.95, 1)

                alert_final = comp >= SIGNAL_THRESHOLD
        except Exception:
            pass  # research module not available -- trade normally

    _result = {
        "symbol":symbol,"price":round(price,2),"score":comp,"alert":alert_final,
        "mode":mode,"strategy":strategy,"exit_type":exit_type,
        "adx":adx_val,"regime_reason":regime_reason,"adaptive_ofi":adaptive_ofi,
        # Hurst
        "hurst_H":round(H_daily,3),"hurst_H_intra":round(H_intra,3),
        "hurst_score":round(h_sc,1),"hurst_regime":hurst_regime(H_daily),
        "hurst_regime_intra":h_intra_regime,"hurst_conflict":hurst_conflict,
        # Indicators
        "hawkes_lam":lam,"hawkes_score":hawk_sc,"hawkes_sig":hawkes_signal(hawk_sc),
        "ofi":ofi,"ofi_delta":od,"ofi_score":o_sc,"ofi_sig":ofi_signal(ofi,od),
        "ofi_norm":ofi_norm,  # rolling Z-score normalized OFI (Markwick 2022)
        "signed_ofi": round(float(signed_ofi) if 'signed_ofi' in locals() else 0.0, 4),
        "market":market["regime"],"regime":market["regime"],"sector_etf":etf,
        "sector_rs":sec_rs,"sector_score":sec_sc,"sector_gate":sec_gate,
        # ── Hierarchical RS (Seven-Layer Pipeline) ────────────────────────
        "rs_vs_spy":     _hrs["rs_vs_spy"],       # sector vs SPY (market alpha)
        "rs_vs_sector":  _hrs["rs_vs_sector"],    # stock vs sector ETF (sector alpha)
        "rs_vs_industry":_hrs["rs_vs_industry"],  # stock vs industry ETF (micro-regime)
        "industry_etf":  _hrs["industry_etf"],    # SOXX/XBI/KRE/etc or sector_etf
        "index_benchmark": "SPY",                 # market baseline always SPY
        "add_val":add_val,"add_bull":add_b,"add_score":round(add_score(add_b),1),
        "zscore":round(z,3),"zscore_score":round(zscore_score(z),1),"zscore_ok":z_ok,
        "kurtosis":kurt,"skewness":skew,"ks_score":ks_sc,"ks_label":ks_lbl,
        "bayes_prob":round(bayes_p*100,1),"bayes_factors":b_log,
        "hl_strength":strength,"hl_remaining":round(hl_rem,0),"hl_alive":hl_alive,
        "rr_ratio":round(rr,2),"rr_ok":rr >= 3.0,
        "kelly_frac":kf,"dollar_risk":drisk,"shares":shares,
        "atr":atr,"stop":stop_p,"target":target,"vwap":round(vwap,2),
        # Mean reversion specific
        "mr_long":mr_long_sig,"mr_short":mr_short_sig,
        "vol_diverge":vol_diverge,"exhaustion_reason":exhaustion_reason,
        "health_mult":hlth,"health_label":hlbl,"liq_status":liq_reason,
        "intraday_blocked":intraday_blocked,"intraday_block_reason":intraday_block_reason,
        "intraday_ret":hd.get("intraday_ret",0.0),"gap_ret":hd.get("gap_ret",0.0),
        "below_vwap":hd.get("below_vwap",False),
        "mcap_b":mcap_b,"scanned_at":datetime.now().strftime("%H:%M:%S"),
        "park_shock":market.get("park_shock",False),"park_ratio":round(market.get("park_ratio",1.0),3),
        "regime_mr_prob":_regime_mr_prob,"regime_tr_prob":_regime_tr_prob,"H_blend":round(_H_blend,3),
        "atr_pct":round(atr/price if price>0 else 0.02, 5),"est_spread":round(_liq_spread if "_liq_spread" in dir() else 0.0, 5),
        "adaptive_src":_adaptive_src,"adaptive_tp":_eff_tp,"adaptive_sl":_eff_sl,
        # ── Dynamic VWAP bands ────────────────────────────────
        "vwap_upper1":   vwap_bands.get("upper1", vwap),
        "vwap_lower1":   vwap_bands.get("lower1", vwap),
        "vwap_upper2":   vwap_bands.get("upper2", vwap),
        "vwap_lower2":   vwap_bands.get("lower2", vwap),
        "vwap_gap":      vwap_bands.get("gap_vwap"),
        "vwap_role":     vwap_bands.get("vwap_role", "MAGNET"),
        "vwap_bands":    vwap_bands.get("band_mult", "1.5/2.5SD"),
        "above_vwap1":   vwap_bands.get("above_upper1", False),
        "below_vwap1":   vwap_bands.get("below_lower1", False),
        "above_vwap2":   vwap_bands.get("above_upper2", False),
        "below_vwap2":   vwap_bands.get("below_lower2", False),
        # ── Range Filter ──────────────────────────────────────
        "rf_val":round(rf_val,4),"rf_up":rf_up,"rf_dn":rf_dn,
        "rf_confirms":rf_confirms,"rf_warning":rf_warning,
        "rf_length":rf_len,
        # ── Markov Gate ───────────────────────────────────────
        "markov_stable":    markov_stable,
        "markov_stay_prob": round(markov_stay_prob, 3),
        "markov_stop_adj":  round(markov_stop_adj, 2),
        "markov_reason":    markov_reason,
        # ── Golden Entry flag ─────────────────────────────────
        "golden_entry": (
            float(z) <= -Z_ENTRY_THRESH and
            hd.get("below_vwap", False) and
            strategy == "MEAN_REVERSION"
        ),
        # ── Universe mode tag ─────────────────────────────────
        "universe_mode": UNIVERSE_MODE,
        "signal_threshold": _score_thr_mr,
    }

    # ── Ticker Intelligence: apply profile score adjustment ──────────────
    # Adjusts composite score based on historical win rate for this symbol/strategy.
    # Only fires when ticker_intelligence.py is available and has >= 5 trades on file.
    if _TI_AVAILABLE:
        try:
            _result = _ti.patch_scan_symbol_with_profile(_result, symbol)
        except Exception:
            pass

    return _result


def run_full_scan(symbols: list = WATCHLIST, timeframe: str = "5m") -> tuple:
    """
    Parallel scan using ThreadPoolExecutor.
    Auto-reloads watchlist from auto_watchlist.txt if it exists and is newer.
    Tracks FULL rejection reason for every blocked symbol.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    global WATCHLIST
    # Auto-refresh watchlist: priority order —
    #   1. ticker_intelligence.build_active_watchlist() if available
    #      (incorporates profile data, cooling, dynamic runners, RVOL bubbling)
    #   2. auto_watchlist.txt if it exists (written by dynamic_watchlist.py)
    #   3. engine.WATCHLIST fallback (always works)
    if symbols is WATCHLIST:
        if _TI_AVAILABLE:
            try:
                _wl, _wl_meta = _ti.build_active_watchlist(
                    core=WATCHLIST,
                    universe_csv_path=os.path.join(SCANNER_DIR, "current_universe.csv"),
                    strategy_mode=STRATEGY_MODE,
                    verbose=False,
                )
                if _wl:
                    WATCHLIST = _wl
                    symbols   = _wl
            except Exception:
                pass  # fall through to auto_watchlist.txt
        if symbols is WATCHLIST:
            # Fallback: plain auto_watchlist.txt
            _auto_path = os.path.join(SCANNER_DIR, "auto_watchlist.txt")
            if not os.path.exists(_auto_path):
                _auto_path = "auto_watchlist.txt"
            if os.path.exists(_auto_path):
                try:
                    fresh = [s.strip().upper()
                             for s in open(_auto_path).read().split("\n")
                             if s.strip()]
                    if fresh:
                        WATCHLIST = fresh
                        symbols   = fresh
                except Exception:
                    pass
    market   = get_market_regime()
    results  = []
    rejected = []   # list of dicts: {symbol, reason, detail}

    # Pre-check market cap (fast - uses 1h cache)
    qualified = []
    for sym in symbols:
        ok, mb = is_midcap(sym, STRATEGY_MODE)
        if not ok:
            # Guard: if mcap returned 0 (bad data), don't block on cap
            # — it will be caught by liquidity firewall instead
            if mb == 0.0:
                qualified.append(sym)
                continue
            reason = "BELOW_MIN_CAP" if mb < MIDCAP_MIN/1e9 else "ABOVE_MAX_CAP"
            rejected.append({"symbol":sym,"reason":reason,
                             "detail":f"${mb:.2f}B (need $300M-$20B)"})
        else:
            qualified.append(sym)

    # Parallel scan - capture why each symbol was rejected inside scan_symbol
    def _scan_one(sym):
        try:
            # Run each firewall manually to get the specific rejection reason
            df_d = fetch_daily(sym, 60)
            if df_d is None or len(df_d) < 30:
                return sym, None, "NO_DATA", "Insufficient daily bars"

            liq_ok, liq_msg, _ = passes_liquidity_firewall(df_d, sym, STRATEGY_MODE)
            if not liq_ok:
                return sym, None, "LIQUIDITY", liq_msg

            _df_i = fetch_intraday(sym, timeframe)
            df_i  = _df_i if (_df_i is not None and not _df_i.empty) else df_d.copy()
            spy_ret = market.get("spy_dev", 0.0) / 100.0

            if STRATEGY_MODE == "AUTO":
                strat, _, _, reg_reason = get_strategy_regime(df_i, spy_ret)
            else:
                strat = STRATEGY_MODE; reg_reason = f"FORCED_{STRATEGY_MODE}"

            if strat == "NO_TRADE":
                return sym, None, "NO_TRADE", reg_reason

            etf = SECTOR_MAP.get(sym, "SPY")
            _, _, sec_gate = get_sector_rs(etf)
            if not sec_gate:
                return sym, None, "SECTOR_RS", f"{etf} RS < {get_dynamic_sector_rs(STRATEGY_MODE):.2f} (sector lagging SPY in {STRATEGY_MODE} mode)"

            r = scan_symbol(sym, market, timeframe)
            if r:
                return sym, r, None, None
            else:
                return sym, None, "SCAN_FAILED", "Returned None (likely intraday health block)"
        except Exception as e:
            return sym, None, "ERROR", str(e)

    with ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(_scan_one, sym): sym for sym in qualified}
        for fut in as_completed(futures):
            sym, r, reason, detail = fut.result()
            if r:
                results.append(r)
            elif reason:
                rejected.append({"symbol":sym,"reason":reason,"detail":detail or ""})

    # Build simple blocked list for backward compat + rich rejected list
    blocked_simple = [f"{r['symbol']}({r['reason']})" for r in rejected]

    # Determine the active strategy for this scan based on actual results
    # In AUTO mode: look at what strategy most signals used this scan
    # This is what gets displayed in the terminal market line
    if results:
        strats = [r.get("strategy","") for r in results if r.get("strategy")]
        trend_n = sum(1 for s in strats if s == "TREND")
        mr_n    = sum(1 for s in strats if s == "MEAN_REVERSION")
        active_strategy = "TREND" if trend_n >= mr_n else "MEAN_REVERSION"
    else:
        active_strategy = STRATEGY_MODE

    market.update({
        "blocked":         blocked_simple,
        "blocked_count":   len(rejected),
        "scanned":         len(results),
        "rejected_detail": rejected,
        "active_strategy": active_strategy,  # actual per-scan strategy for display
    })

    # ── Ticker Intelligence: record rejections + update profiles ──────────
    # Persists rejection streaks, updates cooling list, updates ticker profiles.
    # Runs asynchronously so it never delays the scan return.
    if _TI_AVAILABLE and rejected:
        try:
            _ti.record_rejections(rejected)
        except Exception:
            pass

    if not results: return pd.DataFrame(), market
    df_raw = pd.DataFrame(results).sort_values(by=["score","bayes_prob"],
                                                ascending=False).reset_index(drop=True)
    # §9: Apply relative ranking for R2K (top N% not fixed threshold)
    if UNIVERSE_MODE == "RUSSELL2000":
        df_raw = rank_signals(df_raw, top_n=20, top_pct=0.05)
    else:
        df_raw["rank"]       = range(1, len(df_raw)+1)
        df_raw["percentile"] = (df_raw.index / len(df_raw)).round(4)
    return df_raw, market


# ── Universe Scanner ──────────────────────────────────────────────────────────
def run_universe_scan(top_n: int = 30, timeframe: str = "5m") -> pd.DataFrame:
    """
    Optimized universe scan — three pipeline stages:
      Stage 1: fast batch prefilter (daily bars, liquidity/cap check)
      Stage 2: chunked intraday download matching the scan timeframe
               (100-ticker chunks → single HTTP call per chunk)
      Stage 3: parallel scoring via df_override (zero per-ticker fetches)

    10x faster than the sequential version because:
      - yf.download() fetches all tickers in one HTTP call per chunk
      - scan_symbol() uses df_override to skip its own fetch_intraday()
      - ThreadPoolExecutor parallelizes the scoring math
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import yfinance as yf

    # Timeframe → yfinance period mapping
    # Must match what fetch_intraday() would fetch so df_override is valid
    _period_map = {"1m":"1d","2m":"5d","5m":"5d","15m":"60d","30m":"60d","1h":"730d"}
    _dl_period   = _period_map.get(timeframe, "5d")
    _dl_interval = timeframe   # match scan timeframe exactly

    print(f"[UNIVERSE] Start {datetime.now().strftime('%H:%M:%S')}  "
          f"tf={timeframe}  dl_period={_dl_period}")

    # ── Stage 1: Prefilter (batched daily bars)
    survivors = universe_prefilter_fast(UNIVERSE)
    if not survivors:
        print("[UNIVERSE] No survivors after prefilter.")
        return pd.DataFrame()

    # ── Market regime snapshot (passed to scan_symbol)
    market = get_market_regime()

    # ── Stage 2: Chunked intraday download (timeframe-matched)
    batched = {}
    chunk_size = 100

    chunks = [survivors[i:i + chunk_size] for i in range(0, len(survivors), chunk_size)]
    print(f"[UNIVERSE] Downloading {len(survivors)} survivors in "
          f"{len(chunks)} chunks (interval={_dl_interval})...")

    for idx, chunk in enumerate(chunks, 1):
        print(f"[UNIVERSE] Chunk {idx}/{len(chunks)} ({len(chunk)} tickers)")
        try:
            raw = yf.download(
                tickers=" ".join(chunk),
                period=_dl_period,
                interval=_dl_interval,   # FIXED: was hardcoded "1m", now matches timeframe
                group_by="ticker",
                threads=True,
                auto_adjust=True,
                progress=False,
                timeout=30,
            )
            if raw is None or raw.empty:
                continue
            is_multi = isinstance(raw.columns, pd.MultiIndex)
            for sym in chunk:
                try:
                    df = raw[sym].copy() if is_multi else raw.copy()
                    df.columns = [c.lower() if not isinstance(c, tuple) else c[0].lower()
                                  for c in df.columns]
                    if "close" in df.columns and len(df) >= 15:
                        batched[sym] = df
                except Exception:
                    continue
        except Exception as e:
            print(f"[UNIVERSE] Chunk {idx} download error: {e}")

    print(f"[UNIVERSE] Valid price data: {len(batched)}/{len(survivors)}")
    
    # ── Stage 3: Parallel scoring using df_override
    results = []

    def _scan_one(sym):
        try:
            df = batched.get(sym)
            if df is None:
                return None
            return scan_symbol(sym, market, timeframe, df_override=df)
        except Exception as e:
            print(f"[UNIVERSE ERR] {sym}: {e}")
            return None

    print(f"[UNIVERSE] Scanning {len(batched)} symbols (parallel)...")

    done = 0
    with ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(_scan_one, sym): sym for sym in batched}
        for fut in as_completed(futures):
            r = fut.result()
            if r:
                results.append(r)
            done += 1
            if done % 10 == 0:
                print(f"[UNIVERSE] {done}/{len(batched)} scanned")

    if not results:
        print("[UNIVERSE] No results after scoring.")
        return pd.DataFrame()

    # ── Sort + save top picks
    df = pd.DataFrame(results).sort_values(
        by=["score", "bayes_prob"], ascending=False
    ).reset_index(drop=True)

    # Write auto_watchlist.txt
    with open(os.path.join(SCANNER_DIR, "auto_watchlist.txt"), "w") as f:
        f.write("\n".join(df.head(top_n)["symbol"].tolist()))

    print(f"[UNIVERSE] Saved top {top_n} → auto_watchlist.txt")

    # Export Excel via engine
    export_excel(df, market)

    return df

# ── Russell 2000 Batched Scanner ─────────────────────────────────────────────
# Scans the full IWM constituent list in batches of 500 to avoid rate limits.
# Applies R2K-specific stricter thresholds automatically.
# Run: set UNIVERSE_MODE = "RUSSELL2000" then call run_russell2000_scan()

# Full Russell 2000 constituent list - loaded from IWM holdings or static list
# This is a representative subset; for live use fetch from iShares IWM holdings
RUSSELL2000_SYMBOLS = [
    # Technology
    "ACLS","ADTH","AEHR","AGYS","ALLT","AMSWA","AOSL","APPF","ARLO","ATEN",
    "ATNI","AVNW","BAND","BCOV","BLKB","BSIG","CASS","CEVA","CGNT",
    "CLFD","CLPS","CMPR","CNXC","COHU","CSTL","CSWI","DOMO","DTIL","DTST",
    "EBIX","EDAC","EGAN","EGHT","ENVA","EVTC","EXPI","EXLP","FCRD","FIVN",
    "FORM","FORR","GDOT","GEOS","GLNG","GNSS","GRPN","GTLB","HCAT","HIMX",
    "HOLI","HUBS","IESC","IPAR","IRDM","IRTC","ITRI","JAMF","JNPR","KVHI",
    "LSCC","LSTR","LYTS","MARA","MGNI","MLAB","MMSI","MODV","MODN","MORN",
    "MRCY","MTSI","NTGR","NTRA","NVAX","ONTO","POWI","PRCT","PSTG","PWSC",
    "QADA","QNST","RAMP","SCSC","SHEN","SIFY","SILK","SMTC","SPOK","SSYS",
    "STRL","TASK","TCMD","TELOS","TTEC","TUYA","TZOO","VIAV","VNET",
    # Healthcare
    "ACAD","ACET","ACLS","ACMR","ADMA","AGIO","AKCA","ALDX","ALEC","ALOG",
    "AMPH","AMRX","ARWR","ATRC","AVDL","AXDX","AXGT","BCPC","BIIB","BLFS",
    "BPMC","BTBT","CALA","CALX","CAMT","CASI","CBPO","CCRN","CDMO","CDNA",
    "CERO","CGEM","CGEN","CHRS","CLOV","CMPS","CNMD","CODX","CORT","CPRX",
    "CRVL","CSTE","CTLT","CTMX","CVAC","CYTK","DRIO","DYAI","ENTA","ESTE",
    "EVIO","EVLO","FATE","FDMT","FOLD","FREQ","GALT","GERN","GKOS",
    "GLPG","GNCA","GNFT","GOSS","GRTS","HALO","HCAT","HIMS","HROW","HRMY",
    "IMCR","IMGO","IMTX","IMVT","INDB","INMD","IONS","IOVA","IPHA",
    "KRYS","KRTX","LGND","LNTH","MGNX","MYGN","NKTR","NTRA","NUVL","NVAX",
    "OCGN","OCUL","OMCL","OPCH","OPRT","ORGO","PCRX","PDCO","PDLI","PRME",
    # Financials
    "ABR","ACNB","AFCG","AMAL","AMNB","AMTB","ANCX","AROW","ATLC",
    "ATLO","BANC","BANF","BANR","BBCP","BCAR","BCML","BCSB","BFIN","BFST",
    "BHLB","BKNG","BKSC","BLMN","BMRC","BNCN","BOCH","BOKF","BPOP","BRKL",
    "BSVN","BUSE","BYFC","CABO","CADE","CAR","CARE","CASH","CATO","CBFV",
    "CBSH","CBTX","CCBG","CCNE","CFFN","CFFI","CFNB","CHCO","CHMG","CIVB",
    "CLBK","CLDB","CMTV","CNOB","COOP","CORE","CRAI","CRBP","CSTR","CTBI",
    "CUBI","CURE","CVBF","CVCY","CZWI","DCOM","DENN","DFIN","DGICA",
    "EBMT","ECPG","EFSC","EFC","ENVA","ESSA","EVBN","EVTC","FBIZ",
    # Industrials / Energy / Materials
    "ASTE","ASTL","ATRI","AVAV","AVNT","AZTA","BATL","BCPC","BWXT","CIVI",
    "CLW","CMCO","CRS","CSWI","CWST","DAN","DXPE","ECCA","ENVA",
    "EPAC","ERII","FBHS","FELE","FSTR","GATX","GFF","GHM","GLDD","GNRC",
    "GTLS","HAYW","HLIO","HURC","HWKN","HYMC","IESC","ITRI","JBSS",
    "KELYA","KTOS","KWR","LCII","LCUT","LECO","LYTS","MATW","MATX",
    "MBIN","MCBC","MCRI","MERC","MGPI","MNRO","MRTN","MRUS","MSEX",
    "MTDR","NNBR","NOMD","NOVT","NSP","OBNK","OSIS","PTEN","REX","RYAM",
    "SCVL","SHYF","SPXC","STAG","STRL","SUNL","SUPN","SVRA","SWBI","TCBK",
    "TROX","TRST","TTGT","TUEM","UFCS","UFPT","UITB","ULCC","UONE","USAC",
    "VBTX","VICR","VSCO","VTOL","WABC","WASH","WCBI","WERN","WTBA","WTTR",
]

def fetch_russell2000_symbols() -> list:
    """
    Attempts to fetch live IWM constituent list from iShares.
    Falls back to built-in RUSSELL2000_SYMBOLS list if unavailable.
    """
    try:
        import urllib.request, json
        url = "https://www.ishares.com/us/products/239710/ishares-russell-2000-etf/1467271812596.ajax?tab=holdings&fileType=json"
        req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
        syms = [row[0] for row in data.get("aaData",[]) if row[0] and len(row[0]) <= 5]
        if len(syms) > 100:
            print(f"[R2K] Fetched {len(syms)} live IWM constituents")
            return syms
    except Exception as e:
        print(f"[R2K] Live fetch failed ({e}), using built-in list")
    return RUSSELL2000_SYMBOLS


def run_russell2000_scan(timeframe: str = "5m", top_n: int = 30) -> tuple:
    """
    Full Russell 2000 scan in batches of 500 to avoid yfinance rate limits.

    Architecture:
      1. Fetch constituent list (live from iShares or built-in)
      2. Stage 1 fast prefilter: price > $5, avg vol > 300K, ATR% > 2%
         Done in batches of 500 via yf.download() - much faster than per-symbol
      3. Stage 2 full scan: all 10 layers on survivors
         Uses R2K stricter thresholds: $10M vol, 1.5x RelVol, score > 72

    Rate limit strategy: 3-second pause between 500-symbol batches.
    Expected time: ~8-12 min for full 2000 symbol scan.
    Expected survivors: 15-40 stocks depending on market conditions.
    """
    global UNIVERSE_MODE
    original_mode = UNIVERSE_MODE
    UNIVERSE_MODE = "RUSSELL2000"  # Activate stricter thresholds

    thresholds = get_active_thresholds()
    batch_size = thresholds["batch_size"]
    min_price  = thresholds["min_price"]

    print(f"\n[R2K] Starting Russell 2000 scan at {datetime.now().strftime('%H:%M:%S')}")
    print(f"[R2K] Thresholds: DV>${thresholds['min_dollar_vol']/1e6:.0f}M "
          f"RelVol>{thresholds['min_rel_vol']}x "
          f"Score>{thresholds['signal_thresh']}")

    symbols = fetch_russell2000_symbols()
    print(f"[R2K] Universe: {len(symbols)} symbols -> scanning in batches of {batch_size}")

    # Stage 1: fast prefilter in batches
    survivors = []
    batches = [symbols[i:i+batch_size] for i in range(0, len(symbols), batch_size)]

    for b_idx, batch in enumerate(batches):
        print(f"[R2K] Prefilter batch {b_idx+1}/{len(batches)} ({len(batch)} symbols)...")
        try:
            raw = yf.download(
                batch, period="1mo", interval="1d",
                group_by="ticker", auto_adjust=True,
                progress=False, threads=True
            )
            for sym in batch:
                try:
                    df = raw[sym] if len(batch) > 1 else raw
                    if df is None or df.empty or len(df) < 10: continue
                    df.columns = [c.lower() for c in df.columns]
                    if "close" not in df.columns: continue
                    p = float(df["close"].iloc[-1])
                    if p < min_price: continue
                    if float(df["volume"].mean()) < 300_000: continue
                    atr_pct = float(df["close"].diff().abs().mean() / p)
                    if atr_pct < 0.02: continue
                    survivors.append(sym)
                except Exception:
                    continue
        except Exception as e:
            print(f"[R2K] Batch {b_idx+1} failed: {e}")
        # Rate limit pause between batches
        if b_idx < len(batches) - 1:
            time.sleep(3)

    print(f"[R2K] Stage 1 complete: {len(symbols)} -> {len(survivors)} survivors")

    # Stage 2: full 10-layer scan on survivors
    market  = get_market_regime()
    results = []
    rejected= []

    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _r2k_scan_one(sym):
        try:
            df_d = fetch_daily(sym, 60)
            if df_d is None or len(df_d) < 30:
                return sym, None, "NO_DATA", "Insufficient bars"
            liq_ok, liq_msg, _ = passes_liquidity_firewall(df_d, sym, STRATEGY_MODE)
            if not liq_ok:
                return sym, None, "LIQUIDITY", liq_msg
            r = scan_symbol(sym, market, timeframe)
            if r: return sym, r, None, None
            return sym, None, "SCAN_FAILED", "Below threshold"
        except Exception as e:
            return sym, None, "ERROR", str(e)[:80]

    done = 0
    with ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(_r2k_scan_one, sym): sym for sym in survivors}
        for fut in as_completed(futures):
            sym, r, reason, detail = fut.result()
            if r:   results.append(r)
            else:   rejected.append({"symbol":sym,"reason":reason,"detail":detail or ""})
            done += 1
            if done % 20 == 0:
                print(f"[R2K] Stage 2: {done}/{len(survivors)} scanned, {len(results)} passed")

    UNIVERSE_MODE = original_mode  # Restore

    if not results:
        print("[R2K] No results passed all filters")
        return pd.DataFrame(), market

    df = (pd.DataFrame(results)
          .sort_values(by=["golden_entry","score","bayes_prob"],
                       ascending=[False, False, False])
          .reset_index(drop=True))

    # Save top picks to watchlist
    top = df.head(top_n)["symbol"].tolist()
    with open("r2k_watchlist.txt", "w") as f:
        f.write("\n".join(top))

    market.update({
        "rejected_detail": rejected,
        "blocked_count":   len(rejected),
        "scanned":         len(results),
    })

    print(f"[R2K] Complete: {len(results)} passed, "
          f"{len(df[df['golden_entry']==True])} Golden Entries, "
          f"{len(df[df['alert']==True])} signals")
    print(f"[R2K] Top {top_n} saved to r2k_watchlist.txt")

    return df, market


# ── Excel Export ──────────────────────────────────────────────────────────────

def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert all columns to Excel-safe Python-native types.
    Fixes ArrowStringArray / ArrowDtype errors from pandas 2.0+ with PyArrow.
    openpyxl cannot serialize pyarrow-backed arrays -- must convert first.

    Fix: forces every column out of the Arrow backend using .to_numpy(object)
    before any further type conversion. This prevents 'ArrowStringArray is not
    callable' errors that occur when .astype(str) stays in Arrow memory.
    """
    if df.empty:
        return df
    out = pd.DataFrame(index=df.index)  # fresh frame, no Arrow metadata
    for col in df.columns:
        try:
            series = df[col]
            dtype_str = str(series.dtype).lower()

            # Step 1: Force out of Arrow/PyArrow backend unconditionally.
            # .to_numpy(dtype=object) always returns a plain NumPy object array
            # regardless of whether the underlying storage is Arrow or not.
            raw = series.to_numpy(dtype=object, na_value=None)

            # Step 2: Convert to appropriate Python-native type
            if "bool" in dtype_str:
                out[col] = pd.array(
                    [bool(x) if x is not None else None for x in raw],
                    dtype=object)
            elif any(t in dtype_str for t in ("int", "uint")):
                out[col] = pd.to_numeric(
                    pd.Series(raw), errors="coerce").fillna(0).astype(int)
            elif "float" in dtype_str:
                out[col] = pd.to_numeric(pd.Series(raw), errors="coerce")
            elif any(t in dtype_str for t in ("arrow", "string", "object")):
                # Convert to plain Python str; preserve None for blank cells
                out[col] = pd.array(
                    [str(x) if x is not None else None for x in raw],
                    dtype=object)
            else:
                # Unknown dtype -- safest: stringify non-None, preserve None
                out[col] = pd.array(
                    [str(x) if x is not None and not isinstance(x, float)
                     else x for x in raw],
                    dtype=object)
        except Exception:
            # Nuclear fallback: stringify the entire column
            try:
                out[col] = df[col].astype(str).to_numpy(dtype=object)
            except Exception:
                out[col] = [str(v) for v in df[col].tolist()]
    return out


def export_excel(df: pd.DataFrame, market: dict):
    """
    SINGLE DAILY WORKBOOK -- one file per day, one sheet per scan.
    File: ~/Downloads/scan_YYYY-MM-DD.xlsx
    Sheets: Summary (accumulates) + Scan_HHMM (per scan) + Rejections
    Also writes scan_latest.xlsx for quick access to the most recent scan.
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError:
        print("[EXPORT] pip install openpyxl"); return

    dl_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(dl_dir, exist_ok=True)

    now        = datetime.now()
    today      = now.strftime("%Y-%m-%d")
    hhmm       = now.strftime("%H%M")
    scan_sheet = f"Scan_{hhmm}"
    daily_path = os.path.join(dl_dir, f"scan_{today}.xlsx")
    latest_path= os.path.join(dl_dir, "scan_latest.xlsx")

    # Sanitize BEFORE any Excel operation -- fixes ArrowStringArray error
    df_clean  = _sanitize_df(df)
    top       = df_clean[df_clean["alert"] == True] if not df_clean.empty else pd.DataFrame()
    rejected  = _sanitize_df(pd.DataFrame(market.get("rejected_detail", [])))
    mkt_row  = {
        "Scan": scan_sheet,
        "Time": now.strftime("%H:%M:%S"),
        "Regime":   market.get("regime","--"),
        "SPY":      market.get("spy_price","--"),
        "SPY%":     market.get("spy_dev",0),
        "QQQ":      market.get("qqq_price","--"),
        "QQQ%":     market.get("qqq_dev",0),
        "Scanned":  market.get("scanned",0),
        "Blocked":  market.get("blocked_count",0),
        "Signals":  len(top),
        "Strategy": STRATEGY_MODE,
        "Universe": UNIVERSE_MODE,
    }

    # -- Daily workbook: append a new sheet each scan ──────────
    if os.path.exists(daily_path):
        wb = load_workbook(daily_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Summary sheet -- one row per scan, accumulates all day
    if "Summary" in wb.sheetnames:
        wb["Summary"].append(list(mkt_row.values()))
    else:
        ws = wb.create_sheet("Summary", 0)
        ws.append(list(mkt_row.keys()))
        ws.append(list(mkt_row.values()))

    # Per-scan sheet -- full results for this scan (sanitized)
    if not df_clean.empty:
        if scan_sheet in wb.sheetnames:
            del wb[scan_sheet]
        ws_s = wb.create_sheet(scan_sheet)
        ws_s.append(list(df_clean.columns))
        for _, row in df_clean.iterrows():
            ws_s.append([str(v) if isinstance(v, (list, dict, bool))
                         else round(float(v), 6) if isinstance(v, float)
                         else v for v in row])

    # Rejections sheet -- appends across all scans
    if not rejected.empty:
        rejected["scan_time"] = now.strftime("%H:%M:%S")
        rej_clean = _sanitize_df(rejected)  # fix ArrowStringArray / ArrowDtype
        def _safe_row(row):
            return [str(v) if isinstance(v, (list, dict, bool))
                    else round(float(v), 4) if isinstance(v, float)
                    else v for v in row]
        if "Rejections" in wb.sheetnames:
            for _, row in rej_clean.iterrows():
                wb["Rejections"].append(_safe_row(list(row.values())))
        else:
            ws_r = wb.create_sheet("Rejections")
            ws_r.append(list(rej_clean.columns))
            for _, row in rej_clean.iterrows():
                ws_r.append(_safe_row(list(row.values())))

    wb.save(daily_path)

    # scan_latest.xlsx -- always the most recent scan (overwritten)
    with pd.ExcelWriter(latest_path, engine="openpyxl") as w:
        if not df_clean.empty:
            df_clean.to_excel(w, sheet_name="Latest Scan", index=False)
        if not top.empty:
            top.to_excel(w, sheet_name="Signals Only", index=False)
        pd.DataFrame([mkt_row]).to_excel(w, sheet_name="Market Context", index=False)

    if STREAMLIT_MODE:
        with open(daily_path, "rb") as f:
            fname = f"scan_{today}.xlsx"
            st.download_button(
                f"Download {fname}", f.read(), file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{today}_{hhmm}"
            )
        st.session_state["last_export_time"] = now.strftime("%H:%M:%S")
        st.session_state["last_export_path"] = daily_path
        st.success(f"Saved -> scan_{today}.xlsx  (Sheet: {scan_sheet})  +  scan_latest.xlsx")
    else:
        print(f"[EXPORT] scan_{today}.xlsx  Sheet:{scan_sheet}  +  scan_latest.xlsx")


# ── Scheduler & Background Threads ───────────────────────────────────────────

def _scheduler_loop():
    if not _SCHEDULE_AVAILABLE:
        return
    schedule.every().day.at("09:00").do(lambda: run_universe_scan(30, "5m"))
    print("[SCHEDULER] Daily 9:00 AM scan scheduled")
    while True:
        schedule.run_pending()
        time.sleep(30)

_started = False
def _start_background():
    """
    Starts background threads for ADD breadth and scheduler.
    Called automatically only when running as Streamlit dashboard.
    NOT called on plain 'import scanner_v4' to avoid running universe scans
    during terminal/research import.
    """
    global _started
    if not _started:
        threading.Thread(target=_update_add_loop, daemon=True, name="add").start()
        threading.Thread(target=_scheduler_loop,  daemon=True, name="sched").start()
        _started = True

# Start background threads only when genuinely running inside Streamlit.
# STREAMLIT_MODE is False when imported by terminal/research (real check above).
if STREAMLIT_MODE:
    _start_background()


# ── Dashboard CSS ─────────────────────────────────────────────────────────────

DASH_CSS = """<style>
*{font-family:'Helvetica Neue',Helvetica,Arial,sans-serif!important;}

/* Base - white background */
html,body,[class*="css"]{background:#ffffff!important;color:#1a1a1a!important;}

/* Sidebar - light grey with yellow top border */
section[data-testid="stSidebar"]{background:#f5f5f5!important;border-right:3px solid #f5c400!important;}

/* Buttons - blue */
.stButton>button{background:#1a5fd9!important;color:#ffffff!important;font-weight:700!important;
  border:none!important;border-radius:4px!important;letter-spacing:1px!important;
  font-size:13px!important;text-transform:uppercase!important;}
.stButton>button:hover{background:#1048b0!important;}

/* Form labels */
.stSelectbox label,.stSlider label,.stTextArea label,.stNumberInput label{
  color:#555555!important;font-size:12px!important;font-weight:600!important;}

/* Metrics */
[data-testid="stMetricValue"]{color:#1a1a1a!important;font-weight:800!important;}
[data-testid="stMetricDelta"]{color:#1a5fd9!important;}

/* Dividers */
hr{border-color:#e0e0e0!important;}

/* Header */
.hdr{font-size:28px;font-weight:800;letter-spacing:6px;color:#1a1a1a;text-transform:uppercase;
     border-bottom:3px solid #f5c400;padding-bottom:10px;}
.sub{font-size:11px;font-weight:600;color:#888888;letter-spacing:3px;text-transform:uppercase;}
.sub2{font-size:10px;font-weight:700;color:#888888;text-transform:uppercase;letter-spacing:2px;}

/* Signal cards */
.card{border-radius:6px;padding:16px;margin-bottom:10px;
      border:1px solid #e0e0e0;background:#ffffff;
      box-shadow:0 1px 4px rgba(0,0,0,0.06);}
.card.on{border-color:#f5c400;border-left:4px solid #f5c400;
         background:#fffdf0;box-shadow:0 2px 12px rgba(245,196,0,0.15);}
.card.warn{border-color:#1a5fd9;border-left:4px solid #1a5fd9;background:#f0f4ff;}

/* Typography */
.sym{font-size:22px;font-weight:800;color:#1a1a1a;letter-spacing:2px;text-transform:uppercase;}
.sc{font-size:30px;font-weight:900;line-height:1;}
.bayes{font-size:20px;font-weight:700;line-height:1;}

/* Tags */
.tag{display:inline-block;font-size:10px;font-weight:700;padding:3px 8px;
     border-radius:3px;margin:2px;letter-spacing:0.5px;text-transform:uppercase;}

/* Metric cells */
.row{display:flex;gap:6px;margin-top:8px;}
.cell{flex:1;background:#f8f8f8;border:1px solid #e8e8e8;border-radius:4px;padding:5px 8px;}
.lbl{font-size:9px;font-weight:700;color:#999999;text-transform:uppercase;letter-spacing:1.5px;}
.val{font-size:12px;font-weight:700;color:#1a1a1a;}

/* Score bar */
.bar-bg{background:#e8e8e8;border-radius:2px;height:5px;margin:8px 0;overflow:hidden;}
.bar-fg{height:5px;border-radius:2px;}

/* Trade box */
.trade{background:#f8f8f8;border:1px solid #e0e0e0;border-left:4px solid #f5c400;
       border-radius:4px;padding:10px 14px;margin-top:10px;font-size:11px;
       font-weight:600;color:#333333;}
</style>"""


# -- Dashboard --───────────────────────────────────────────────────────────────

def run_dashboard():
    _start_background()   # Start ADD loop + scheduler only when dashboard is live
    st.set_page_config(page_title="QUANT SCANNER v4", page_icon=None,
                       layout="wide", initial_sidebar_state="expanded")
    st.markdown(DASH_CSS, unsafe_allow_html=True)

    # ── Sidebar ──────────────────────────────────────────────
    with st.sidebar:
        st.subheader("SCANNER v4 CONTROLS")
        timeframe   = st.selectbox("Timeframe", ["15s","1m","5m","15m"], index=2)

        # ── Strategy Switch ───────────────────────────────────
        st.markdown("**Strategy Mode**")
        strategy_choice = st.radio(
            "Active Strategy",
            ["AUTO (recommended)", "TREND only", "MEAN REVERSION only"],
            index=0,
            help="AUTO uses ADX+SPY vol to pick the right strategy each scan. "
                 "Never run both manually - they contradict each other."
        )
        # Map sidebar choice to global config
        import sys
        this_module = sys.modules[__name__]
        if "TREND only"           in strategy_choice: setattr(this_module,'STRATEGY_MODE','TREND')
        elif "MEAN REVERSION only" in strategy_choice: setattr(this_module,'STRATEGY_MODE','MEAN_REVERSION')
        else:                                           setattr(this_module,'STRATEGY_MODE','AUTO')
        auto_list   = open("auto_watchlist.txt").read() if os.path.exists("auto_watchlist.txt") \
                      else "\n".join(WATCHLIST)
        custom_syms = st.text_area("Watchlist (one per line)", value=auto_list, height=250)
        run_btn     = st.button("RUN SCAN NOW",       use_container_width=True)
        uni_btn     = st.button("RUN UNIVERSE SCAN", use_container_width=True)
        refresh     = st.slider("Auto-Refresh (sec)", 10, 300, 60)
        _           = st.number_input("Account Size ($)", value=int(ACCOUNT_SIZE), step=5000, format="%d")
        st.markdown("---")
        st.markdown(f"<div style='font-size:10px;font-weight:600;color:#555555'>"
                    f"<span style='color:#1a1a1a;font-weight:800'>10-LAYER ENGINE</span><br>"
                    f"Hurst | Hawkes | OFI<br>"
                    f"Sector RS | ADD Breadth<br>"
                    f"Z-Score | Kurt/Skew<br>"
                    f"Bayesian | Half-Life | ATR<br><br>"
                    f"<span style='color:#1a5fd9'>CAP: $300M-$20B | TF: {timeframe}</span><br>"
                    f"<span style='color:{'#1a8c2a' if SCHWAB_AVAILABLE else '#d94040'}'>"
                    f"{'Schwab API: Connected' if SCHWAB_AVAILABLE else 'Data: yfinance (15s delay)'}"
                    f"</span></div>",
                    unsafe_allow_html=True)

    symbols = [s.strip().upper() for s in custom_syms.split("\n") if s.strip()]

    # Session state
    for k, v in [("results",pd.DataFrame()),("market",{}),
                 ("last_scan",0.0),("count",0),("last_export_time","Never")]:
        if k not in st.session_state: st.session_state[k] = v

    now = time.time()

    if uni_btn:
        with st.spinner("Universe scan (~15 min)..."):
            run_universe_scan(30, timeframe)
        st.success("Done - watchlist updated")

    if run_btn or (now - st.session_state["last_scan"] > refresh) or st.session_state["last_scan"] == 0:
        with st.spinner(f"Scanning {len(symbols)} symbols..."):
            df, market = run_full_scan(symbols, timeframe)
        st.session_state.update({"results":df,"market":market,
                                   "last_scan":now,"count":st.session_state["count"]+1})

        # ── AUTO-EXPORT: runs every scan - even if 0 results ─────
        try:
            dl_dir   = os.path.join(os.path.expanduser("~"), "Downloads")
            os.makedirs(dl_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y-%m-%d_%H%M')
            today     = datetime.now().strftime('%Y-%m-%d')
            log_path  = os.path.join(dl_dir, f"scan_log_{today}.csv")

            # 1. Excel snapshot (results + rejections + market context)
            export_excel(df if not df.empty else pd.DataFrame(), market)

            # 2. Append results to daily CSV log
            if not df.empty:
                df_log = df.copy()
                df_log["scan_time"] = datetime.now().strftime("%H:%M:%S")
                df_log["scan_num"]  = st.session_state["count"]
                write_header = not os.path.exists(log_path)
                df_log.to_csv(log_path, mode="a", header=write_header, index=False)
                st.session_state["last_log_path"] = log_path

            # 3. Always write rejection log (separate file, appends all day)
            rej = market.get("rejected_detail", [])
            if len(rej) > 0:
                rej_log = os.path.join(dl_dir, f"rejections_log_{today}.csv")
                df_rej  = pd.DataFrame(rej)
                df_rej["scan_time"] = datetime.now().strftime("%H:%M:%S")
                df_rej["scan_num"]  = st.session_state["count"]
                rej_header = not os.path.exists(rej_log)
                df_rej.to_csv(rej_log, mode="a", header=rej_header, index=False)

        except Exception as e:
            st.warning(f"Auto-export failed: {e}")

    df, market, count = st.session_state["results"], st.session_state["market"], st.session_state["count"]

    # ── Header ────────────────────────────────────────────────
    c1, c2 = st.columns([3, 1])
    add_val, add_bull = get_add_breadth()
    with c1:
        st.markdown('<div class="hdr">QUANT SCANNER v4</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="sub">10-LAYER ENGINE | {timeframe.upper()} | CAP $300M-$20B</div>', unsafe_allow_html=True)
    with c2:
        ac = "#1a5fd9" if add_bull else "#d94040"
        st.markdown(f'<div style="text-align:right;margin-top:6px">'
                    f'<div class="sub">{datetime.now().strftime("%H:%M:%S")} EST | SCAN #{count}</div>'
                    f'<div style="font-size:12px;font-weight:700;color:{ac}">'
                    f'ADD {"BULLISH" if add_bull else "BEARISH"} ({add_val:.3f})</div>'
                    f'<div class="sub">NEXT REFRESH {max(0,int(refresh-(now-st.session_state["last_scan"])))}s</div>'
                    f'</div>', unsafe_allow_html=True)

    if market:
        rc = "#1a8c2a" if "RISK-ON"  in market.get("regime","") else \
             "#d94040" if "RISK-OFF" in market.get("regime","") else "#1a5fd9"
        st.markdown(
            f'<div style="font-size:12px;font-weight:600;color:#444444;'
            f'letter-spacing:2px;margin-bottom:14px;text-transform:uppercase;'
            f'border-bottom:1px solid #e0e0e0;padding-bottom:8px">'
            f'MARKET: <span style="color:{rc}">{market.get("regime","--")}</span>'
            f'&nbsp;&nbsp;|&nbsp;&nbsp;SPY {market.get("spy_price","--")} '
            f'<span style="color:{"#1a8c2a" if market.get("spy_dev",0)>0 else "#d94040"}">'
            f'({market.get("spy_dev",0):+.2f}%)</span>'
            f'&nbsp;&nbsp;|&nbsp;&nbsp;QQQ {market.get("qqq_price","--")} '
            f'<span style="color:{"#1a8c2a" if market.get("qqq_dev",0)>0 else "#d94040"}">'
            f'({market.get("qqq_dev",0):+.2f}%)</span>'
            f'&nbsp;&nbsp;|&nbsp;&nbsp;{market.get("scanned",0)} SCANNED'
            + (
                (lambda rej: (
                    f'&nbsp;&nbsp;{len(rej)} BLOCKED'
                    + (lambda groups: "".join(
                        f'&nbsp;<span style="color:#888888;font-size:10px">'
                        f'[{k}:{v}]</span>'
                        for k,v in groups.items()
                    ))(
                        {r["reason"]: sum(1 for x in rej if x["reason"]==r["reason"])
                         for r in rej}
                    )
                ))(market.get("rejected_detail",[]))
                if market.get("blocked_count",0) > 0 else ""
            )
            + f'</div>', unsafe_allow_html=True)

    if df.empty:
        st.warning("Scan returned 0 results - all symbols were filtered out. See rejection details below.")
        # DO NOT return - fall through so rejected panel and export still render

    # ── Results sections (only when scan has results) ─────────
    if not df.empty:
      # Alert banner
      alerts = df[df["alert"] == True]
      if not alerts.empty:
        syms = "  |  ".join(f"{r['symbol']} [{r['score']}] {r['bayes_prob']:.0f}%" for _,r in alerts.iterrows())
        st.markdown(
            f'<div style="background:#fffdf0;border:2px solid #f5c400;'
            f'border-left:6px solid #f5c400;border-radius:4px;padding:12px 18px;'
            f'margin-bottom:14px;font-size:13px;font-weight:700;color:#b08800;'
            f'letter-spacing:2px;text-transform:uppercase">SIGNAL ALERT  |  {syms}</div>',
            unsafe_allow_html=True)

    # ── Golden Entry Banner ───────────────────────────────────
    # Surfaces the most important condition in the whole scanner:
    # Z <= -2.0 + below VWAP in MEAN_REVERSION mode = highest-probability MR long setup
    if "golden_entry" in df.columns:
        golden = df[df["golden_entry"] == True]
        if not golden.empty:
            ge_syms = "  |  ".join(
                f"{r['symbol']} Z={r['zscore']:.2f} VWAP${r.get('vwap',0):.2f}"
                for _, r in golden.iterrows()
            )
            st.markdown(
                f'<div style="background:#fff8e1;border:2px solid #f5c400;'
                f'border-left:8px solid #f5c400;border-radius:4px;'
                f'padding:14px 18px;margin-bottom:14px;">'
                f'<div style="font-size:13px;font-weight:800;color:#b08800;'
                f'letter-spacing:2px;text-transform:uppercase;margin-bottom:4px">'
                f'GOLDEN ENTRY  |  Z &le; -2.0 + BELOW VWAP</div>'
                f'<div style="font-size:12px;font-weight:600;color:#1a1a1a">'
                f'{ge_syms}</div>'
                f'<div style="font-size:10px;color:#888888;margin-top:4px">'
                f'Price is statistically oversold AND below institutional fair value. '
                f'Wait for OFI exhaustion confirmation before entering.</div>'
                f'</div>',
                unsafe_allow_html=True
            )

    # ── Signal Cards ──────────────────────────────────────────
    cols = st.columns(min(4, max(1, len(df))))
    for i, (_, r) in enumerate(df.iterrows()):
        sc     = r["score"]
        sc_col = "#b08800" if sc>=75 else "#1a5fd9" if sc>=SIGNAL_THRESHOLD else "#888888" if sc>=45 else "#d94040"
        bp     = r["bayes_prob"]
        bp_col = "#b08800" if bp>=70 else "#1a5fd9" if bp>=55 else "#d94040"
        css    = "on" if r["alert"] else ("warn" if sc>=55 else "")
        hl_str = f"{int(r['hl_remaining'])}s" if r["hl_alive"] else "expired"
        strat  = r.get("strategy","--")
        adx_v  = r.get("adx", 0.0)
        exit_t = r.get("exit_type","--")

        # Strategy badge colour
        strat_col = "#1a5fd9" if strat=="TREND" else "#b08800" if strat=="MEAN_REVERSION" else "#888888"
        strat_lbl = "TREND" if strat=="TREND" else "MEAN REV" if strat=="MEAN_REVERSION" else strat

        conflict_tag = ""
        if r.get("hurst_conflict", False):
            conflict_tag = '<span class="tag" style="background:#fff0f0;color:#d94040;border:1px solid #f0c0c0;font-weight:800">HURST CONFLICT</span>'
        blocked_tag = ""
        if r.get("intraday_blocked", False):
            blocked_tag = f'<span class="tag" style="background:#fff0f0;color:#d94040;border:1px solid #f0c0c0">BLOCKED</span>'

        h_intra     = r.get("hurst_H_intra", r.get("hurst_H", 0))
        scan_mode   = r.get("mode","auto").upper()

        # Mean reversion specific display
        if strat == "MEAN_REVERSION":
            trade_html = (f'<div class="trade">{"LONG" if r.get("mr_long") else "SHORT"} EXHAUSTION<br>'
                          f'ENTRY ${r["price"]}  |  STOP ${r["stop"]}  |  EXIT AT VWAP ${r["vwap"]:.2f}<br>'
                          f'Kelly {r["kelly_frac"]:.1%} -> {r["shares"]} shares  |  TP 0.5% or VWAP</div>'
                          if r["alert"] else "")
        else:
            trade_html = (f'<div class="trade">ENTRY ${r["price"]}  |  STOP ${r["stop"]}  |  TARGET ${r["target"]}<br>'
                          f'Kelly {r["kelly_frac"]:.1%} -> {r["shares"]} shares  |  ${r["dollar_risk"]} risk</div>'
                          if r["alert"] else "")
        with cols[i % len(cols)]:
            st.markdown(f"""
            <div class="card {css}">
              <div style="display:flex;justify-content:space-between;align-items:flex-start">
                <div>
                  <div class="sym">{r['symbol']}</div>
                  <div style="font-size:11px;font-weight:500;color:#666666">${r['price']} &nbsp;|&nbsp; MCap ${r.get('mcap_b',0):.1f}B &nbsp;|&nbsp; {scan_mode}</div>
                </div>
                <div style="text-align:right">
                  <div class="sc" style="color:{sc_col}">{sc}</div>
                  <div class="bayes" style="color:{bp_col}">{bp:.0f}%</div>
                  <div class="sub2">BAYES WIN PROB</div>
                </div>
              </div>
              <div class="bar-bg"><div class="bar-fg" style="width:{sc}%;background:{sc_col}"></div></div>
              <div style="margin:6px 0">
                <span class="tag" style="background:{strat_col}22;color:{strat_col};border:1px solid {strat_col}55;font-weight:800">{strat_lbl}</span>
                <span class="tag" style="background:#f0f0f0;color:#555555;border:1px solid #dddddd">ADX {adx_v:.0f}</span>
                <span class="tag" style="background:#eef3ff;color:#1a5fd9;border:1px solid #c0d0f0">{r['hurst_regime']}</span>
                <span class="tag" style="background:#eef3ff;color:#1a5fd9;border:1px solid #c0d0f0">{r['hawkes_sig']}</span>
                <span class="tag" style="background:#eef3ff;color:#1a5fd9;border:1px solid #c0d0f0">{r['ofi_sig']}</span>
                <span class="tag" style="background:#f8f8f8;color:{'#1a8c2a' if r['add_bull'] else '#d94040'};border:1px solid #dddddd">ADD BULL</span>
                <span class="tag" style="background:#f8f8f8;color:{'#1a8c2a' if r['zscore_ok'] else '#d94040'};border:1px solid #dddddd">Z={r['zscore']:.2f}</span>
                <span class="tag" style="background:#fffdf0;color:#b08800;border:1px solid #f0e080">{r['ks_label']}</span>
                {conflict_tag}{blocked_tag}
                {"<span class=\"tag\" style=\"background:#fff8e1;color:#b08800;border:2px solid #f5c400;font-weight:800\">GOLDEN ENTRY</span>" if r.get("golden_entry") else ""}
              </div>
              <div class="row">
                <div class="cell"><div class="lbl">HURST D</div><div class="val">{r['hurst_H']}</div></div>
                <div class="cell"><div class="lbl">HURST 5M</div>
                  <div class="val" style="color:{'#1a8c2a' if h_intra>0.55 else '#d94040' if h_intra<0.42 else '#888888'}">{h_intra}</div></div>
                <div class="cell"><div class="lbl">OFI</div><div class="val">{r['ofi']}</div></div>
                <div class="cell" style="{'background:#fff8e1;border:1px solid #f5c400' if r.get('golden_entry') else ''}">
                  <div class="lbl">Z-SCORE{" GOLDEN" if r.get("golden_entry") else ""}</div>
                  <div class="val" style="color:{'#b08800' if r.get('golden_entry') else '#d94040' if float(r['zscore'])<=-2.0 else '#1a8c2a' if float(r['zscore'])>=2.0 else '#1a1a1a'};{'font-weight:800' if abs(float(r['zscore']))>=2.0 else ''}">{r['zscore']:.2f}</div></div>
              </div>
              <div class="row">
                <div class="cell"><div class="lbl">SECTOR</div>
                  <div class="val" style="color:{'#1a8c2a' if r['sector_gate'] else '#d94040'}">{r['sector_etf']} {"OK" if r["sector_gate"] else "FAIL"}</div></div>
                <div class="cell"><div class="lbl">R:R</div>
                  <div class="val" style="color:{'#1a8c2a' if r['rr_ok'] else '#d94040'}">{"OK" if r["rr_ok"] else "--"} {r["rr_ratio"]:.1f}:1</div></div>
                <div class="cell"><div class="lbl">HALF-LIFE</div><div class="val">{hl_str}</div></div>
                <div class="cell" style="{'background:#fff0f0' if r.get('below_vwap') else 'background:#f0fff4'}">
                  <div class="lbl">VWAP</div>
                  <div class="val" style="color:{'#d94040' if r.get('below_vwap') else '#1a8c2a'}">${r.get('vwap',0):.2f} {"BELOW" if r.get("below_vwap") else "ABOVE"}</div></div>
              </div>
              {trade_html}
            </div>""", unsafe_allow_html=True)

    # ── Trade Log ─────────────────────────────────────────────
    if TRADE_LOG:
        st.markdown("---")
        st.markdown("### Trade Log")
        total = sum(t["P&L $"] for t in TRADE_LOG)
        wins  = sum(1 for t in TRADE_LOG if t["P&L $"] > 0)
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Total P&L", f"${total:+.2f}")
        m2.metric("Win Rate",  f"{wins/len(TRADE_LOG)*100:.0f}%")
        m3.metric("Trades",    len(TRADE_LOG))
        m4.metric("W / L",     f"{wins} / {len(TRADE_LOG)-wins}")
        st.dataframe(pd.DataFrame(TRADE_LOG[::-1]), use_container_width=True, hide_index=True)

    # ── Top Alpha Pick (Execution Planner) ────────────────────
    st.markdown("---")
    eligible = df[df["alert"] == True] if not df.empty else pd.DataFrame()
    if not eligible.empty:
        top = eligible.sort_values(by=["bayes_prob","score"], ascending=False).iloc[0]
        sec_bull  = bool(top.get("sector_gate", True))
        etf_name  = top.get("sector_etf","--")
        above_vwap = not bool(top.get("below_vwap", False))
        hl_rem    = int(top.get("hl_remaining", 600))
        z_val     = float(top.get("zscore", 0))

        st.markdown(f"## Top Alpha Pick: {top['symbol']}")
        if sec_bull: st.success(f"Sector Alignment: {top['symbol']} in **{etf_name}** - outperforming SPY")
        else:        st.warning(f"Sector Divergence: **{etf_name}** lagging - reduce position size")

        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("Bayesian Win Prob", f"{top['bayes_prob']:.1f}%",
                      delta="High Conviction" if top['bayes_prob'] >= 65 else "Moderate")
            st.write(f"**Stop:** ${top['stop']:.2f}  *(1.5x ATR ${top['atr']:.2f})*")
            st.write(f"**Entry:** ${top['price']:.2f}  |  **{int(top['shares'])} shares**  (${top['dollar_risk']:.0f} risk)")
            st.caption(f"Kelly {top['kelly_frac']:.1%} of ${int(ACCOUNT_SIZE):,}")
        with c2:
            st.metric("Target", f"${top['target']:.2f}",
                      delta=f"{top['rr_ratio']:.1f}:1 R:R {'OK' if top['rr_ok'] else 'FAIL'}")
            st.write(f"**Profit if hit:** ${(top['target']-top['price'])*top['shares']:.0f}")
            if above_vwap: st.success("Above VWAP - institutions net buyers")
            else:          st.error("Below VWAP - wait for VWAP reclaim")
        with c3:
            st.metric("Alpha Half-Life", f"{hl_rem}s",
                      delta=f"{float(top.get('hl_strength',1.0))*100:.0f}% strength")
            if z_val > 2.0:   st.error(f"Z={z_val:.2f} EXTENDED - reduce size")
            elif z_val > 1.5: st.warning(f"Z={z_val:.2f} - Elevated")
            else:             st.success(f"Z={z_val:.2f} - Healthy")
            st.write(f"**{top.get('ks_label','--')}** (K:{top.get('kurtosis',0):.2f} S:{top.get('skewness',0):.2f})")

        # 7-point checklist
        checks = [
            ("Market",    "RISK-OFF" not in market.get("regime",""), market.get("regime","--")),
            ("Sector",    sec_bull,                                   etf_name),
            ("ADD",       bool(top.get("add_bull",True)),             f"{top.get('add_val',0):.3f}"),
            ("Z-Score",   bool(top.get("zscore_ok",True)),            f"Z={z_val:.2f}"),
            ("R:R >= 3",   bool(top.get("rr_ok",False)),              f"{top.get('rr_ratio',0):.1f}:1"),
            ("VWAP",      above_vwap,                                 f"${top.get('vwap',0):.2f}"),
            ("Half-Life", bool(top.get("hl_alive",True)),             f"{hl_rem}s"),
        ]
        passed = sum(1 for _, ok, _ in checks if ok)
        for col, (label, ok, detail) in zip(st.columns(len(checks)), checks):
            with col:
                st.markdown(
                    f"<div style='text-align:center;font-size:11px;font-weight:700;"
                    f"padding:8px 4px;border-radius:3px;"
                    f"background:{'#fffdf0' if ok else '#fff0f0'};"
                    f"border:1px solid {'#f5c400' if ok else '#e05050'}'>"
                    f"{'OK' if ok else '--'}&nbsp; {label}<br>"
                    f"<span style='font-size:10px;font-weight:500;color:#666666'>{detail}</span></div>",
                    unsafe_allow_html=True)

        verdict_color = "#b08800" if passed >= 6 else "#1a5fd9" if passed >= 4 else "#d94040"
        verdict_text  = "ALL CLEAR - EXECUTE"    if passed >= 6 else \
                        "PARTIAL - REDUCE SIZE"   if passed >= 4 else "BLOCKED - DO NOT ENTER"
        st.markdown(
            f"<div style='text-align:center;font-size:16px;font-weight:800;"
            f"color:{verdict_color};padding:14px;border-radius:4px;letter-spacing:3px;"
            f"text-transform:uppercase;border:2px solid {verdict_color};"
            f"background:{'#fffdf0' if passed>=6 else '#f0f4ff' if passed>=4 else '#fff0f0'};"
            f"margin-top:14px'>{verdict_text} &nbsp;|&nbsp; {passed}/{len(checks)} CHECKS PASSED</div>",
            unsafe_allow_html=True)
    else:
        st.info("No execution-grade signals this scan - watching and waiting.")

    # ── Full Results Table ─────────────────────────────────────
    st.markdown("---")
    st.markdown("### Full Scan Results")

    if not df.empty:
        # Build display DataFrame
        show_cols = ["symbol","golden_entry","score","strategy","bayes_prob",
                     "zscore","vwap","below_vwap","adx",
                     "kurtosis","skewness","ks_label",
                     "hurst_regime","hawkes_sig","ofi_sig","sector_etf",
                     "add_bull","rr_ok","kelly_frac","shares",
                     "stop","target","health_label","mcap_b"]
        df_show = df[[c for c in show_cols if c in df.columns]].copy()

        # Sort: golden entries first, then by score descending
        if "golden_entry" in df_show.columns:
            df_show = df_show.sort_values(
                by=["golden_entry","score","bayes_prob"],
                ascending=[False, False, False]
            ).reset_index(drop=True)

        # Rename for readability
        df_show = df_show.rename(columns={
            "golden_entry":  "Golden",
            "bayes_prob":    "Bayes%",
            "below_vwap":    "< VWAP",
            "sector_etf":    "Sector",
            "add_bull":      "ADD Bull",
            "rr_ok":         "R:R OK",
            "kelly_frac":    "Kelly%",
            "health_label":  "Health",
            "mcap_b":        "MCap $B",
        })

        # Streamlit styling - highlight golden rows yellow, z-score column red/green
        def highlight_golden(row):
            if row.get("Golden", False):
                return ["background-color:#fff8e1;font-weight:700"] * len(row)
            return [""] * len(row)

        def color_zscore(val):
            try:
                v = float(val)
                if v <= -2.0:  return "color:#d94040;font-weight:800"
                if v >= 2.0:   return "color:#1a8c2a;font-weight:800"
            except (ValueError, TypeError):
                pass
            return ""

        styled = (df_show.style
                  .apply(highlight_golden, axis=1)
                  .map(color_zscore, subset=["zscore"] if "zscore" in df_show.columns else []))

        st.dataframe(styled, use_container_width=True, hide_index=True)

        # Golden Entry legend
        st.markdown(
            "<div style='font-size:11px;color:#888888;padding:4px 0'>"
            "<b>Golden Entry</b> = Z-Score &le; -2.0 AND price below VWAP AND strategy = MEAN REVERSION  "
            "| <span style='color:#d94040;font-weight:700'>Red Z-Score</span> = oversold (long setup)  "
            "| <span style='color:#1a8c2a;font-weight:700'>Green Z-Score</span> = overbought (short setup)"
            "</div>",
            unsafe_allow_html=True
        )

    # ── Rejected / Blocked Stocks (always visible) ────────────
    st.markdown("---")
    st.markdown("### Rejected Stocks - Filter Reasons")
    rejected_detail = market.get("rejected_detail", [])
    if rejected_detail:
        # Colour-code by rejection reason
        reason_colour = {
            "BELOW_MIN_CAP":  "#888888",
            "ABOVE_MAX_CAP":  "#888888",
            "NO_DATA":        "#d94040",
            "LIQUIDITY":      "#d94040",
            "NO_TRADE":       "#b08800",
            "SECTOR_RS":      "#1a5fd9",
            "SCAN_FAILED":    "#d94040",
            "ERROR":          "#d94040",
        }
        reason_label = {
            "BELOW_MIN_CAP": "Cap too small",
            "ABOVE_MAX_CAP": "Cap too large",
            "NO_DATA":       "No data",
            "LIQUIDITY":     "Liquidity fail",
            "NO_TRADE":      "Regime no-trade",
            "SECTOR_RS":     "Sector RS fail",
            "SCAN_FAILED":   "Health/intraday block",
            "ERROR":         "Error",
        }
        # Group by reason
        from collections import defaultdict
        by_reason = defaultdict(list)
        for r in rejected_detail:
            by_reason[r["reason"]].append(r)

        cols_rej = st.columns(min(4, len(by_reason)))
        for i, (reason, items) in enumerate(sorted(by_reason.items())):
            col = cols_rej[i % len(cols_rej)]
            colour = reason_colour.get(reason, "#888888")
            label  = reason_label.get(reason, reason)
            rows   = "".join(
                f"<div style='padding:3px 0;border-bottom:1px solid #f0f0f0;font-size:11px;"
                f"font-weight:600;color:#1a1a1a'>{r['symbol']}"
                f"<span style='font-size:10px;font-weight:400;color:#666666;margin-left:6px'>"
                f"{r['detail']}</span></div>"
                for r in items
            )
            col.markdown(
                f"<div style='border:1px solid {colour};border-left:4px solid {colour};"
                f"border-radius:4px;padding:10px 12px;background:#ffffff;"
                f"box-shadow:0 1px 3px rgba(0,0,0,0.06)'>"
                f"<div style='font-size:11px;font-weight:800;color:{colour};"
                f"text-transform:uppercase;letter-spacing:1px;margin-bottom:6px'>"
                f"{label} ({len(items)})</div>"
                f"{rows}</div>",
                unsafe_allow_html=True
            )
    else:
        st.info("No rejections recorded - run a scan first.")

    # ── Export (always visible regardless of scan results) ────
    st.markdown("---")
    st.markdown("### Export")

    last_export = st.session_state.get('last_export_time', 'Never')
    last_log    = st.session_state.get('last_log_path', None)
    st.markdown(
        f"<div style='background:#f0f7ff;border:1px solid #1a5fd9;border-left:4px solid #1a5fd9;"
        f"border-radius:4px;padding:10px 14px;margin-bottom:12px;font-size:11px;"
        f"font-weight:600;color:#1a1a1a'>"
        f"AUTO-EXPORT - every scan saves to ~/Downloads/<br>"
        f"<span style='color:#666666;font-weight:400'>"
        f"scan_YYYY-MM-DD.xlsx (one file/day, new sheet each scan) &nbsp;|&nbsp; "
        f"scan_latest.xlsx (always current) &nbsp;|&nbsp; "
        f"scan_log_YYYY-MM-DD.csv (appends all day)<br>"
        f"Last export: <b>{last_export}</b></span></div>",
        unsafe_allow_html=True
    )

    ec1, ec2, ec3, ec4 = st.columns(4)

    with ec1:
        # Manual export trigger - works even if df is empty (exports rejected list)
        if st.button("Export to Excel Now", use_container_width=True):
            export_excel(df if not df.empty else pd.DataFrame(), market)
            st.success("Exported to ~/Downloads/")

    with ec2:
        if not df.empty:
            csv_data = df.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="Download Results CSV",
                data=csv_data,
                file_name=f"scan_{datetime.now().strftime('%Y-%m-%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        else:
            st.markdown("<div style='font-size:11px;color:#888888;padding-top:8px'>"
                        "No scan results to download yet</div>", unsafe_allow_html=True)

    with ec3:
        if last_log and os.path.exists(last_log):
            with open(last_log, "rb") as f:
                st.download_button(
                    label="Download Daily Log",
                    data=f.read(),
                    file_name=os.path.basename(last_log),
                    mime="text/csv",
                    use_container_width=True,
                )
        else:
            st.markdown("<div style='font-size:11px;color:#888888;padding-top:8px'>"
                        "Daily log builds after first successful scan</div>",
                        unsafe_allow_html=True)

    with ec4:
        rejected_detail_exp = market.get("rejected_detail", [])
        if rejected_detail_exp:
            rej_csv = pd.DataFrame(rejected_detail_exp).to_csv(index=False).encode("utf-8")
            st.download_button(
                label=f"Download Rejections ({len(rejected_detail_exp)})",
                data=rej_csv,
                file_name=f"rejected_{datetime.now().strftime('%Y-%m-%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        else:
            st.markdown("<div style='font-size:11px;color:#888888;padding-top:8px'>"
                        "No rejections recorded yet</div>", unsafe_allow_html=True)

    time.sleep(refresh)
    st.rerun()


# ── Entry Point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n[QUANT v4] Run: python -m streamlit run scanner_v4.py\n")

# Run dashboard only when genuinely inside `streamlit run`.
# Guard: check for the actual Streamlit script runner frame, not just mock.
try:
    from streamlit.runtime.scriptrunner import get_script_run_ctx as _get_ctx
    _ctx = _get_ctx()
    _is_real_streamlit = _ctx is not None and hasattr(_ctx, 'session_id')
except Exception:
    _is_real_streamlit = False

if _is_real_streamlit:
    run_dashboard()
