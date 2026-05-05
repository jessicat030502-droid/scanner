"""
QUANT SCANNER v4 - Terminal Mode
=================================
Run:  python scanner_terminal_v4.py                    # single scan
      python scanner_terminal_v4.py --loop               # auto-refresh every 60s
      python scanner_terminal_v4.py --loop --interval 30 # refresh every 30s
      python scanner_terminal_v4.py --tf 1m              # 1-minute bars
      python scanner_terminal_v4.py --r2k                # Russell 2000 scan
      python scanner_terminal_v4.py --once               # scan once and exit

Deps: pip install yfinance pandas numpy scipy openpyxl schedule colorama

Imports all engine logic from scanner_v4.py - no duplication.
Outputs terminal table + auto-exports to Excel/CSV on every scan.
"""

import os, sys, time, argparse
from datetime import datetime

# -- Locate scanner_v4.py --
_dir = os.path.dirname(os.path.abspath(__file__))
if _dir not in sys.path:
    sys.path.insert(0, _dir)

try:
    import scanner_v4 as engine
except ImportError as e:
    print(f"[ERROR] Cannot import scanner_v4.py: {e}")
    print("Make sure scanner_v4.py is in the same folder as this script.")
    sys.exit(1)

# Ticker Intelligence layer
try:
    import ticker_intelligence as _ti
    _TI_AVAILABLE = True
except ImportError:
    _ti = None
    _TI_AVAILABLE = False

import pandas as pd
import numpy as np

# -- Optional colour support --
try:
    from colorama import init, Fore, Style
    init(autoreset=True)
    C_RESET  = Style.RESET_ALL
    C_YELLOW = Fore.YELLOW
    C_GREEN  = Fore.GREEN
    C_RED    = Fore.RED
    C_CYAN   = Fore.CYAN
    C_WHITE  = Fore.WHITE
    C_DIM    = Style.DIM
    C_BOLD   = Style.BRIGHT
    HAS_COLOR = True
except ImportError:
    C_RESET=C_YELLOW=C_GREEN=C_RED=C_CYAN=C_WHITE=C_DIM=C_BOLD=""
    HAS_COLOR = False


# -- Terminal helpers --

W = 90  # terminal width

def sep(char="-"): print(char * W)

def header_line():
    now    = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    src    = "SCHWAB LIVE" if engine.SCHWAB_AVAILABLE else "yfinance (15s delay)"
    umode  = getattr(engine, "UNIVERSE_MODE", "WATCHLIST")
    kill   = getattr(engine, "GLOBAL_KILL_SWITCH", False)
    lock   = getattr(engine, "GLOBAL_REGIME_LOCK", None)
    limit_o= getattr(engine, "LIMIT_ORDERS_ONLY", True)

    # ── Active strategy mode ──────────────────────────────────
    # AUTO = scanner selects TREND or MR each bar based on ADX + SPY vol
    # TREND / MEAN_REVERSION = forced via --regime flag
    cfg_mode = getattr(engine, "STRATEGY_MODE", "AUTO")
    if cfg_mode == "AUTO":
        mode_display = f"{C_CYAN}AUTO{C_RESET}"
    elif cfg_mode == "TREND":
        mode_display = f"{C_GREEN}TREND (forced){C_RESET}"
    else:
        mode_display = f"{C_YELLOW}MR (forced){C_RESET}"

    flags = [f"MODE: {mode_display}"]
    if kill:    flags.append(f"{C_RED}[KILL SWITCH]{C_RESET}")
    if lock:    flags.append(f"{C_YELLOW}[LOCKED: {lock}]{C_RESET}")
    if limit_o: flags.append(f"{C_GREEN}[LIMIT ORDERS]{C_RESET}")
    else:       flags.append(f"{C_RED}[MARKET ORDERS - WARNING]{C_RESET}")

    try:
        if engine.is_eod_soft_exit():        flags.append(f"{C_YELLOW}[EOD SOFT EXIT]{C_RESET}")
        allowed, _ = engine.is_trading_allowed()
        if not allowed:                       flags.append(f"{C_RED}[TRADING HALTED]{C_RESET}")
    except Exception:
        pass

    print("=" * W)
    print(f"  QUANT SCANNER v4   |   {now}   |   {src}   |   {umode}")
    print("  " + "  ".join(flags))
    print("=" * W)

def market_line(market: dict):
    regime = clean(market.get("regime","--"))
    rc  = C_GREEN if "RISK-ON" in regime else C_RED if "RISK-OFF" in regime else C_YELLOW
    spy = market.get("spy_price","--")
    spd = market.get("spy_dev", 0)
    qqq = market.get("qqq_price","--")
    qd  = market.get("qqq_dev", 0)
    # Show the actual active strategy for this scan (per-bar resolution)
    # In AUTO mode this shows what strategy the scanner actually used
    spy_today = market.get("spy_today", 0)
    active_strategy = market.get("active_strategy", getattr(engine,"STRATEGY_MODE","AUTO"))
    if active_strategy == "TREND":
        strat_c = C_GREEN; strat_s = "TREND"
    elif active_strategy == "MEAN_REVERSION":
        strat_c = C_YELLOW; strat_s = "MR"
    elif active_strategy == "NO_TRADE":
        strat_c = C_RED; strat_s = "NO_TRADE"
    else:
        strat_c = C_CYAN; strat_s = "AUTO"
    print(f"  MARKET: {rc}{regime:<14}{C_RESET}"
          f"  |  SPY  {spy}  {pct_color(spd)}({spd:+.2f}%){C_RESET}"
          f"  |  QQQ  {qqq}  {pct_color(qd)}({qd:+.2f}%){C_RESET}"
          f"  |  {strat_c}[{strat_s}]{C_RESET}")

def blocked_line(market: dict):
    rej = market.get("rejected_detail", [])
    if not rej:
        return
    from collections import Counter
    counts  = Counter(r["reason"] for r in rej)
    summary = "  ".join(f"{k}: {v}" for k,v in sorted(counts.items()))
    print(f"{C_DIM}  Filtered {len(rej)} symbols  |  {summary}{C_RESET}")
    cap_b = [r for r in rej if "CAP" in r["reason"]]
    if cap_b:
        names = "  ".join(f"{clean(r['symbol'])} ({clean(r['detail'])})"
                          for r in cap_b[:10])
        extra = f"  +{len(cap_b)-10} more" if len(cap_b) > 10 else ""
        print(f"{C_DIM}  Cap blocked:  {names}{extra}{C_RESET}")


def fmt_health(h: dict, health_label: str) -> str:
    """Format health string like: OK +2.1% | <VWAP$16.09"""
    return health_label[:28] if health_label else "--"

def clean(s: str) -> str:
    """Strip all non-ASCII characters."""
    return "".join(c for c in str(s) if ord(c) < 128)

def pad(s: str, width: int) -> str:
    return clean(s)[:width].ljust(width)

def yn(val) -> str:
    return "YES" if val else "NO "

def ok_fail(val) -> str:
    return "OK  " if val else "FAIL"

def pct_color(v: float) -> str:
    return C_GREEN if v > 0 else C_RED if v < 0 else C_RESET

def fmt_sig(raw: str, keywords: dict, width: int) -> str:
    cleaned = clean(raw).upper()
    for key, label in keywords.items():
        if key in cleaned:
            return label[:width].ljust(width)
    return cleaned[:width].ljust(width)

def fmt_hawkes(sig: str) -> str:
    return fmt_sig(sig, {
        "CLUSTERING":"CLUSTERING","BUILDING":"BUILDING","IDLE":"IDLE",
        "FADING":"FADING","SELL":"SELL PRESS"
    }, 14)

def fmt_ofi(sig: str) -> str:
    return fmt_sig(sig, {
        "ACCUMULATING":"ACCUMULATING","TOPPING":"TOPPING",
        "DISTRIBUTING":"DISTRIBUTING","SELLING":"SELLING","NEUTRAL":"NEUTRAL"
    }, 14)

def score_color(sc: float) -> str:
    if sc >= 75: return C_GREEN
    if sc >= 65: return C_CYAN
    if sc >= 45: return C_YELLOW
    return C_RED

def z_color(z: float) -> str:
    if z <= -2.0: return C_RED
    if z >= 2.0:  return C_GREEN
    return C_RESET


def print_results_table(df: pd.DataFrame, market: dict):
    """Print the main results table matching the screenshot format."""
    sep()
    # Column header
    print(f"  {'SYM':<7} {'SCORE':>6} {'STRAT':<5} "
          f"{'HEALTH':<26} {'HAWKES':<14} {'OFI':<14} "
          f"{'Z':>6} {'KELLY':>6} {'SHRS':>5}  STATUS")
    print("-" * W)

    if df.empty:
        print(f"  {C_YELLOW}No results - all symbols were filtered out."
              f"  Check the FILTER lines above.{C_RESET}")
        print("-" * W)
        return

    for _, r in df.iterrows():
        sym     = str(r["symbol"])
        score   = float(r.get("score", 0))
        raw_sc  = float(r.get("hurst_score", score))
        health  = fmt_health({}, str(r.get("health_label","--")))
        hawkes  = fmt_hawkes(str(r.get("hawkes_sig","--")))
        ofi     = fmt_ofi(str(r.get("ofi_sig","--")))
        kelly   = float(r.get("kelly_frac", 0)) * 100
        shares  = int(r.get("shares", 0))
        golden  = r.get("golden_entry", False)
        strat   = str(r.get("strategy","--"))
        vwap    = float(r.get("vwap", 0))
        below_v = bool(r.get("below_vwap", False))
        z       = float(r.get("zscore", 0))
        alert   = bool(r.get("alert", False))
        hl_alive  = bool(r.get("hl_alive", False))
        hl_rem    = float(r.get("hl_remaining", 0))
        i_blocked = bool(r.get("intraday_blocked", False))

        sc_c  = score_color(score)
        z_c   = z_color(z)
        al_mk = ""  # handled in status column
        ge_mk = ""  # handled in status column

        # Health with VWAP tag
        vwap_tag = f" <VWAP${vwap:.2f}" if below_v else ""
        health_str = health + vwap_tag

        # Half-life tag
        if i_blocked:
            hl_tag = f" {C_RED}[BLK]{C_RESET}"
        elif hl_alive:
            hl_tag = f" {C_GREEN}[{int(hl_rem)}s]{C_RESET}"
        else:
            hl_tag = f" {C_DIM}[exp]{C_RESET}"

        # Status column text - no emoji
        if i_blocked:
            status = "BLOCKED"
            sc_flag = C_RED
        elif not hl_alive:
            status = "EXPIRED"
            sc_flag = C_DIM
        elif alert and golden:
            status = "SIGNAL  GOLDEN"
            sc_flag = C_YELLOW
        elif alert:
            status = "SIGNAL"
            sc_flag = C_GREEN
        else:
            status = "WATCH"
            sc_flag = C_DIM

        strat_tag = "MR" if strat == "MEAN_REVERSION" else "TR"
        hs_pad    = pad(health_str, 26)
        hk_pad    = hawkes[:14].ljust(14)
        of_pad    = ofi[:14].ljust(14)

        print(f"  {C_BOLD}{sym:<7}{C_RESET}"
              f" {sc_c}{score:>6.1f}{C_RESET}"
              f" {strat_tag:<5}"
              f" {hs_pad}"
              f" {hk_pad}"
              f" {of_pad}"
              f" {z_c}{z:>+6.2f}{C_RESET}"
              f" {kelly:>6.1f}"
              f" {shares:>5}"
              f"  {sc_flag}{status}{C_RESET}")

    print("-" * W)
    alerts    = df[df["alert"] == True] if not df.empty else pd.DataFrame()
    goldens   = df[df["golden_entry"] == True] if "golden_entry" in df.columns else pd.DataFrame()
    gate      = "OPEN" if market.get("allows_long") else "CLOSED"
    gate_c    = C_GREEN if market.get("allows_long") else C_RED

    # Strategy breakdown -- shows what AUTO actually selected this scan
    strat_c   = df["strategy"].value_counts().to_dict() if "strategy" in df.columns else {}
    cfg_mode  = getattr(engine, "STRATEGY_MODE", "AUTO")
    act_strat = market.get("active_strategy", cfg_mode)

    if cfg_mode == "AUTO":
        trend_n = strat_c.get("TREND", 0)
        mr_n    = strat_c.get("MEAN_REVERSION", 0)
        t_c = C_GREEN if trend_n >= mr_n else C_DIM
        m_c = C_YELLOW if mr_n > trend_n else C_DIM
        strat_str = (f"{C_CYAN}AUTO{C_RESET}: "
                     f"{t_c}TREND:{trend_n}{C_RESET} "
                     f"{m_c}MR:{mr_n}{C_RESET}")
    else:
        sc = C_GREEN if cfg_mode == "TREND" else C_YELLOW
        strat_str = f"{sc}{cfg_mode} (forced){C_RESET}"

    print(f"  Signals: {C_GREEN}{len(alerts)}{C_RESET}/{len(df)}"
          f"  |  Golden entries: {C_YELLOW}{len(goldens)}{C_RESET}"
          f"  |  Gate: {gate_c}{gate}{C_RESET}"
          f"  |  {strat_str}")


def print_top_picks(df: pd.DataFrame):
    """Print detailed breakdown of alert-level signals."""
    if df.empty: return
    alerts = df[df["alert"] == True].head(5)
    if alerts.empty:
        print(f"\n{C_DIM}  No signals above threshold this scan.{C_RESET}")
        return

    print()
    print("-" * W)
    print(f"  TOP PICKS  ({len(alerts)} signal(s))")
    print("-" * W)

    for idx, (_, r) in enumerate(alerts.iterrows(), 1):
        sym    = r["symbol"]
        price  = float(r.get("price", 0))
        mcap   = float(r.get("mcap_b", 0))
        score  = float(r.get("score", 0))
        strat  = r.get("strategy","--")
        golden = bool(r.get("golden_entry", False))

        golden_note = "   *** GOLDEN ENTRY  Z <= -2.0 + BELOW VWAP ***" if golden else ""
        print(f"\n  [{idx}]  {C_BOLD}{sym}{C_RESET}   "
              f"${price:.2f}   MCap ${mcap:.1f}B   "
              f"Score: {score_color(score)}{score:.1f}{C_RESET}"
              f"   Strategy: {clean(strat)}"
              f"{C_YELLOW}{golden_note}{C_RESET}")

        # Intraday
        intra_r = float(r.get("intraday_ret", 0))
        gap_r   = float(r.get("gap_ret", 0))
        vwap    = float(r.get("vwap", 0))
        below_v = bool(r.get("below_vwap", False))
        vwap_pos   = "BELOW" if below_v else "ABOVE"
        vwap_role  = str(r.get("vwap_role", "MAGNET"))
        vwap_bands = str(r.get("vwap_bands", "1.5/2.5SD"))
        vwap_l1    = float(r.get("vwap_lower1", vwap))
        vwap_u1    = float(r.get("vwap_upper1", vwap))
        vwap_l2    = float(r.get("vwap_lower2", vwap))
        vwap_u2    = float(r.get("vwap_upper2", vwap))
        vwap_gap   = r.get("vwap_gap")

        print(f"  {'Intraday':<18} {pct_color(intra_r)}{intra_r:+.2f}%{C_RESET} from open"
              f"   Gap: {pct_color(gap_r)}{gap_r:+.2f}%{C_RESET}"
              f"   VWAP ${vwap:.2f}  ({vwap_pos} VWAP)")
        vwap_c = C_YELLOW if vwap_role == "FLOOR" else C_CYAN
        print(f"  {'VWAP Role':<18} {vwap_c}{vwap_role}{C_RESET}  "
              f"bands={vwap_bands}  "
              f"1SD: ${vwap_l1:.2f}-${vwap_u1:.2f}  "
              f"2SD: ${vwap_l2:.2f}-${vwap_u2:.2f}")
        if vwap_gap:
            print(f"  {'Gap VWAP':<18} ${vwap_gap:.2f}  "
                  f"(gap-open anchor -- institutional support level)")

        # Health
        hlth  = float(r.get("health_mult", 1.0))
        hlbl  = str(r.get("health_label","--"))
        hc    = C_GREEN if hlth > 0.7 else C_YELLOW if hlth > 0.3 else C_RED
        print(f"  {'Health':<18} {hc}{hlbl}{C_RESET}  ({hlth:.2f}x)")

        # Indicators
        H      = float(r.get("hurst_H", 0))
        h_reg  = str(r.get("hurst_regime","--"))
        Hi     = float(r.get("hurst_H_intra", 0))
        hawk_l = float(r.get("hawkes_lam", 0))
        hawk_s = str(r.get("hawkes_sig","--"))
        ofi_v  = float(r.get("ofi", 0))
        ofi_s  = str(r.get("ofi_sig","--"))
        z_v    = float(r.get("zscore", 0))
        adx_v  = float(r.get("adx", 0))
        print(f"  {'Hurst (daily)':<18} H={H:.3f}  {clean(h_reg)}"
              f"   5m H={Hi:.3f}   ADX={adx_v:.1f}")
        zc = z_color(z_v)
        print(f"  {'Hawkes':<18} lam={hawk_l:.4f}   {fmt_hawkes(hawk_s)}")
        print(f"  {'OFI':<18} val={ofi_v:.4f}   {fmt_ofi(ofi_s)}")
        print(f"  {'Z-Score':<18} {zc}{z_v:+.3f}{C_RESET}"
              f"   (threshold +/- {engine.Z_ENTRY_THRESH:.1f})")

        # Sector + ADD
        sec_etf  = str(r.get("sector_etf","--"))
        sec_rs   = float(r.get("sector_rs", 1.0))
        sec_gate = bool(r.get("sector_gate", False))
        add_bull = bool(r.get("add_bull", True))
        add_val  = float(r.get("add_val", 0))
        sec_c    = C_GREEN if sec_gate else C_RED
        add_c    = C_GREEN if add_bull else C_RED
        print(f"  {'Sector RS':<18} {sec_c}{sec_etf}  RS={sec_rs:.4f}  ({ok_fail(sec_gate)}){C_RESET}")
        print(f"  {'ADD Breadth':<18} {add_c}{'BULLISH' if add_bull else 'BEARISH'}"
              f"  val={add_val:.3f}{C_RESET}")

        # Strategy-specific entry info
        if strat == "MEAN_REVERSION":
            mr_l = bool(r.get("mr_long", False))
            mr_s = bool(r.get("mr_short", False))
            exh  = str(r.get("exhaustion_reason","--"))
            vol_div = bool(r.get("vol_diverge", False))
            adapt_ofi = float(r.get("adaptive_ofi", 0.30))
            dir_str = "LONG" if mr_l else "SHORT" if mr_s else "--"
            dc = C_GREEN if mr_l else C_RED if mr_s else C_DIM
            print(f"  {'MR Direction':<18} {dc}{dir_str}{C_RESET}"
                  f"   OFI threshold: {adapt_ofi:.2f}"
                  f"   Vol divergence: {yn(vol_div)}")
            print(f"  {'Exhaustion':<18} {exh[:60]}")
        else:
            h_sc = float(r.get("hurst_score", 0))
            print(f"  TREND - Hurst score: {h_sc:.1f}  |  ADX={adx_v:.1f}")

        # Half-life decay
        hl_rem   = float(r.get("hl_remaining", 0))
        hl_alive = bool(r.get("hl_alive", False))
        hl_str   = float(r.get("hl_strength", 0))
        hl_c     = C_GREEN if hl_alive else C_RED
        print(f"  Signal half-life: {hl_c}{'ALIVE' if hl_alive else 'EXPIRED'}{C_RESET}"
              f"  {hl_rem:.0f}s remaining  |  Strength: {hl_str*100:.0f}%")

        # Intraday block reason (if blocked)
        i_blocked = bool(r.get("intraday_blocked", False))
        i_reason  = str(r.get("intraday_block_reason",""))
        if i_blocked and i_reason:
            print(f"  {C_RED}[BLOCKED]  {i_reason}{C_RESET}")

        # Regime reason + liquidity status
        # In AUTO mode this shows WHY the scanner picked TREND vs MR for this bar
        reg_reason = str(r.get("regime_reason","--"))
        liq_status = str(r.get("liq_status","--"))
        strategy   = str(r.get("strategy","--"))
        cfg_mode   = getattr(engine, "STRATEGY_MODE", "AUTO")

        # Color-code by strategy
        strat_c = C_GREEN if strategy == "TREND" else C_YELLOW if strategy == "MEAN_REVERSION" else C_CYAN
        auto_tag = " (AUTO-selected)" if cfg_mode == "AUTO" else " (forced)"
        print(f"  Strategy: {strat_c}{strategy}{auto_tag}{C_RESET}  |  Reason: {reg_reason}")
        print(f"  Liquidity: {liq_status}")

        # Range Filter confirmation
        rf_val      = float(r.get("rf_val", 0))
        rf_up       = bool(r.get("rf_up", False))
        rf_confirms = bool(r.get("rf_confirms", False))
        rf_warning  = bool(r.get("rf_warning", False))
        rf_len      = int(r.get("rf_length", 25))
        rf_dir   = "UPWARD" if rf_up else "DOWNWARD"
        rf_label = "CONFIRMED" if rf_confirms else "NOT CONFIRMED"
        rf_c     = C_GREEN if rf_confirms else C_YELLOW
        print(f"  {'Range Filter':<18} len={rf_len}   val=${rf_val:.4f}   "
              f"trend={rf_dir}   {rf_c}{rf_label}{C_RESET}")
        if rf_warning:
            print(f"  {C_YELLOW}  [WARN] Score above threshold but Range Filter"
                  f" not yet confirming -- wait.{C_RESET}")

        # ── Markov Gate ───────────────────────────────────────
        mk_stable  = bool(r.get("markov_stable", True))
        mk_prob    = float(r.get("markov_stay_prob", 1.0))
        mk_adj     = float(r.get("markov_stop_adj", 1.0))
        mk_reason  = str(r.get("markov_reason", "MARKOV_DISABLED"))
        mk_enabled = getattr(engine, "MARKOV_GATE_ENABLED", False)

        if mk_enabled:
            is_active  = mk_reason.startswith("ACTIVE")
            is_no_data = mk_reason.startswith("NO_DATA")

            if is_active:
                mk_c = C_GREEN if mk_stable else C_YELLOW
                mk_label = "STABLE" if mk_stable else "UNSTABLE"
                print(f"  {'Markov Gate':<18} "
                      f"stay={mk_prob:.1%}  stop_adj={mk_adj:.0%}  "
                      f"{mk_c}{mk_label}{C_RESET}")
                if not mk_stable:
                    print(f"  {'':<18} {C_YELLOW}[WARN] Regime unstable -- "
                          f"score penalised 10%, stop tightened{C_RESET}")
            elif is_no_data:
                print(f"  {'Markov Gate':<18} {C_DIM}{mk_reason}{C_RESET}")
            else:
                print(f"  {'Markov Gate':<18} {C_DIM}{mk_reason[:65]}{C_RESET}")
        else:
            print(f"  {'Markov Gate':<18} {C_DIM}DISABLED  "
                  f"(run --markov-gate {clean(str(r.get('symbol','')))} "
                  f"--days 60 to test){C_RESET}")
        stop_p  = float(r.get("stop", 0))
        target  = float(r.get("target", vwap))
        rr      = float(r.get("rr_ratio", 0))
        kf      = float(r.get("kelly_frac", 0))
        sh      = int(r.get("shares", 0))
        drisk   = float(r.get("dollar_risk", 0))
        exit_t  = str(r.get("exit_type","--")).replace("_"," ")
        bayes   = float(r.get("bayes_prob", 0))
        print(f"  {'TRADE PLAN':-<50}")
        print(f"  {'Entry':<18} ${price:.2f}")
        print(f"  {'Stop':<18} ${stop_p:.2f}")
        print(f"  {'Target':<18} ${target:.2f}   ({exit_t})")
        print(f"  {'Risk/Reward':<18} {rr:.1f}:1")
        print(f"  {'Kelly':<18} {kf*100:.1f}%  ->  {sh} shares   (${drisk:.0f} at risk)")
        print(f"  {'Bayes Win Prob':<18} {C_CYAN}{bayes:.1f}%{C_RESET}")
        print()


def export_results(df: pd.DataFrame, market: dict):
    """
    SINGLE DAILY WORKBOOK - one file per day, one sheet per scan.

    File:   ~/Downloads/scan_YYYY-MM-DD.xlsx   (e.g. scan_2026-04-30.xlsx)
    Sheets added each scan:
      Summary          - one row per scan, accumulates all day
      Scan_HHMM        - full results for this specific scan
      Rejections       - all blocked stocks (appends across scans)

    Also writes scan_latest.xlsx (always current scan, single sheet).
    Rejections CSV still appends all day as a separate file.

    Nothing is deleted. Every scan adds a sheet. At end of day you have
    one clean workbook with the full day's history inside.
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError:
        print(f"  {C_YELLOW}[EXPORT] pip install openpyxl{C_RESET}")
        _export_csv_only(df, market)
        return

    dl_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(dl_dir, exist_ok=True)
    now   = datetime.now()
    hhmm  = now.strftime("%H%M")
    today = now.strftime("%Y-%m-%d")
    rej   = market.get("rejected_detail", [])

    # ── Always export rejections CSV ──────────────────────────
    if len(rej) > 0:
        rej_df            = pd.DataFrame(rej)
        rej_df["scan_time"] = now.strftime("%H:%M:%S")
        rej_path          = os.path.join(dl_dir, f"rejections_log_{today}.csv")
        rej_df.to_csv(rej_path, mode="a",
                      header=not os.path.exists(rej_path), index=False)
        print(f"  {C_GREEN}[EXPORT] Rejections ->{C_RESET} "
              f"rejections_log_{today}.csv  ({len(rej)} blocked)")
    else:
        rej_df = pd.DataFrame()

    if df.empty:
        print(f"  {C_YELLOW}[SKIP] No stocks passed filters -- "
              f"daily workbook not updated.{C_RESET}")
        return

    # ── Daily workbook: scan_YYYY-MM-DD.xlsx ──────────────────
    daily_path  = os.path.join(dl_dir, f"scan_{today}.xlsx")
    scan_sheet  = f"Scan_{hhmm}"

    top     = df[df["alert"] == True]
    mkt_row = {
        "Scan":     scan_sheet,
        "Time":     now.strftime("%H:%M:%S"),
        "Regime":   market.get("regime","--"),
        "SPY":      market.get("spy_price","--"),
        "SPY%":     market.get("spy_dev",0),
        "QQQ":      market.get("qqq_price","--"),
        "QQQ%":     market.get("qqq_dev",0),
        "Scanned":  market.get("scanned",0),
        "Blocked":  market.get("blocked_count",0),
        "Signals":  len(top),
        "Strategy": engine.STRATEGY_MODE,
        "Universe": engine.UNIVERSE_MODE,
    }

    # Load existing workbook or create fresh
    if os.path.exists(daily_path):
        wb = load_workbook(daily_path)
    else:
        wb = openpyxl.Workbook()
        # Remove the default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Sheet 1: Summary (one row per scan, always first) ──────
    if "Summary" in wb.sheetnames:
        ws_sum = wb["Summary"]
        ws_sum.append(list(mkt_row.values()))
    else:
        ws_sum = wb.create_sheet("Summary", 0)
        ws_sum.append(list(mkt_row.keys()))   # header
        ws_sum.append(list(mkt_row.values())) # first data row

    # ── Sheet 2+: Scan_HHMM (this scan's full results) ────────
    # If sheet name already exists (re-run at same minute), overwrite it
    if scan_sheet in wb.sheetnames:
        del wb[scan_sheet]
    ws_scan = wb.create_sheet(scan_sheet)

    # Write column headers
    cols = list(df.columns)
    ws_scan.append(cols)
    for _, row in df.iterrows():
        ws_scan.append([
            str(v) if isinstance(v, (list, dict, bool)) else
            round(float(v), 6) if isinstance(v, float) else
            int(v) if isinstance(v, (int,)) else v
            for v in [row[c] for c in cols]
        ])

    # ── Rejections sheet (appends across all scans) ────────────
    if not rej_df.empty:
        rej_df_w = rej_df.copy()
        if "Rejections" in wb.sheetnames:
            ws_rej = wb["Rejections"]
            for _, row in rej_df_w.iterrows():
                ws_rej.append(list(row.values()))
        else:
            ws_rej = wb.create_sheet("Rejections")
            ws_rej.append(list(rej_df_w.columns))
            for _, row in rej_df_w.iterrows():
                ws_rej.append(list(row.values()))

    wb.save(daily_path)
    sheets = [s for s in wb.sheetnames if s != "Summary"]
    print(f"\n  {C_GREEN}[EXPORT] Daily workbook ->{C_RESET} "
          f"scan_{today}.xlsx  "
          f"({len(sheets)} scan(s) today  |  Sheet: {scan_sheet})")

    # ── scan_latest.xlsx — always shows the most recent scan ──
    latest_path = os.path.join(dl_dir, "scan_latest.xlsx")
    with pd.ExcelWriter(latest_path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Latest Scan", index=False)
        if not top.empty:
            top.to_excel(w, sheet_name="Signals Only", index=False)
        pd.DataFrame([mkt_row]).to_excel(w, sheet_name="Market Context", index=False)
        if not rej_df.empty:
            rej_df.to_excel(w, sheet_name="Rejected", index=False)
    print(f"  {C_GREEN}[EXPORT] Latest ->{C_RESET} scan_latest.xlsx  (overwritten)")

    # -- scan_log_DATE.csv: appending CSV all day ──────────────
    # This is the SOURCE FILE for hypothesis testing.
    # tag_outcomes reads this file to fetch actual outcomes.
    # Contains every column from the scan including score,
    # hurst_score, zscore, target, stop, price, alert, strategy.
    log_path = os.path.join(dl_dir, f"scan_log_{today}.csv")
    df_log   = df.copy()
    df_log["scan_time"] = now.strftime("%H:%M:%S")
    df_log["scan_num"]  = getattr(run_scan, "_count", 1)
    df_log.to_csv(log_path, mode="a",
                  header=not os.path.exists(log_path), index=False)
    print(f"  {C_GREEN}[EXPORT] Scan log ->{C_RESET} "
          f"scan_log_{today}.csv  (appended for hypothesis testing)")


def _export_csv_only(df: pd.DataFrame, market: dict):
    dl_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(dl_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M")
    if not df.empty:
        p = os.path.join(dl_dir, f"scan_{ts}.csv")
        df.to_csv(p, index=False)
        print(f"  {C_GREEN}[EXPORT] CSV ->{C_RESET} {p}")
    rej = market.get("rejected_detail",[])
    if len(rej) > 0:  # Fixed: was "if rej:" which fails on DataFrame
        rp = os.path.join(dl_dir, f"rejected_{ts}.csv")
        pd.DataFrame(rej).to_csv(rp, index=False)
        print(f"  {C_GREEN}[EXPORT] Rejections CSV ->{C_RESET} {rp}")


def run_scan(timeframe: str = "5m"):
    """Run one full scan, print results, export files."""
    if not hasattr(run_scan, "_count"):
        run_scan._count = 1
    os.system("cls" if os.name == "nt" else "clear")
    sep("=")
    header_line()
    sep("=")

    print(f"\n  {C_DIM}Scanning {len(engine.WATCHLIST)} symbols on {timeframe} bars...{C_RESET}")
    df, market = engine.run_full_scan(engine.WATCHLIST, timeframe)

    # Market + filter summary
    print()
    market_line(market)
    blocked_line(market)
    print()

    # Results table
    print_results_table(df, market)

    # Top picks detail
    print_top_picks(df)

    # Export
    sep()
    export_results(df, market)

    # -- Single daily Excel log (all scans in one workbook) --
    try:
        import scanner_research_v4 as research
        scan_n = getattr(run_scan, "_count", 1)
        research.append_scan_to_daily_excel(df, market, scan_num=scan_n)
        run_scan._count = scan_n + 1
    except Exception as e:
        pass  # Research module optional

    sep("=")

    return df, market



# -- Help text shown with --help --

HELP_TEXT = """
COMMANDS:
  Single scan (runs once and exits):
    python scanner_terminal_v4.py
    python scanner_terminal_v4.py --once

  Continuous loop (re-scans every N seconds):
    python scanner_terminal_v4.py --loop
    python scanner_terminal_v4.py --loop --interval 30
    python scanner_terminal_v4.py --loop --interval 120

  Hybrid dynamic watchlist (recommended):
    python scanner_terminal_v4.py --loop --dynamic
    python scanner_terminal_v4.py --loop --dynamic --sync-interval 30
    python scanner_terminal_v4.py --loop --dynamic --auto-sync
    Adds high-RVOL/momentum runners to your core watchlist every hour.
    --auto-sync automatically tightens to 30min sync when SPY moves >1.5%% today
    or regime is RISK-OFF, reverts to 60min when market calms. No extra API calls.
    Writes auto_watchlist.txt + current_universe.csv each sync.

  Change timeframe:
    python scanner_terminal_v4.py --tf 1m    # 1-minute bars
    python scanner_terminal_v4.py --tf 5m    # 5-minute (default)
    python scanner_terminal_v4.py --tf 15m   # 15-minute

  Russell 2000 full scan (8-12 min, runs once):
    python scanner_terminal_v4.py --r2k
    python scanner_terminal_v4.py --r2k --tf 5m

  Override universe mode:
    python scanner_terminal_v4.py --universe WATCHLIST
    python scanner_terminal_v4.py --universe RUSSELL2000

  Combine options:
    python scanner_terminal_v4.py --loop --interval 60 --tf 5m
    python scanner_terminal_v4.py --loop --interval 300 --r2k

RESEARCH COMMANDS (separate script):
    python scanner_research_v4.py --profile ALKT CRUS BOOT --days 60
    python scanner_research_v4.py --backtest --days 30
    python scanner_research_v4.py --sensitivity --days 20
    python scanner_research_v4.py --consistency
    python scanner_research_v4.py --twap ALKT 100 --side LONG --price 16.27

OUTPUT FILES (in ~/Downloads/):
    scan_YYYY-MM-DD_HHMM.xlsx   -- timestamped Excel snapshot per scan
    scan_latest.xlsx            -- always the most recent scan
    scan_log_YYYY-MM-DD.csv     -- daily rolling log (appends every scan)
    rejections_log_YYYY-MM-DD.csv -- blocked stocks with reasons
    daily_YYYY-MM-DD.xlsx       -- single daily workbook (all scans in one)
"""

# -- Entry point --

def main():
    parser = argparse.ArgumentParser(
        description="Quant Scanner v4 -- Terminal Mode",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=HELP_TEXT)
    parser.add_argument("--tf",       default="5m",        help="Timeframe: 15s/1m/5m/15m")
    parser.add_argument("--loop",     action="store_true", help="Auto-refresh loop")
    parser.add_argument("--once",     action="store_true", help="Single scan and exit")
    parser.add_argument("--interval", type=int, default=60,help="Refresh interval seconds")
    parser.add_argument("--r2k",      action="store_true", help="Run full Russell 2000 scan")
    parser.add_argument("--universe",  default=None,
                        help="WATCHLIST | RUSSELL2000 -- override UNIVERSE_MODE config")
    parser.add_argument("--kill",      action="store_true",
                        help="Set GLOBAL_KILL_SWITCH=True (halt all new entries)")
    parser.add_argument("--resume",    action="store_true",
                        help="Set GLOBAL_KILL_SWITCH=False (re-enable entries)")
    parser.add_argument("--regime",    default=None,
                        help="Lock strategy: TREND | MEAN_REVERSION | AUTO")
    parser.add_argument("--no-color",  action="store_true",
                        help="Disable colors (for log files or basic terminals)")
    parser.add_argument("--dynamic",       action="store_true",
                        help="Enable hybrid dynamic watchlist (RVOL+momentum runners merged hourly)")
    parser.add_argument("--auto-sync",    action="store_true", default=False,
                        help="Auto-tighten universe sync to 30min when market is volatile "
                             "(SPY move > 1.5%% today or RISK-OFF regime). "
                             "Reverts to 60min when market calms. "
                             "Only active when --dynamic is also set.")
    parser.add_argument("--sync-interval", type=int, default=60,
                        help="Minutes between dynamic watchlist refreshes (default 60)")
    args = parser.parse_args()

    # -- Apply CLI overrides --
    if args.no_color:
        global C_RESET,C_YELLOW,C_GREEN,C_RED,C_CYAN,C_WHITE,C_DIM,C_BOLD
        C_RESET=C_YELLOW=C_GREEN=C_RED=C_CYAN=C_WHITE=C_DIM=C_BOLD=""
        print("  Colors disabled")

    if args.universe:
        engine.UNIVERSE_MODE = args.universe.upper()
        print(f"  Universe: {engine.UNIVERSE_MODE}")

    if args.kill:
        engine.GLOBAL_KILL_SWITCH = True
        print(f"  [ON]  KILL SWITCH ENABLED -- no new entries will fire")

    if args.resume:
        engine.GLOBAL_KILL_SWITCH = False
        print(f"  [OFF] KILL SWITCH DISABLED -- entries re-enabled")

    if args.regime:
        reg = args.regime.upper()
        if reg in ("TREND","MEAN_REVERSION","AUTO"):
            engine.GLOBAL_REGIME_LOCK = None if reg == "AUTO" else reg
            engine.STRATEGY_MODE      = reg
            print(f"  Strategy locked to: {reg}")
        else:
            print(f"  Unknown regime '{reg}' -- use TREND, MEAN_REVERSION, or AUTO")

    if args.r2k or engine.UNIVERSE_MODE == "RUSSELL2000":
        print(f"\n  {C_YELLOW}{'='*60}")
        print(f"  RUSSELL 2000 MODE -- Full universe scan")
        print(f"  Stricter thresholds: $20M DV . 1.5x RelVol . Score>72")
        print(f"  Expected time: 8-12 min . Rate limit pauses between batches")
        print(f"  {'='*60}{C_RESET}\n")
        engine.UNIVERSE_MODE = "RUSSELL2000"
        df, market = engine.run_russell2000_scan(args.tf)
        print_results_table(df, market)
        print_top_picks(df)
        sep()
        export_results(df, market)
        sep("=")
        return

    if args.once or not args.loop:
        run_scan(args.tf)
        return

    # ── Continuous loop ───────────────────────────────────────────────────────
    # With --dynamic: runs generate_and_save() every --sync-interval minutes
    # to refresh auto_watchlist.txt with high-RVOL + momentum runners.
    # The file-read below then loads that merged list before each scan.
    #
    # Without --dynamic: same as before -- reads auto_watchlist.txt only if
    # you wrote it manually (backward-compatible).

    from datetime import timedelta

    dynamic_enabled       = args.dynamic
    sync_interval_min     = args.sync_interval      # current active interval (may auto-adjust)
    sync_interval_base    = args.sync_interval      # user's baseline (30 or 60 from CLI)
    sync_interval_volatile= max(15, args.sync_interval // 2)  # half of base, min 15 min
    auto_sync_enabled     = args.auto_sync and args.dynamic   # only useful with --dynamic
    _last_volatile_state  = None                              # track transitions to print once
    last_sync_time        = datetime.min   # ensures first-scan sync when --dynamic

    # Import dynamic generator (only needed with --dynamic, but safe to import always)
    _dynamic_gen = None
    if dynamic_enabled:
        try:
            import dynamic_watchlist as _dyn_mod
            _dynamic_gen = _dyn_mod.generate_and_save
            print(f"  {C_GREEN}[DYNAMIC]{C_RESET} Hybrid mode ON  "
                  f"|  Sync every {sync_interval_min} min  "
                  f"|  First sync will run before scan #1")
        except ImportError:
            print(f"  {C_YELLOW}[DYNAMIC] WARNING:{C_RESET} dynamic_watchlist.py not found. "
                  f"Running in standard mode.")
            dynamic_enabled = False

    auto_wl_path = os.path.join(_dir, "auto_watchlist.txt")

    # ── Ticker Intelligence startup ─────────────────────────────────────────
    if _TI_AVAILABLE:
        try:
            # Bootstrap from any historical rejection CSVs (idempotent, fast on repeats)
            _ti.bootstrap_from_historical_csvs()
            # Print a brief intelligence report at startup
            _intel_rpt = _ti.watchlist_intelligence_report(engine.WATCHLIST)
            print(_intel_rpt)
        except Exception as _ti_err:
            print(f"  [TI] Intelligence layer init warning: {_ti_err}")

    _auto_sync_note = (f"  | Auto-sync: ON (volatile={sync_interval_volatile}min, "
                        f"calm={sync_interval_base}min)")  if auto_sync_enabled else ""
    print(f"  Auto-scan every {args.interval}s on {args.tf} bars. "
          f"Universe: {engine.UNIVERSE_MODE}.{_auto_sync_note}  Press Ctrl+C to stop.\n")
    scan_count = 0
    while True:
        try:
            now = datetime.now()

            # ── Dynamic sync gate ──────────────────────────────────────────────
            # Fires on first scan and then every sync_interval_min minutes.
            # Writes auto_watchlist.txt -- the file-read below picks it up.
            if dynamic_enabled and _dynamic_gen is not None:
                due = (now - last_sync_time) > timedelta(minutes=sync_interval_min)
                if due:
                    print(f"\n  {C_CYAN}[SYNC]{C_RESET} "
                          f"{now.strftime('%H:%M:%S')}  Refreshing universe "
                          f"(RVOL + momentum screen)...")
                    try:
                        merged = _dynamic_gen(verbose=False)
                        last_sync_time = now
                        print(f"  {C_GREEN}[SYNC]{C_RESET} Universe updated: "
                              f"{len(merged)} symbols  "
                              f"(core + RVOL runners)  "
                              f"-> auto_watchlist.txt")
                    except Exception as _sync_err:
                        print(f"  {C_YELLOW}[SYNC] WARNING:{C_RESET} "
                              f"Dynamic sync failed: {_sync_err}")
                        print(f"  Continuing with current watchlist.")

            # ── Build active watchlist (intelligence-aware) ───────────────
            # Priority: profile-sorted + cooling-aware > auto_watchlist.txt > WATCHLIST
            if _TI_AVAILABLE:
                try:
                    _wl, _wl_meta = _ti.build_active_watchlist(
                        core=engine.WATCHLIST,
                        strategy_mode=engine.STRATEGY_MODE,
                        verbose=False,
                    )
                    if _wl:
                        engine.WATCHLIST = _wl
                except Exception:
                    pass  # fall through to file-based load
            if not _TI_AVAILABLE:
                if os.path.exists(auto_wl_path):
                    syms = [s.strip().upper()
                            for s in open(auto_wl_path).read().split("\n")
                            if s.strip()]
                    if syms:
                        engine.WATCHLIST = syms
                elif os.path.exists("auto_watchlist.txt"):
                    syms = [s.strip().upper()
                            for s in open("auto_watchlist.txt").read().split("\n")
                            if s.strip()]
                    if syms:
                        engine.WATCHLIST = syms

            scan_count += 1
            # Show cooled symbols if any
            if _TI_AVAILABLE:
                try:
                    _cooled = _ti.CoolingList().all_cooled()
                    _cooled_str = f"  | Cooled: {_cooled}" if _cooled else ""
                except Exception:
                    _cooled_str = ""
            else:
                _cooled_str = ""
            print(f"\n  -- Scan #{scan_count} --  "
                  f"({len(engine.WATCHLIST)} symbols in watchlist){_cooled_str}")
            _scan_df, _scan_market = run_scan(args.tf)

            # ── Auto-sync: adjust universe refresh rate based on volatility ──
            # Only active when --auto-sync flag is set. Zero impact otherwise.
            # Checks market data returned by this scan -- no extra API calls.
            if auto_sync_enabled and _scan_market:
                _spy_move   = abs(_scan_market.get("spy_today", 0))   # today's SPY % move
                _regime     = _scan_market.get("regime", "NEUTRAL")
                _is_volatile = _spy_move > 1.5 or _regime == "RISK-OFF"

                if _is_volatile and _last_volatile_state is not True:
                    # Just became volatile -- tighten sync
                    sync_interval_min   = sync_interval_volatile
                    _last_volatile_state = True
                    print(f"\n  {C_YELLOW}[AUTO-SYNC]{C_RESET} Market volatile "
                          f"(SPY {_scan_market.get('spy_today',0):+.1f}%  "
                          f"regime={_regime})  "
                          f"-- universe sync tightened to {sync_interval_min}min")
                elif not _is_volatile and _last_volatile_state is not False:
                    # Just calmed down -- restore normal sync
                    sync_interval_min   = sync_interval_base
                    _last_volatile_state = False
                    print(f"\n  {C_GREEN}[AUTO-SYNC]{C_RESET} Market calm "
                          f"(SPY {_scan_market.get('spy_today',0):+.1f}%  "
                          f"regime={_regime})  "
                          f"-- universe sync restored to {sync_interval_min}min")

            print(f"\n  Next scan in {args.interval}s...  (Ctrl+C to stop)")
            time.sleep(args.interval)
        except KeyboardInterrupt:
            print(f"\n\n  Stopped after {scan_count} scan(s). Files saved to ~/Downloads/")
            break


if __name__ == "__main__":
    main()
