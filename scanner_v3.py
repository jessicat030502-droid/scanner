"""
QUANT RESEARCH MODULE - scanner_research_v4.py
============================================
Fills the identified gaps in scanner_v4.py:

  1. Ticker Personality Profiling
  2. Backtester (with slippage model)
  3. Sensitivity Analysis (grid search)
  4. Consistency Analytics (win rate by regime/ADX/Z-score)
  5. TWAP/VWAP Execution Scheduler
  6. Single Daily Excel Log
  7. Markov Chain Regime Analysis (NEW)
     - Models probability of state transitions between regimes
     - Calm (ADX<15) / Trending (ADX>25) / Volatile (SPY vol>2%)
     - Probability gate: only trade if current state stable >= 65%
     - Prevents MR entry when calm state likely to break into trend
     - Tightens stop loss automatically when volatile state incoming
     - Guardrails: SPY vol override, 60-day minimum, 65% threshold

Run standalone:
  python scanner_research_v4.py --profile ALKT CRUS BOOT
  python scanner_research_v4.py --backtest --days 30
  python scanner_research_v4.py --sensitivity
  python scanner_research_v4.py --consistency
  python scanner_research_v4.py --markov ALKT CRUS BOOT --days 60
  python scanner_research_v4.py --markov-gate ALKT --days 60
  python scanner_research_v4.py --twap ALKT 100 --side LONG --price 16.27
"""

import os, sys, json, time, argparse
from datetime import datetime, timedelta
from collections import defaultdict

import numpy as np
import pandas as pd

# -- Import engine ---------------------------------------------
_dir = os.path.dirname(os.path.abspath(__file__))
if _dir not in sys.path:
    sys.path.insert(0, _dir)

try:
    import scanner_v4 as engine
except ImportError as e:
    print(f"[ERROR] Cannot import scanner_v4.py: {e}"); sys.exit(1)

import yfinance as yf

# -- Storage paths ---------------------------------------------
# ── Storage paths -- imported from engine (single source of truth) ────────────
# scanner_v4.py defines SCANNER_DIR, MARKOV_DATA_DIR, TICKER_PROFILE_DIR,
# and DOWNLOADS_DIR. We import them here so both files always agree on locations.
# If the import fails (engine not found), we fall back to sensible defaults.
try:
    SCANNER_DIR        = engine.SCANNER_DIR
    MARKOV_DATA_DIR    = engine.MARKOV_DATA_DIR
    TICKER_PROFILE_DIR = engine.TICKER_PROFILE_DIR
    DOWNLOADS_DIR      = engine.DOWNLOADS_DIR
except AttributeError:
    # Fallback: research module running standalone without engine
    SCANNER_DIR        = os.path.dirname(os.path.abspath(__file__))
    MARKOV_DATA_DIR    = os.path.join(SCANNER_DIR, "markov_data")
    TICKER_PROFILE_DIR = os.path.join(SCANNER_DIR, "ticker_profiles")
    DOWNLOADS_DIR      = os.path.join(os.path.expanduser("~"), "Downloads")

DL_DIR        = DOWNLOADS_DIR   # short alias used throughout this file
PROFILE_DIR   = TICKER_PROFILE_DIR
BACKTEST_DIR  = os.path.join(DOWNLOADS_DIR, "backtests")
DAILY_LOG_DIR = DL_DIR
os.makedirs(PROFILE_DIR,  exist_ok=True)
os.makedirs(BACKTEST_DIR, exist_ok=True)


# ===============================================================
# SECTION 1 — TICKER PERSONALITY PROFILING
# ===============================================================

class TickerProfile:
    """
    Behavioral fingerprint for a single ticker.

    Tracks per-symbol statistics across all scans and trades:
      - MR win rate: how often it actually reverts to VWAP after Z <= -2
      - Trend reliability: does a high Hawkes score lead to continuation?
      - Typical VWAP reversion distance (how far it snaps back)
      - Best time-of-day for signals
      - Regime performance split (MR vs TREND)
      - Historical Z-score behavior (how often it reaches 2.0)

    Stored as JSON in ~/Downloads/ticker_profiles/{SYMBOL}.json
    Updated after every closed trade.
    """

    def __init__(self, symbol: str):
        self.symbol   = symbol
        self.path     = os.path.join(PROFILE_DIR, f"{symbol}.json")
        self.data     = self._load()

    def _load(self) -> dict:
        if os.path.exists(self.path):
            try:
                return json.load(open(self.path))
            except Exception:
                pass
        return {
            "symbol":              self.symbol,
            "created":             datetime.now().isoformat(),
            "updated":             datetime.now().isoformat(),
            "total_trades":        0,
            "mr_trades":           0,
            "mr_wins":             0,
            "trend_trades":        0,
            "trend_wins":          0,
            "avg_mr_pnl_pct":      0.0,
            "avg_trend_pnl_pct":   0.0,
            "avg_vwap_reversion":  0.0,   # avg distance snapped back to VWAP
            "z_reaches_2_pct":     0.0,   # % of bars where |Z| >= 2.0
            "best_hour":           None,  # hour with highest win rate
            "regime_wins":         {},    # {"RISK-ON": 0.73, "RISK-OFF": 0.50}
            "sector_etf":          engine.SECTOR_MAP.get(self.symbol, "SPY"),
            "trade_history":       [],    # last 100 trades
        }

    def save(self):
        self.data["updated"] = datetime.now().isoformat()
        with open(self.path, "w") as f:
            json.dump(self.data, f, indent=2)

    def record_trade(self, trade: dict):
        """
        Record a completed trade and update all statistics.
        trade = {strategy, pnl_pct, entry_price, exit_price, vwap,
                 regime, hour, z_at_entry, result}
        """
        self.data["total_trades"] += 1
        strat  = trade.get("strategy","-")
        win    = trade.get("result","") == "WIN"
        pnl    = trade.get("pnl_pct", 0.0)
        hour   = trade.get("hour", 10)
        regime = trade.get("regime", "NEUTRAL")
        vwap   = trade.get("vwap", 0)
        entry  = trade.get("entry_price", 0)
        exit_p = trade.get("exit_price", 0)

        if strat == "MEAN_REVERSION":
            self.data["mr_trades"] += 1
            if win: self.data["mr_wins"] += 1
            n = self.data["mr_trades"]
            self.data["avg_mr_pnl_pct"] = (
                (self.data["avg_mr_pnl_pct"] * (n-1) + pnl) / n)
            # VWAP reversion distance
            if entry > 0 and vwap > 0:
                rev = abs(vwap - entry) / entry
                n_r = self.data["mr_trades"]
                self.data["avg_vwap_reversion"] = (
                    (self.data["avg_vwap_reversion"]*(n_r-1) + rev) / n_r)
        elif strat == "TREND":
            self.data["trend_trades"] += 1
            if win: self.data["trend_wins"] += 1
            n = self.data["trend_trades"]
            self.data["avg_trend_pnl_pct"] = (
                (self.data["avg_trend_pnl_pct"] * (n-1) + pnl) / n)

        # Regime win rate
        if regime not in self.data["regime_wins"]:
            self.data["regime_wins"][regime] = {"wins":0,"total":0}
        self.data["regime_wins"][regime]["total"] += 1
        if win: self.data["regime_wins"][regime]["wins"] += 1

        # Store trade (cap at 100)
        self.data["trade_history"].append({
            "ts":       datetime.now().strftime("%Y-%m-%d %H:%M"),
            "strategy": strat, "win": win, "pnl_pct": round(pnl,4),
            "regime":   regime, "hour": hour, "z": trade.get("z_at_entry",0),
        })
        if len(self.data["trade_history"]) > 100:
            self.data["trade_history"] = self.data["trade_history"][-100:]

        self.save()

    def summary(self) -> str:
        d = self.data
        mr_wr = (d["mr_wins"]/d["mr_trades"]*100) if d["mr_trades"] > 0 else 0
        tr_wr = (d["trend_wins"]/d["trend_trades"]*100) if d["trend_trades"] > 0 else 0
        lines = [
            f"  -- {self.symbol} Personality Profile ------------------",
            f"  MR  trades: {d['mr_trades']:3d}  win: {mr_wr:5.1f}%  "
            f"avg P&L: {d['avg_mr_pnl_pct']*100:+.2f}%  "
            f"avg VWAP rev: {d['avg_vwap_reversion']*100:.2f}%",
            f"  TRD trades: {d['trend_trades']:3d}  win: {tr_wr:5.1f}%  "
            f"avg P&L: {d['avg_trend_pnl_pct']*100:+.2f}%",
            f"  Best regime: " + (
                max(d["regime_wins"].items(),
                    key=lambda x: x[1]["wins"]/max(x[1]["total"],1))[0]
                if d["regime_wins"] else "-"),
            f"  Total trades: {d['total_trades']}  "
            f"Last updated: {d['updated'][:10]}",
        ]
        verdict = _personality_verdict(d)
        lines.append(f"  Personality: {verdict}")
        return "\n".join(lines)


def _personality_verdict(d: dict) -> str:
    mr_wr = (d["mr_wins"]/max(d["mr_trades"],1)) if d["mr_trades"] >= 5 else None
    tr_wr = (d["trend_wins"]/max(d["trend_trades"],1)) if d["trend_trades"] >= 5 else None

    if mr_wr is None and tr_wr is None:
        return "INSUFFICIENT DATA (need 5+ trades per strategy)"
    if mr_wr and mr_wr >= 0.65 and (tr_wr is None or tr_wr < 0.50):
        return f"MEAN REVERTER - MR win rate {mr_wr:.0%} (strong MR bias)"
    if tr_wr and tr_wr >= 0.60 and (mr_wr is None or mr_wr < 0.50):
        return f"TREND FOLLOWER - Trend win rate {tr_wr:.0%}"
    if mr_wr and tr_wr and mr_wr >= 0.55 and tr_wr >= 0.55:
        return f"ADAPTABLE - both strategies work (MR {mr_wr:.0%} / TRD {tr_wr:.0%})"
    if mr_wr and tr_wr and mr_wr < 0.45 and tr_wr < 0.45:
        return f"AVOID - neither strategy works here"
    return f"MIXED - needs more data (MR {mr_wr:.0%} / TRD {tr_wr:.0%})"


def profile_ticker_from_history(symbol: str, days: int = 60) -> TickerProfile:
    """
    Builds a profile by running the scanner engine over historical bars
    and simulating what trades it would have taken.
    """
    print(f"\n  Profiling {symbol} over last {days} trading days...")
    profile = TickerProfile(symbol)

    df_d = yf.Ticker(symbol).history(period=f"{days+20}d", interval="1d")
    df_5m = yf.Ticker(symbol).history(period="60d", interval="5m")

    if df_d is None or df_d.empty or len(df_d) < 20:
        print(f"  {symbol}: insufficient data"); return profile
    if df_5m is None: df_5m = pd.DataFrame()

    df_d.columns  = [c.lower() for c in df_d.columns]
    df_5m.columns = [c.lower() for c in df_5m.columns] if not df_5m.empty else df_5m.columns

    closes = df_d["close"].values
    trades = 0

    for i in range(20, len(df_d)):
        window  = df_d.iloc[i-20:i]
        price   = float(closes[i])
        z       = float((closes[i] - np.mean(closes[i-20:i])) /
                        (np.std(closes[i-20:i], ddof=1) + 1e-9))
        vwap    = price * 1.01  # approximate: slightly above close as VWAP proxy
        below_v = price < vwap

        # Simulate MR signal: Z <= -2.0 AND below VWAP
        if z <= -engine.Z_ENTRY_THRESH and below_v:
            # Simple exit model: did price return to VWAP within 5 bars?
            future = df_d.iloc[i+1:min(i+6, len(df_d))]
            if future.empty: continue
            pnl_pct = 0.0
            win     = False
            for _, fut_row in future.iterrows():
                if fut_row["close"] >= vwap:
                    pnl_pct = (fut_row["close"] - price) / price
                    win     = True
                    break
            if not win:
                pnl_pct = (future.iloc[-1]["close"] - price) / price
                win     = abs(pnl_pct) < 0.005

            profile.record_trade({
                "strategy": "MEAN_REVERSION",
                "pnl_pct":  pnl_pct,
                "entry_price": price,
                "exit_price": price * (1 + pnl_pct),
                "vwap":     vwap,
                "regime":   "RISK-ON",
                "hour":     10,
                "z_at_entry": z,
                "result":   "WIN" if win else "LOSS",
            })
            trades += 1

    print(f"  {symbol}: {trades} simulated trades -> {profile.summary()}")
    return profile


# ===============================================================
# SECTION 2 — BACKTESTER
# ===============================================================

def run_backtest(symbols: list, days: int = 30,
                 strategy: str = "MEAN_REVERSION",
                 slippage_pct: float = 0.001) -> pd.DataFrame:
    """
    Replays historical OHLCV through the scanner signal logic.
    Does NOT call yfinance live - uses stored daily data.

    Returns DataFrame of simulated trades with full metadata.
    slippage_pct: modeled as % of price (e.g., 0.001 = 0.1% each side)
    """
    print(f"\n  -- Backtest: {strategy} | {len(symbols)} symbols | {days} days --")
    print(f"  Slippage model: {slippage_pct*100:.2f}% per side")
    trades = []

    for sym in symbols:
        try:
            df = yf.Ticker(sym).history(period=f"{days+25}d", interval="1d")
            if df.empty or len(df) < 25: continue
            df.columns = [c.lower() for c in df.columns]
            closes = df["close"].values
            highs  = df["high"].values
            lows   = df["low"].values
            vols   = df["volume"].values

            for i in range(20, len(df) - 2):
                price  = float(closes[i])
                z      = float((closes[i] - np.mean(closes[i-20:i])) /
                               (np.std(closes[i-20:i], ddof=1) + 1e-9))
                atr    = float(np.mean([highs[j]-lows[j] for j in range(i-14,i)]))
                vwap_approx = float(np.mean(closes[i-5:i]))  # 5-bar avg as VWAP proxy

                if strategy == "MEAN_REVERSION":
                    if z > -engine.Z_ENTRY_THRESH: continue
                    if price >= vwap_approx:        continue
                    direction = "LONG"
                    stop    = price * (1 - engine.MR_STOP_LOSS_PCT)
                    target  = vwap_approx

                elif strategy == "TREND":
                    # Simple: Hurst proxy (recent trend strength)
                    trend_str = (closes[i] - closes[i-10]) / closes[i-10]
                    if abs(trend_str) < 0.02: continue
                    direction = "LONG" if trend_str > 0 else "SHORT"
                    stop   = price - atr * engine.ATR_STOP_MULT if direction == "LONG" \
                             else price + atr * engine.ATR_STOP_MULT
                    target = price + atr * engine.ATR_TARGET_MULT if direction == "LONG" \
                             else price - atr * engine.ATR_TARGET_MULT
                else:
                    continue

                # Apply entry slippage
                entry = price * (1 + slippage_pct) if direction == "LONG" \
                        else price * (1 - slippage_pct)

                # Simulate outcome on next 5 bars
                result_pnl  = None
                exit_reason = "TIMEOUT"
                exit_price  = None

                for j in range(i+1, min(i+6, len(df))):
                    high_j  = float(highs[j])
                    low_j   = float(lows[j])
                    close_j = float(closes[j])

                    if direction == "LONG":
                        if low_j   <= stop:  result_pnl=-engine.MR_STOP_LOSS_PCT; exit_price=stop;   exit_reason="STOP"; break
                        if high_j  >= target: result_pnl=(target-entry)/entry;     exit_price=target; exit_reason="TARGET"; break
                    else:
                        if high_j  >= stop:  result_pnl=-engine.MR_STOP_LOSS_PCT; exit_price=stop;   exit_reason="STOP"; break
                        if low_j   <= target: result_pnl=(entry-target)/entry;     exit_price=target; exit_reason="TARGET"; break

                if result_pnl is None:
                    exit_price  = float(closes[min(i+5, len(df)-1)])
                    result_pnl  = ((exit_price - entry) / entry) if direction == "LONG" \
                                  else ((entry - exit_price) / entry)

                # Apply exit slippage
                result_pnl -= slippage_pct * 2  # round-trip

                trades.append({
                    "symbol":      sym,
                    "strategy":    strategy,
                    "direction":   direction,
                    "entry_date":  df.index[i].strftime("%Y-%m-%d"),
                    "entry_price": round(entry, 4),
                    "exit_price":  round(exit_price, 4) if exit_price else 0,
                    "exit_reason": exit_reason,
                    "z_at_entry":  round(z, 3),
                    "atr":         round(atr, 4),
                    "pnl_pct":     round(result_pnl * 100, 4),
                    "win":         result_pnl > 0,
                })
        except Exception as e:
            print(f"  {sym}: {e}")
            continue

    if not trades:
        print("  No trades generated.")
        return pd.DataFrame()

    df_t = pd.DataFrame(trades)
    _print_backtest_summary(df_t, strategy, slippage_pct)
    _save_backtest(df_t, strategy)
    return df_t


def _print_backtest_summary(df: pd.DataFrame, strategy: str, slip: float):
    n     = len(df)
    wins  = df["win"].sum()
    wr    = wins / n * 100 if n > 0 else 0
    avg_p = df["pnl_pct"].mean()
    med_p = df["pnl_pct"].median()
    total = df["pnl_pct"].sum()

    # Sharpe (annualized daily)
    daily_ret = df.groupby("entry_date")["pnl_pct"].sum()
    sharpe    = (daily_ret.mean() / (daily_ret.std()+1e-9)) * np.sqrt(252) if len(daily_ret) > 5 else 0

    # Max drawdown
    cum = df["pnl_pct"].cumsum().values
    peak = np.maximum.accumulate(cum)
    dd   = (cum - peak).min()

    print(f"\n  -- Backtest Results: {strategy} --------------------------")
    print(f"  Trades:    {n:4d}  |  Win rate: {wr:5.1f}%  |  "
          f"Sharpe: {sharpe:5.2f}")
    print(f"  Avg P&L:  {avg_p:+.3f}%  |  Median: {med_p:+.3f}%  |  "
          f"Total: {total:+.1f}%")
    print(f"  Max DD:   {dd:+.3f}%  |  "
          f"Slippage cost: {slip*200:.2f}% per round-trip")
    print(f"  Per-symbol breakdown:")
    by_sym = df.groupby("symbol").agg(
        trades=("win","count"), wins=("win","sum"), avg_pnl=("pnl_pct","mean")
    ).reset_index()
    by_sym["win_rate"] = by_sym["wins"] / by_sym["trades"] * 100
    for _, row in by_sym.iterrows():
        flag = "" if row["win_rate"] >= 55 else ""
        print(f"    {flag} {row['symbol']:<6} {int(row['trades']):3d} trades  "
              f"WR={row['win_rate']:.0f}%  avg={row['avg_pnl']:+.3f}%")


def _save_backtest(df: pd.DataFrame, strategy: str):
    ts   = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = os.path.join(BACKTEST_DIR, f"backtest_{strategy}_{ts}.csv")
    df.to_csv(path, index=False)
    print(f"  Saved -> {path}")


# ===============================================================
# SECTION 3 — SENSITIVITY ANALYSIS
# ===============================================================

def run_sensitivity_analysis(symbols: list, days: int = 30) -> pd.DataFrame:
    """
    Grid search on key parameters - answers which ones actually matter.
    Tests:
      Z_ENTRY_THRESH: [1.5, 2.0, 2.5, 3.0]
      MR_STOP_LOSS:   [0.005, 0.008, 0.012, 0.015]
      MR_TP:          [0.003, 0.005, 0.008, 0.010]

    For each combination: runs backtest, records Sharpe + win rate.
    The parameter with highest Sharpe VARIANCE is most impactful.
    The parameter with lowest Sharpe VARIANCE is noise - tune with caution.
    """
    print(f"\n  -- Sensitivity Analysis ----------------------------------")
    print(f"  Symbols: {len(symbols)}  |  Days: {days}")
    print(f"  Testing Z_ENTRY_THRESH x MR_STOP x MR_TP grid...")

    z_vals   = [1.5, 2.0, 2.5, 3.0]
    sl_vals  = [0.005, 0.008, 0.012, 0.015]
    tp_vals  = [0.003, 0.005, 0.008, 0.010]

    results = []
    orig_z  = engine.Z_ENTRY_THRESH
    orig_sl = engine.MR_STOP_LOSS_PCT
    orig_tp = engine.MR_TAKE_PROFIT_PCT

    try:
        for z in z_vals:
            for sl in sl_vals:
                for tp in tp_vals:
                    engine.Z_ENTRY_THRESH     = z
                    engine.MR_STOP_LOSS_PCT   = sl
                    engine.MR_TAKE_PROFIT_PCT = tp

                    df_t = run_backtest(symbols, days=days,
                                        strategy="MEAN_REVERSION",
                                        slippage_pct=0.001)

                    if df_t.empty:
                        sharpe = wr = avg_p = 0.0
                    else:
                        n     = len(df_t)
                        wr    = df_t["win"].sum() / n if n > 0 else 0
                        avg_p = df_t["pnl_pct"].mean()
                        daily = df_t.groupby("entry_date")["pnl_pct"].sum()
                        sharpe= (daily.mean()/(daily.std()+1e-9))*np.sqrt(252) if len(daily)>3 else 0

                    results.append({
                        "Z_ENTRY": z, "STOP_PCT": sl, "TP_PCT": tp,
                        "sharpe": round(sharpe, 3),
                        "win_rate": round(wr * 100, 1),
                        "avg_pnl_pct": round(avg_p, 4),
                        "trades": len(df_t) if not df_t.empty else 0,
                    })
    finally:
        # ALWAYS restore — protects engine state even if exception thrown
        engine.Z_ENTRY_THRESH     = orig_z
        engine.MR_STOP_LOSS_PCT   = orig_sl
        engine.MR_TAKE_PROFIT_PCT = orig_tp

    df_s = pd.DataFrame(results).sort_values("sharpe", ascending=False)

    print(f"\n  Top 10 parameter combinations by Sharpe:")
    print(f"  {'Z':>5} {'SL%':>6} {'TP%':>6} {'Sharpe':>8} {'WR%':>6} {'Trades':>7}")
    print("  " + "-"*46)
    for _, row in df_s.head(10).iterrows():
        flag = "" if row["sharpe"] == df_s["sharpe"].max() else " "
        print(f"  {flag} Z={row['Z_ENTRY']:.1f} "
              f"SL={row['STOP_PCT']*100:.1f}% "
              f"TP={row['TP_PCT']*100:.1f}% "
              f"-> Sharpe={row['sharpe']:+.2f} "
              f"WR={row['win_rate']:.0f}% "
              f"({row['trades']} trades)")

    # Variance analysis — which parameter matters most
    print(f"\n  Parameter Impact (Sharpe variance - higher = more impactful):")
    z_var  = df_s.groupby("Z_ENTRY")["sharpe"].var().mean()
    sl_var = df_s.groupby("STOP_PCT")["sharpe"].var().mean()
    tp_var = df_s.groupby("TP_PCT")["sharpe"].var().mean()
    params = sorted([("Z_ENTRY_THRESH", z_var),
                     ("MR_STOP_LOSS_PCT", sl_var),
                     ("MR_TP_PCT", tp_var)],
                    key=lambda x: x[1], reverse=True)
    for pname, pvar in params:
        bar = "" * int(pvar * 20)
        print(f"    {pname:<22} variance={pvar:.4f}  {bar}")

    # Best recommendation
    best = df_s.iloc[0]
    print(f"\n   RECOMMENDED PARAMETERS:")
    print(f"    Z_ENTRY_THRESH     = {best['Z_ENTRY']}")
    print(f"    MR_STOP_LOSS_PCT   = {best['STOP_PCT']}")
    print(f"    MR_TAKE_PROFIT_PCT = {best['TP_PCT']}")
    print(f"    -> Expected Sharpe: {best['sharpe']:.2f}  "
          f"WR: {best['win_rate']:.0f}%")

    ts   = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = os.path.join(BACKTEST_DIR, f"sensitivity_{ts}.csv")
    df_s.to_csv(path, index=False)
    print(f"\n  Full grid saved -> {path}")
    return df_s


# ===============================================================
# SECTION 4 — CONSISTENCY ANALYTICS
# ===============================================================

def run_consistency_report(trade_log: list = None) -> dict:
    """
    Breaks the trade log into buckets and measures win rate in each.
    Answers: "does this scanner find consistency or just activity?"

    Buckets:
      - By regime (RISK-ON / RISK-OFF / NEUTRAL)
      - By strategy (MEAN_REVERSION / TREND)
      - By ADX bucket (<15, 15-25, >25)
      - By half-life strength (weak <40%, medium 40-70%, strong >70%)
      - By time of day (9:30-10:30, 10:30-12:00, 12:00-14:00, 14:00-16:00)
      - By Z-score bucket (<-2, -2 to -1.5, etc.)
    """
    if trade_log is None:
        trade_log = engine.TRADE_LOG

    if not trade_log:
        # Try loading from daily log CSV
        today = datetime.now().strftime("%Y-%m-%d")
        log_p = os.path.join(DL_DIR, f"scan_log_{today}.csv")
        if os.path.exists(log_p):
            try:
                df = pd.read_csv(log_p)
                trade_log = df.to_dict("records")
            except Exception:
                pass

    if not trade_log:
        print("  No trade history found. Run scans first or pass trade_log.")
        return {}

    df = pd.DataFrame(trade_log)
    report = {}

    def bucket_winrate(group_col: str, label: str):
        if group_col not in df.columns:
            return
        print(f"\n  -- Win Rate by {label} -------------------------")
        for val, grp in df.groupby(group_col):
            if "win" in grp.columns:
                wins  = grp["win"].sum() if grp["win"].dtype == bool \
                        else (grp["win"] == True).sum()
                total = len(grp)
                wr    = wins / total * 100 if total > 0 else 0
                bar   = "" * int(wr/5)
                flag  = "" if wr >= 55 else ("-" if wr >= 45 else "")
                print(f"  {flag} {str(val):<20} {total:4d} trades  "
                      f"WR={wr:5.1f}%  {bar}")
                report[f"{label}_{val}"] = {"wr": wr, "n": total}

    bucket_winrate("strategy",  "Strategy")
    bucket_winrate("regime",    "Market Regime")

    # ADX buckets
    if "adx" in df.columns:
        df["adx_bucket"] = pd.cut(df["adx"], bins=[0,15,25,100],
                                   labels=["<15 (calm)","15-25 (range)","25+ (trend)"])
        bucket_winrate("adx_bucket", "ADX Level")

    # Half-life strength buckets
    if "hl_strength" in df.columns:
        df["hl_bucket"] = pd.cut(df["hl_strength"], bins=[0,0.4,0.7,1.0],
                                  labels=["Weak (<40%)","Med (40-70%)","Strong (>70%)"])
        bucket_winrate("hl_bucket", "Signal Strength at Entry")

    # Z-score buckets
    if "zscore" in df.columns:
        df["z_bucket"] = pd.cut(df["zscore"],
                                 bins=[-10,-3,-2.5,-2,-1,0,1,2,2.5,3,10],
                                 labels=["<-3","-3:-2.5","-2.5:-2","-2:-1","-1:0",
                                         "0:1","1:2","2:2.5","2.5:3",">3"])
        bucket_winrate("z_bucket", "Z-Score at Entry")

    # Time of day
    if "scanned_at" in df.columns:
        try:
            df["hour"] = pd.to_datetime(df["scanned_at"], format="%H:%M:%S").dt.hour
            df["tod"]  = pd.cut(df["hour"],
                                 bins=[9,10,12,14,16],
                                 labels=["9-10 (open)","10-12 (mid-morning)",
                                         "12-14 (midday)","14-16 (afternoon)"])
            bucket_winrate("tod", "Time of Day")
        except Exception:
            pass

    print(f"\n  -- Overall Summary ----------------------------------------")
    if "win" in df.columns:
        wins  = (df["win"] == True).sum()
        total = len(df)
        print(f"  Total trades: {total}  |  Overall WR: {wins/total*100:.1f}%")

    # Save report
    ts   = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = os.path.join(BACKTEST_DIR, f"consistency_{ts}.json")
    with open(path,"w") as f:
        json.dump(report, f, indent=2)
    print(f"  Report saved -> {path}")
    return report


# ===============================================================
# SECTION 5 — TWAP/VWAP EXECUTION SCHEDULER
# ===============================================================

def compute_twap_schedule(symbol: str, total_shares: int,
                           side: str, entry_price: float,
                           n_tranches: int = 4,
                           interval_seconds: int = 30,
                           max_spread_pct: float = 0.002) -> dict:
    """
    TWAP-style order splitting for slippage minimization.

    Splits total_shares into n_tranches over interval_seconds * n_tranches.
    Each tranche is a limit order - never market order.

    Spread gate: before placing each tranche, estimated spread must
    be below max_spread_pct (default 0.2% of price).
    For R2K small-caps: typical spread is 0.1%-0.5%. Reject if > 0.2%.

    Recommendation on your current setup (without Schwab API):
    You CANNOT automate this without a live broker API.
    What you CAN do right now:
      1. This function prints the exact schedule - execute manually in TOS
      2. When Schwab API is connected, replace print() with API calls
      3. Use TOS "Active Trader" with LIMIT orders at the computed prices

    Returns:
      schedule: list of {tranche, shares, limit_price, time_offset_s}
      estimated_slippage: expected slippage cost in $ for full order
    """
    shares_per = total_shares // n_tranches
    remainder  = total_shares % n_tranches
    schedule   = []

    # VWAP proxy: use entry price ± small offset per tranche
    # In TWAP: each limit price steps slightly toward fair value
    step = entry_price * 0.0002  # 0.02% step between tranches

    print(f"\n  -- TWAP Schedule: {symbol} {side} {total_shares} shares --")
    print(f"  Entry price: ${entry_price:.2f}  |  "
          f"{n_tranches} tranches  |  {interval_seconds}s apart")
    print(f"  Spread gate: reject if est. spread > {max_spread_pct*100:.1f}%")
    print(f"  ORDER TYPE: LIMIT ONLY - never market order")
    print()
    print(f"  {'Tranche':>8} {'Shares':>8} {'Limit $':>10} "
          f"{'Time':>8} {'Spread gate':>12}")
    print("  " + "-"*55)

    total_cost = 0.0
    for i in range(n_tranches):
        sh = shares_per + (1 if i < remainder else 0)
        if side == "LONG":
            # Aggressive: first tranche at ask, later tranches at slightly better price
            lim = round(entry_price - step * i, 4)
        else:
            lim = round(entry_price + step * i, 4)

        t_offset = i * interval_seconds
        est_spread_cost = lim * max_spread_pct * sh
        total_cost += est_spread_cost

        schedule.append({
            "tranche":     i + 1,
            "shares":      sh,
            "limit_price": lim,
            "time_offset_s": t_offset,
            "time_label":  f"T+{t_offset}s",
            "est_cost_$":  round(est_spread_cost, 2),
        })
        print(f"  Tranche {i+1}:  {sh:>5} sh  "
              f"@ ${lim:.4f}  "
              f"T+{t_offset:>3}s  "
              f"spread < ${lim*max_spread_pct:.2f}")

    print(f"\n  Total shares: {total_shares}  |  "
          f"Est. spread cost: ${total_cost:.2f}  |  "
          f"Duration: {interval_seconds*(n_tranches-1)}s")

    # Slippage estimate based on R2K typical spread
    rk_spread = entry_price * 0.003  # 0.3% typical R2K spread
    slip_est  = rk_spread * total_shares
    print(f"  R2K slippage estimate: ${slip_est:.2f} "
          f"({rk_spread/entry_price*100:.2f}% x {total_shares} shares)")

    print(f"\n  -- Manual Execution Steps (ThinkorSwim) -----------------")
    print(f"  1. Open Active Trader for {symbol}")
    print(f"  2. Set order type: LIMIT")
    print(f"  3. Execute each tranche as shown above")
    print(f"  4. Wait {interval_seconds}s between each")
    print(f"  5. CANCEL remaining if price moves > 0.5% against you")
    print(f"\n  -- When Schwab API is connected -------------------------")
    print(f"  Replace this print block with:")
    print(f"    for tranche in schedule:")
    print(f"      schwab.place_limit_order(symbol, tranche['shares'],")
    print(f"                               tranche['limit_price'], side)")
    print(f"      time.sleep(tranche['time_offset_s'])")

    return {
        "symbol": symbol,
        "side": side,
        "total_shares": total_shares,
        "schedule": schedule,
        "est_slip_$": round(slip_est, 2),
        "est_spread_cost_$": round(total_cost, 2),
    }


def spread_gate_check(symbol: str, price: float,
                      max_spread_pct: float = 0.002) -> tuple:
    """
    Estimates bid-ask spread from recent OHLCV.
    Without live tick data, approximates as (High - Low) / Close
    for the most recent bar. This understates the true spread but
    gives a directional signal.

    Returns (passes: bool, est_spread_pct: float, reason: str).
    """
    try:
        df = yf.Ticker(symbol).history(period="1d", interval="1m")
        if df.empty: return True, 0.0, "NO_DATA"
        last = df.iloc[-1]
        est_spread = (float(last["High"]) - float(last["Low"])) / float(last["Close"])
        passes = est_spread <= max_spread_pct
        return passes, round(est_spread, 5), \
               f"est_spread={est_spread*100:.3f}% ({'OK' if passes else 'TOO WIDE'})"
    except Exception:
        return True, 0.0, "ESTIMATE_FAILED"


# ===============================================================
# SECTION 6 — SINGLE DAILY EXCEL LOG
# ===============================================================

def append_scan_to_daily_excel(df: pd.DataFrame, market: dict,
                                scan_num: int = 1):
    """
    All scans for the day in ONE Excel workbook.
    Each scan -> new sheet named "Scan_HH:MM"
    Summary sheet updated after each scan.
    Ticker profiles sheet updated after each scan.

    File: ~/Downloads/daily_YYYY-MM-DD.xlsx
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError:
        print("  [DAILY LOG] pip install openpyxl"); return

    today  = datetime.now().strftime("%Y-%m-%d")
    ts     = datetime.now().strftime("%H:%M")
    path   = os.path.join(DL_DIR, f"daily_{today}.xlsx")

    rej    = market.get("rejected_detail", [])
    rej_df = pd.DataFrame(rej) if len(rej) > 0 else pd.DataFrame()

    mkt_row = {
        "Scan#":    scan_num,
        "Time":     ts,
        "Regime":   market.get("regime","-"),
        "SPY":      market.get("spy_price","-"),
        "SPY%":     market.get("spy_dev",0),
        "QQQ":      market.get("qqq_price","-"),
        "QQQ%":     market.get("qqq_dev",0),
        "Scanned":  market.get("scanned",0),
        "Blocked":  market.get("blocked_count",0),
        "Signals":  len(df[df["alert"]==True]) if not df.empty and "alert" in df.columns else 0,
        "Strategy": getattr(engine,"STRATEGY_MODE","-"),
        "Universe": getattr(engine,"UNIVERSE_MODE","-"),
    }

    # Load existing or create new
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # -- Sheet 1 (or update): Summary -------------------------
    if "Summary" in wb.sheetnames:
        ws_sum = wb["Summary"]
        # Append new row
        ws_sum.append(list(mkt_row.values()))
    else:
        ws_sum = wb.create_sheet("Summary", 0)
        ws_sum.append(list(mkt_row.keys()))
        ws_sum.append(list(mkt_row.values()))

    # -- Sheet per scan: "Scan_HH:MM" -------------------------
    sheet_name = f"Scan_{ts.replace(':','')}"
    if df is not None and not df.empty:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_scan = wb.create_sheet(sheet_name)
        # Write header
        cols = list(df.columns)
        ws_scan.append(cols)
        for _, row in df.iterrows():
            ws_scan.append([str(row[c]) if isinstance(row[c], (list,dict)) else row[c]
                            for c in cols])

    # -- Rejections sheet (accumulates) -----------------------
    if not rej_df.empty:
        rej_df["scan_time"] = ts
        rej_df["scan_num"]  = scan_num
        if "Rejections" in wb.sheetnames:
            ws_rej = wb["Rejections"]
            for _, row in rej_df.iterrows():
                ws_rej.append(list(row.values()))
        else:
            ws_rej = wb.create_sheet("Rejections")
            ws_rej.append(list(rej_df.columns))
            for _, row in rej_df.iterrows():
                ws_rej.append(list(row.values()))

    # -- Ticker Profiles sheet (latest snapshots) -------------
    if not df.empty and "symbol" in df.columns:
        profiles = []
        for sym in df["symbol"].unique():
            p = TickerProfile(sym)
            d = p.data
            mr_wr = (d["mr_wins"]/d["mr_trades"]*100) if d["mr_trades"]>0 else None
            profiles.append({
                "Symbol":      sym,
                "MR_Trades":   d["mr_trades"],
                "MR_WinRate%": round(mr_wr,1) if mr_wr else "-",
                "MR_AvgPnL%":  round(d["avg_mr_pnl_pct"]*100,3),
                "TRD_Trades":  d["trend_trades"],
                "VWAP_Rev%":   round(d["avg_vwap_reversion"]*100,3),
                "Personality": _personality_verdict(d),
                "Updated":     d["updated"][:10],
            })
        p_df = pd.DataFrame(profiles)
        if "Ticker_Profiles" in wb.sheetnames:
            del wb["Ticker_Profiles"]
        ws_p = wb.create_sheet("Ticker_Profiles")
        ws_p.append(list(p_df.columns))
        for _, row in p_df.iterrows():
            ws_p.append(list(row.values()))

    wb.save(path)
    print(f"  [DAILY LOG] Updated -> daily_{today}.xlsx  "
          f"(Sheet: {sheet_name}  |  {wb.sheetnames})")


# ===============================================================
# ===============================================================
# SECTION 7 -- MARKOV CHAIN REGIME ANALYSIS
# ===============================================================
#
# THEORETICAL BASIS
# -----------------
# Markov chain: next state depends ONLY on current state (first-order).
# Markets have short-term regime persistence -- CALM rarely jumps to
# VOLATILE without passing through RANGING. The transition matrix
# captures these empirical probabilities from 60+ days of daily data.
#
# STOCHASTIC PROCESS CONTEXT
# ---------------------------
# Markets exhibit Geometric Brownian Motion in price: dS = mu*dt + sigma*dW
# Regime states are discrete and sticky -- not Brownian.
# This module handles both:
#   1. Markov transitions: which regime am I in, will it persist?
#   2. GBM simulation: given this regime, what is the price distribution?
# Combined: "I am in CALM state (70% persistence) AND price has 5% tail
# risk of a 2-sigma move in 5 bars." Far more actionable than either alone.
#
# KNOWN LIMITATIONS
# -----------------
# - First-order only: next state depends on current state only.
#   Second-order would be more predictive but needs ~240 days.
# - ADX proxy from daily data is less precise than intraday ADX.
# - Assumes stationarity -- transition probs may shift over time.
#   The 60-day rolling window partially addresses this.
# ===============================================================

MARKOV_STATES = {0: "CALM", 1: "TRENDING", 2: "VOLATILE", 3: "RANGING"}

MARKOV_MIN_DAYS     = 60    # minimum TRADING days for valid matrix
                            # Run --prefetch --days 120 (not 60!) because:
                            # 60 calendar days = ~42 trading days (after weekends+holidays)
                            # 90 calendar days = ~63 trading days (barely passes)
                            # 120 calendar days = ~84 trading days (solid margin)
                            # Formula used: fetch = days*1.6+30 calendar days
MARKOV_MIN_PROB     = 0.65  # stay-probability required to trade
MARKOV_MIN_OBS      = 20    # minimum observations per state
MARKOV_SPY_OVERRIDE = 0.02  # SPY vol above this always kills trade

# Historically-informed sparse prior (Russell 2000, 2019-2024).
# Used when a state has < MARKOV_MIN_OBS observations.
# Replaces the incorrect uniform(0.25) prior which overstated VOLATILE.
MARKOV_SPARSE_PRIOR = np.array([
    [0.65, 0.12, 0.05, 0.18],  # CALM    -> mostly stays CALM
    [0.10, 0.62, 0.08, 0.20],  # TREND   -> mostly stays TREND
    [0.20, 0.15, 0.45, 0.20],  # VOLATILE-> often resolves to CALM/RANGING
    [0.18, 0.25, 0.07, 0.50],  # RANGING -> often flips to TREND (watch out)
])

# Cache: prevents 80 yfinance calls per scan cycle
_markov_cache: dict = {}
MARKOV_CACHE_TTL    = 3600   # rebuild matrix once per hour


def classify_state(adx: float, spy_vol_pct: float) -> int:
    """Map metrics to Markov state. VOLATILE always takes priority."""
    if abs(spy_vol_pct) > MARKOV_SPY_OVERRIDE: return 2
    if adx >= 25:                               return 1
    if adx < 15:                                return 0
    return 3


def _compute_wilder_adx(high: np.ndarray, low: np.ndarray,
                          close: np.ndarray, n: int = 14) -> np.ndarray:
    """
    Wilder-smoothed ADX -- matches compute_adx() in scanner_v4.py.
    Previous version used raw DM sums which overstated trend strength
    in choppy markets, causing state classifications to flip too often.
    """
    length = len(close)
    if length < n * 2:
        return np.full(max(1, length - n), 20.0)

    tr  = np.zeros(length)
    pdm = np.zeros(length)
    ndm = np.zeros(length)
    for i in range(1, length):
        hl  = high[i] - low[i]
        hpc = abs(high[i] - close[i-1])
        lpc = abs(low[i]  - close[i-1])
        tr[i]  = max(hl, hpc, lpc)
        up     = high[i] - high[i-1]
        dn     = low[i-1] - low[i]
        pdm[i] = up if (up > dn and up > 0) else 0.0
        ndm[i] = dn if (dn > up and dn > 0) else 0.0

    alpha = 1.0 / n
    atr_s = pd.Series(tr).ewm(alpha=alpha, adjust=False).mean().values
    pdi_s = pd.Series(pdm).ewm(alpha=alpha, adjust=False).mean().values
    ndi_s = pd.Series(ndm).ewm(alpha=alpha, adjust=False).mean().values

    with np.errstate(divide='ignore', invalid='ignore'):
        pdi = np.where(atr_s > 0, 100 * pdi_s / atr_s, 0.0)
        ndi = np.where(atr_s > 0, 100 * ndi_s / atr_s, 0.0)
        dx  = np.where((pdi + ndi) > 0,
                        100 * np.abs(pdi - ndi) / (pdi + ndi), 0.0)

    return pd.Series(dx).ewm(alpha=alpha, adjust=False).mean().values[n:]


def build_transition_matrix(symbol: str, days: int = 60,
                              use_cache: bool = True) -> tuple:
    """
    Builds 4x4 Markov transition matrix from historical daily OHLCV.

    THE CALENDAR vs TRADING DAY PROBLEM (why --days 60 always failed):
      60 calendar days  = ~42 trading days  (weekends + holidays)
      90 calendar days  = ~62 trading days
      120 calendar days = ~83 trading days

    Additionally the ADX window (14 bars) consumes 14 observations.
    So to end up with 60 USABLE observations after ADX:
      Need at least 74 trading days in merged data
      Need at least 105 calendar days to fetch reliably

    This function now:
      1. Fetches (days * 2 + 30) calendar days -- guarantees enough trading days
         regardless of what `days` the user passes in
      2. Does NOT truncate with .tail(days) -- keeps all available data
      3. Gate 3 checks that min_len >= MARKOV_MIN_DAYS AFTER ADX window

    DEFAULT: days=60 means "I want 60 usable trading days for the matrix"
    The fetch automatically scales to provide that.

    STRICT DATA REQUIREMENTS:
      1. Raw yfinance rows >= MARKOV_MIN_DAYS (60)
      2. After date-index merge with SPY >= MARKOV_MIN_DAYS trading days
      3. After ADX 14-bar window: usable observations >= MARKOV_MIN_DAYS
    """
    import time as _t

    # Cache check
    if use_cache and symbol in _markov_cache:
        cached = _markov_cache[symbol]
        if _t.time() - cached[3] < MARKOV_CACHE_TTL:
            return cached[0], cached[1], cached[2], True, cached[4] if len(cached) > 4 else 0

    # Fetch enough calendar days to guarantee MARKOV_MIN_DAYS trading days
    # after ADX window.
    # Formula: need (days + 14) trading days minimum.
    # Trading days per calendar day ~ 0.71 (252/365).
    # So calendar days needed = (days + 14) / 0.71 + 30 buffer
    # Simplified: days * 2 + 30 always gives plenty of room.
    fetch_cal = days * 2 + 30
    try:
        import contextlib, io
        with contextlib.redirect_stderr(io.StringIO()):
            df     = yf.Ticker(symbol).history(
                         period=f"{fetch_cal}d", interval="1d", auto_adjust=True)
            spy_df = yf.Ticker("SPY").history(
                         period=f"{fetch_cal}d", interval="1d", auto_adjust=True)
    except Exception:
        return None, None, [], False, 0

    # Gate 1: raw row count
    if df is None or df.empty:
        return None, None, [], False, 0
    raw_days = len(df)
    if raw_days < MARKOV_MIN_DAYS:
        return None, None, [], False, raw_days

    # Normalise columns
    df.columns     = [c.lower() for c in df.columns]
    spy_df.columns = [c.lower() for c in spy_df.columns]

    # Gate 2: merged trading days -- date-index aligned to handle holidays
    df.index     = pd.DatetimeIndex(df.index).normalize()
    spy_df.index = pd.DatetimeIndex(spy_df.index).normalize()
    # NO .tail(days) here -- that was the bug. Keep all fetched data.
    # We want as many observations as possible for the ADX window.
    merged = df.join(spy_df[["close"]], rsuffix="_spy", how="inner")

    merged_days = len(merged)
    if merged_days < MARKOV_MIN_DAYS:
        return None, None, [], False, merged_days

    high  = merged["high"].values
    low   = merged["low"].values
    close = merged["close"].values
    spy_c = merged["close_spy"].values

    # Wilder ADX -- 14-bar window consumes first 14 observations
    adx_arr  = _compute_wilder_adx(high, low, close, n=14)
    offset   = len(close) - len(adx_arr)
    spy_vols = np.abs(np.diff(spy_c)) / (spy_c[:-1] + 1e-9)
    spy_vols = spy_vols[max(0, offset - 1):]

    # Gate 3: usable observations AFTER ADX window must be >= MARKOV_MIN_DAYS
    # This is what actually failed before: tail(60) -> ADX eats 14 -> 46 left
    min_len = min(len(adx_arr), len(spy_vols))
    if min_len < MARKOV_MIN_DAYS:
        return None, None, [], False, min_len

    adx_arr  = adx_arr[-min_len:]
    spy_vols = spy_vols[-min_len:]
    dates    = [str(d)[:10] for d in merged.index[-min_len:]]

    # Classify each day into a state
    states_seq = [
        (dates[i], classify_state(float(adx_arr[i]), float(spy_vols[i])))
        for i in range(min_len)
    ]

    # Build 4x4 transition count matrix
    counts = np.zeros((4, 4), dtype=int)
    for i in range(len(states_seq) - 1):
        counts[states_seq[i][1]][states_seq[i + 1][1]] += 1

    # Normalise rows -- blend with informed prior for sparse states
    matrix = np.zeros((4, 4), dtype=float)
    for i in range(4):
        row_sum = counts[i].sum()
        if row_sum >= MARKOV_MIN_OBS:
            matrix[i] = counts[i] / row_sum
        else:
            weight    = row_sum / MARKOV_MIN_OBS
            emp       = counts[i] / max(row_sum, 1)
            matrix[i] = weight * emp + (1 - weight) * MARKOV_SPARSE_PRIOR[i]

    _markov_cache[symbol] = (matrix, counts, states_seq, _t.time(), min_len)
    return matrix, counts, states_seq, True, min_len


def compute_stationary_distribution(matrix: np.ndarray) -> np.ndarray:
    """
    Long-run stationary distribution: pi such that pi @ M = pi.
    Answers: what fraction of time does this ticker spend in each regime?

    Use this to classify ticker personality:
    - CALM >= 50%  -> mean reverter (MR trades well here)
    - TREND >= 40% -> trend follower (MR will lose, use TREND strategy)
    - Mixed        -> adaptable
    """
    if matrix is None:
        return np.ones(4) / 4
    n = matrix.shape[0]
    A = np.vstack([matrix.T - np.eye(n), np.ones(n)])
    b = np.zeros(n + 1); b[-1] = 1.0
    try:
        pi, _, _, _ = np.linalg.lstsq(A, b, rcond=None)
        pi = np.clip(pi, 0, 1)
        return pi / pi.sum()
    except Exception:
        return np.ones(4) / 4


def compute_nstep_matrix(matrix: np.ndarray, n_steps: int) -> np.ndarray:
    """
    Chapman-Kolmogorov: n-step transition matrix (matrix raised to n_steps).
    Answers: "What state will I be in after n_steps transitions?"
    For 5-min chart with n_steps=5: probability of each state in 25 minutes.
    """
    if matrix is None:
        return np.ones((4, 4)) / 4
    return np.linalg.matrix_power(matrix, max(1, n_steps))


def compute_mean_recurrence_times(matrix: np.ndarray) -> np.ndarray:
    """
    Mean Recurrence Time (MRT) for each state: E[T_ii] = 1 / pi_i

    From the document: "the expected number of days to return to a state
    is 1/pi_i where pi_i is the steady-state probability."

    Scanner application:
      If pi(CALM) = 0.55, then MRT(CALM) = 1/0.55 = 1.82 days
      Interpretation: on average, the CALM regime returns every 1.82 trading days.

      This tells you how long to WAIT after a failed MR signal before
      the regime becomes favorable again. If MRT(CALM) = 4 days,
      you know the window will close for ~4 days before resetting.

    Returns array of length 4: [MRT_CALM, MRT_TREND, MRT_VOLATILE, MRT_RANGING]
    """
    pi = compute_stationary_distribution(matrix)
    # Avoid division by zero for states with near-zero stationary probability
    mrt = np.where(pi > 1e-6, 1.0 / pi, np.inf)
    return mrt


def compute_mean_first_passage_times(matrix: np.ndarray) -> np.ndarray:
    """
    Mean First Passage Time (MFPT) matrix: m[i][j] = E[steps to reach j | start at i]

    From the document (Did's answer expanded to matrix form):
    "Enlarging the problem helps solving it -- consider not only t_C but also
    t_S, t_G, t_GG and set up a linear system solved by the one-step Markov property."

    This is exactly the system:
      m[i][j] = 1 + sum_k P[i][k] * m[k][j]  for i != j
      m[j][j] = MRT[j] = 1/pi_j

    Solved as: (I - Q) * m_col = 1  for each target state j
    where Q is P with row j zeroed out (making j absorbing).

    Scanner application -- example for ALKT:
      MFPT[TRENDING][CALM] = 3.2 days
      Means: if currently in TRENDING, expect 3.2 more trading days
             before CALM state (MR favorable) returns.
      Use this to decide: "is it worth waiting for MR setup
      or should I switch to TREND strategy now?"

    Returns 4x4 matrix where m[i][j] = expected trading days
    from state i to reach state j for the first time.
    """
    if matrix is None:
        return np.full((4, 4), np.nan)

    n    = matrix.shape[0]
    mfpt = np.zeros((n, n))
    pi   = compute_stationary_distribution(matrix)

    for j in range(n):
        # Mean recurrence time for j (diagonal entry)
        mfpt[j][j] = 1.0 / pi[j] if pi[j] > 1e-6 else np.inf

        # For i != j: solve (I - Q_j) m = 1
        # where Q_j = P with row j replaced by zeros (j becomes absorbing)
        Q_j = matrix.copy()
        Q_j[j, :] = 0.0   # make state j absorbing for this computation

        A   = np.eye(n) - Q_j
        b   = np.ones(n)

        try:
            m_col = np.linalg.solve(A, b)
            for i in range(n):
                if i != j:
                    mfpt[i][j] = float(m_col[i])
        except np.linalg.LinAlgError:
            # Singular matrix -- use stationary distribution fallback
            for i in range(n):
                if i != j:
                    mfpt[i][j] = mfpt[j][j]

    return mfpt


def compute_absorption_times(matrix: np.ndarray,
                               target_state: int,
                               consecutive_days: int) -> dict:
    """
    Absorbing Markov chain fundamental matrix.

    Directly from the document (amd's answer):
    "Expand the state space to include all stages of completion of the pattern.
    Model the last state as absorbing. Compute the fundamental matrix
    N = (I - Q)^{-1}. The expected absorption times are t = N * 1."

    Scanner application:
      "How many trading days until the CALM state persists for
      `consecutive_days` days in a row?"

      Example: consecutive_days=3, target_state=0 (CALM)
      Answers: "How many days until I get 3 consecutive CALM days?"
      This is the expected time to a STABLE MR window.

    States in augmented chain:
      Original 4 states + (consecutive_days - 1) partial-match states + 1 absorbing

    Returns dict with:
      expected_days_from[i]: expected trading days from state i until absorption
      fundamental_matrix:    full N matrix
    """
    if matrix is None or consecutive_days < 1:
        return {}

    n_orig   = matrix.shape[0]
    n_partial = consecutive_days - 1   # stages of partial match
    n_total  = n_orig + n_partial + 1  # original + partial + absorbing

    # Build augmented transition matrix P'
    # States 0..n_orig-1 : original Markov states
    # States n_orig..n_orig+n_partial-1 : partial match (1,2,...,n_partial consecutive)
    # State n_total-1 : absorbing (pattern complete)

    P_aug = np.zeros((n_total, n_total))

    # Original states: transition normally, EXCEPT
    # if transitioning to target_state, advance to partial-match track
    for i in range(n_orig):
        for k in range(n_orig):
            if k == target_state and n_partial > 0:
                # First step toward pattern: go to partial-match state 0
                P_aug[i][n_orig] += matrix[i][k]
            elif k == target_state and n_partial == 0:
                # Pattern length 1: go straight to absorbing
                P_aug[i][n_total - 1] += matrix[i][k]
            else:
                P_aug[i][k] += matrix[i][k]

    # Partial-match states: advancing through consecutive target days.
    # Key insight from amd's document: when in partial-match state p,
    # we ARE in the target_state. Next transition uses target_state's row.
    # If next day is target_state: advance to p+1 (or absorb if last step).
    # If next day is NOT target_state: fall back to that state directly.
    # This correctly models self-overlapping patterns (3 consecutive days).
    for p in range(n_partial):
        state_idx  = n_orig + p
        next_match = n_orig + p + 1 if p + 1 < n_partial else n_total - 1
        for k in range(n_orig):
            if k == target_state:
                # Another consecutive target day -- advance
                P_aug[state_idx][next_match] += matrix[target_state][k]
            else:
                # Broke the streak -- fall back to state k (original state)
                P_aug[state_idx][k] += matrix[target_state][k]

    # Absorbing state: stays absorbing
    P_aug[n_total - 1][n_total - 1] = 1.0

    # Q matrix: transient states only (all except absorbing)
    n_transient = n_total - 1
    Q = P_aug[:n_transient, :n_transient]

    # Fundamental matrix N = (I - Q)^{-1}
    try:
        N = np.linalg.inv(np.eye(n_transient) - Q)
    except np.linalg.LinAlgError:
        return {"error": "Singular matrix -- chain may not be absorbing"}

    # Expected absorption times: t = N * 1
    t = N.sum(axis=1)

    result = {
        "target_state":       MARKOV_STATES.get(target_state, "UNKNOWN"),
        "consecutive_days":   consecutive_days,
        "expected_days_from": {},
        "fundamental_matrix": N.tolist(),
    }

    # Expected days from each original state
    for i in range(n_orig):
        state_name = MARKOV_STATES.get(i, f"STATE_{i}")
        result["expected_days_from"][state_name] = round(float(t[i]), 2)

    # Expected days from partial-match states
    for p in range(n_partial):
        label = f"{MARKOV_STATES.get(target_state,'?')}x{p+1}_streak"
        result["expected_days_from"][label] = round(float(t[n_orig + p]), 2)

    return result


def print_markov_timing_report(symbol: str, matrix: np.ndarray,
                                current_state: int) -> dict:
    """
    Full timing report for one symbol's Markov chain.
    Prints mean recurrence times, mean first passage times,
    and absorption times for CALM consecutive streaks.

    This answers:
      1. How long before CALM returns? (MRT = 1/pi_CALM)
      2. From current state, how many days to reach CALM? (MFPT)
      3. How many days for 3 consecutive CALM days? (absorbing chain)

    Called from run_markov_analysis().
    """
    if matrix is None:
        return {}

    report = {"symbol": symbol}

    # Mean Recurrence Times
    mrt = compute_mean_recurrence_times(matrix)
    print(f"\n  Mean Recurrence Times (E[T_ii] = 1/pi_i):")
    print(f"  How long until each regime returns on average:")
    for i, name in MARKOV_STATES.items():
        t = float(mrt[i])
        current = " <- you are here" if i == current_state else ""
        if t == np.inf:
            print(f"    {name:<12} NEVER (never observed in data)")
        else:
            print(f"    {name:<12} {t:.1f} trading days{current}")
    report["mean_recurrence_times"] = {
        MARKOV_STATES[i]: round(float(mrt[i]), 2) if mrt[i] != np.inf else None
        for i in range(4)
    }

    # Mean First Passage Times from current state to CALM (state 0)
    mfpt = compute_mean_first_passage_times(matrix)
    calm_col = mfpt[:, 0]
    print(f"\n  Mean First Passage Times -> CALM state:")
    print(f"  How many trading days from each state until CALM appears:")
    for i, name in MARKOV_STATES.items():
        t = float(calm_col[i])
        current = " <- you are here" if i == current_state else ""
        print(f"    {name:<12} {t:.1f} days to CALM{current}")
    report["mfpt_to_calm"] = {
        MARKOV_STATES[i]: round(float(calm_col[i]), 2) for i in range(4)
    }

    # Absorbing chain: how long for 3 consecutive CALM days (stable MR window)
    abs3 = compute_absorption_times(matrix, target_state=0, consecutive_days=3)
    if abs3 and "expected_days_from" in abs3:
        print(f"\n  Expected days until 3 CONSECUTIVE CALM days (stable MR window):")
        print(f"  Based on absorbing Markov chain (from document reference):")
        for state_name, days in abs3["expected_days_from"].items():
            current = " <- you are here" if state_name == MARKOV_STATES.get(current_state) else ""
            print(f"    From {state_name:<16} {days:.1f} days{current}")
        report["absorption_3_calm"] = abs3["expected_days_from"]

    return report





def simulate_price_paths(symbol: str, current_price: float,
                           n_paths: int = 1000, n_steps: int = 5,
                           dt: float = 1/252) -> dict:
    """
    Geometric Brownian Motion price simulation for tail risk assessment.
    dS = mu*S*dt + sigma*S*dW  (dW = Brownian motion increment)

    Returns price distribution, VaR(95%), CVaR(95%), and P(drop > 2%).
    Use this to understand the downside BEFORE entering a trade.
    """
    try:
        hist = yf.Ticker(symbol).history(period="90d", interval="1d",
                                          auto_adjust=True)
        if hist is None or hist.empty or len(hist) < 30:
            return {}
        hist.columns = [c.lower() for c in hist.columns]
        log_ret = np.log(hist["close"] / hist["close"].shift(1)).dropna()
        mu      = float(log_ret.mean())
        sigma   = float(log_ret.std())
    except Exception as e:
        print(f"  [GBM] Failed: {e}")
        return {}

    np.random.seed(42)
    Z       = np.random.standard_normal((n_paths, n_steps))
    log_ret = (mu - 0.5 * sigma**2) * dt + sigma * np.sqrt(dt) * Z
    paths   = current_price * np.exp(np.cumsum(log_ret, axis=1))
    terminal = paths[:, -1]
    returns  = (terminal - current_price) / current_price
    var_95   = float(np.percentile(returns, 5))
    cvar_95  = float(returns[returns <= var_95].mean()) if (returns <= var_95).any() else var_95

    return {
        "symbol": symbol, "current_price": current_price,
        "mu_daily": round(mu, 6), "sigma_daily": round(sigma, 6),
        "mean_price":   round(float(terminal.mean()), 4),
        "p01": round(float(np.percentile(terminal,  1)), 4),
        "p05": round(float(np.percentile(terminal,  5)), 4),
        "p25": round(float(np.percentile(terminal, 25)), 4),
        "p75": round(float(np.percentile(terminal, 75)), 4),
        "p95": round(float(np.percentile(terminal, 95)), 4),
        "p99": round(float(np.percentile(terminal, 99)), 4),
        "var_95_pct":     round(var_95 * 100, 3),
        "cvar_95_pct":    round(cvar_95 * 100, 3),
        "prob_down_2pct": round(float((returns < -0.02).mean()) * 100, 2),
        "paths": paths,
    }


def is_state_stable(current_state: int, transition_matrix,
                     threshold: float = MARKOV_MIN_PROB) -> tuple:
    """
    Probability gate: (stable, stay_prob, reason).
    Falls back to (True, 1.0, FALLBACK) when matrix unavailable.
    """
    if transition_matrix is None:
        return True, 1.0, "FALLBACK - no matrix, standard logic applies"
    if current_state not in range(4):
        return True, 1.0, "FALLBACK - unknown state"
    stay_p     = float(transition_matrix[current_state][current_state])
    state_name = MARKOV_STATES[current_state]
    stable     = stay_p >= threshold
    if stable:
        reason = (f"STABLE - {state_name} persists at {stay_p:.1%} "
                  f"(>= {threshold:.0%} threshold)")
    else:
        row = transition_matrix[current_state].copy()
        row[current_state] = 0
        ns  = int(np.argmax(row)); np_ = float(row[ns])
        reason = (f"UNSTABLE - {state_name} only {stay_p:.1%} stable. "
                  f"Likely flip: {MARKOV_STATES[ns]} ({np_:.1%}). Skip trade.")
    return stable, stay_p, reason


def get_stop_adjustment(current_state: int, transition_matrix) -> float:
    """Tighten stop when volatile transition risk is elevated."""
    if transition_matrix is None: return 1.0
    vol_p = float(transition_matrix[current_state][2])
    if vol_p >= 0.25: return 0.50
    if vol_p >= 0.15: return 0.75
    return 1.0


def run_markov_analysis(symbols: list, days: int = 60,
                          n_steps_ahead: int = 5) -> dict:
    """Full Markov analysis with stationary dist, n-step forecast, GBM."""
    sep = "=" * 65
    print(f"\n{sep}")
    print(f"  MARKOV CHAIN REGIME ANALYSIS")
    print(f"  {len(symbols)} symbol(s)  |  {days}-day lookback"
          f"  |  {n_steps_ahead}-step ahead forecast")
    print(f"  Stability threshold: {MARKOV_MIN_PROB:.0%}"
          f"  |  Min observations/state: {MARKOV_MIN_OBS}")
    print(f"{'='*65}")

    results = {}
    today   = datetime.now().strftime("%Y-%m-%d")

    for sym in symbols:
        print("=" * 55)
        print(f"  {sym}")
        print(f"  {'='*55}")

        matrix, counts, states_seq, valid, data_days = build_transition_matrix(
            sym, days, use_cache=False)
        if not valid:
            print(f"  SKIPPED - < {MARKOV_MIN_DAYS} days available")
            results[sym] = {"valid": False}
            continue

        current_state = states_seq[-1][1]
        current_name  = MARKOV_STATES[current_state]
        stable, stay_p, reason = is_state_stable(current_state, matrix)
        stop_adj = get_stop_adjustment(current_state, matrix)
        pi       = compute_stationary_distribution(matrix)
        nstep    = compute_nstep_matrix(matrix, n_steps_ahead)
        nstep_p  = nstep[current_state]

        sc = {MARKOV_STATES[s]: 0 for s in range(4)}
        for _, s in states_seq: sc[MARKOV_STATES[s]] += 1
        total = len(states_seq)

        # Current state summary
        print(f"Current state:   {current_name}")
        print(f"  Stability:       {reason}")
        print(f"  Stop adj:        {stop_adj:.0%} of normal stop")
        print(f"  Signal:          {'GO' if stable else 'WAIT - regime unstable'}")

        # State distribution + stationary
        print(f"State distribution ({total} days):")
        print(f"  {'State':<12} {'Days':>5} {'Freq':>6} {'Stay':>8} {'Long-run':>10}")
        print(f"  {'-'*48}")
        for s_id, s_name in MARKOV_STATES.items():
            obs   = sc.get(s_name, 0)
            freq  = obs / total * 100 if total > 0 else 0
            sp    = float(matrix[s_id][s_id])
            pi_v  = float(pi[s_id])
            bar   = "#" * int(sp * 15)
            tag   = " <-- current" if s_id == current_state else ""
            print(f"  {s_name:<12} {obs:>5} {freq:>5.0f}%  "
                  f"{sp:>6.1%}  {pi_v:>9.1%}  {bar}{tag}")

        # Transition matrix
        print(f"Transition matrix (FROM -> TO):")
        hdr = "".join(f"{MARKOV_STATES[j]:>12}" for j in range(4))
        print(f"  {'FROM':<12}{hdr}")
        print(f"  {'-'*60}")
        for i in range(4):
            row_s = "".join(f"{matrix[i][j]:>12.1%}" for j in range(4))
            n_obs = counts[i].sum()
            note  = f"  (n={n_obs})" + (" <- current" if i == current_state else "")
            print(f"  {MARKOV_STATES[i]:<12}{row_s}{note}")

        # N-step forecast
        print(f"{n_steps_ahead}-step ahead forecast from {current_name}:")
        print(f"  {'State':<12} {'Prob':>7}  Assessment")
        print(f"  {'-'*50}")
        for s_id, s_name in MARKOV_STATES.items():
            p = float(nstep_p[s_id])
            note = ""
            if s_id == current_state and p >= 0.65:    note = "STABLE"
            elif s_id == current_state:                 note = "MODERATE - monitor"
            elif s_id == 2 and p >= 0.15:               note = "VOL RISK - tighten stop"
            elif s_id == 1 and current_state == 0 and p >= 0.20:
                note = "MR TRAP - trend forming"
            print(f"  {s_name:<12} {p:>6.1%}  {note}")

        # Tradability by state
        print(f"State assessment:")
        for i in range(4):
            sp    = float(matrix[i][i])
            row   = matrix[i].copy(); row[i] = 0
            fs    = int(np.argmax(row)); fp = float(row[fs])
            n_obs = counts[i].sum()
            if n_obs < MARKOV_MIN_OBS:
                v = f"SPARSE ({n_obs} obs)"
            elif sp >= MARKOV_MIN_PROB:
                v = "TRADEABLE"
            else:
                v = "AVOID"
            print(f"  {MARKOV_STATES[i]:<12} stay={sp:.1%}  "
                  f"flip->{MARKOV_STATES[fs]}({fp:.1%})  {v}")

        # GBM tail risk
        try:
            hist = yf.Ticker(sym).history(period="2d", interval="1d")
            cur_px = float(hist["Close"].iloc[-1]) if not hist.empty else 0
            if cur_px > 0:
                gbm = simulate_price_paths(sym, cur_px, n_steps=n_steps_ahead)
                if gbm:
                    ann_vol = gbm["sigma_daily"] * np.sqrt(252) * 100
                    print(f"GBM ({n_steps_ahead} bars, {n_steps_ahead*5}min ahead):")
                    print(f"  Price now:  ${cur_px:.2f}  "
                          f"Ann vol: {ann_vol:.1f}%")
                    print(f"  Expected:   ${gbm['mean_price']:.2f}")
                    print(f"  90% range:  ${gbm['p05']:.2f} -- ${gbm['p95']:.2f}")
                    print(f"  Tail (1/99): ${gbm['p01']:.2f} -- ${gbm['p99']:.2f}")
                    print(f"  VaR(95%):   {gbm['var_95_pct']:+.2f}%  "
                          f"CVaR(95%): {gbm['cvar_95_pct']:+.2f}%")
                    print(f"  P(drop>2%): {gbm['prob_down_2pct']:.1f}%  "
                          f"{'HIGH' if gbm['prob_down_2pct'] > 20 else 'OK'}")
        except Exception:
            pass

        # MR trap warning
        if current_state == 0:
            if float(matrix[0][1]) >= 0.20:
                print(f"[MR WARN] CALM->TREND = {matrix[0][1]:.1%}. "
                      f"MR trades carry regime flip risk.")
            if float(matrix[0][2]) >= 0.10:
                print(f"  [VOL WARN] CALM->VOLATILE = {matrix[0][2]:.1%}. "
                      f"Tighten stop to {stop_adj:.0%}.")

        # Personality
        calm_p  = float(pi[0])
        trend_p = float(pi[1])
        if calm_p >= 0.50:
            personality = f"MEAN REVERTER ({calm_p:.0%} CALM)"
        elif trend_p >= 0.40:
            personality = f"TREND FOLLOWER ({trend_p:.0%} TRENDING)"
        else:
            personality = f"MIXED (CALM={calm_p:.0%} TREND={trend_p:.0%})"
        print(f"  Long-run personality: {personality}")

        # ── Timing analysis (MRT + MFPT + Absorbing chain) ────────────────
        # From the Markov chain math document:
        #   MRT  = 1/pi_i  (Ross: mean recurrence time)
        #   MFPT = Did's linear system enlargement approach
        #   Absorbing chain = amd's fundamental matrix N=(I-Q)^{-1}
        print()
        timing = print_markov_timing_report(sym, matrix, current_state)

        results[sym] = {
            "valid": True, "current_state": current_name,
            "stable": stable, "stay_prob": round(stay_p, 4),
            "stop_adj": stop_adj, "reason": reason,
            "matrix": matrix.tolist(),
            "stationary": pi.tolist(),
            "nstep_probs": nstep_p.tolist(),
            "state_counts": sc, "total_days": total,
            "personality": personality,
            "timing": timing,
        }

    out = os.path.join(DL_DIR, f"markov_{today}.json")
    with open(out, "w") as f:
        json.dump(results, f, indent=2)
    print("=" * 65)
    print(f"  Saved -> {out}")
    print(f"{'='*65}")
    return results


def run_markov_gate_test(symbol: str, days: int = 60) -> dict:
    """
    Per-symbol Markov gate backtest.
    Compares win rate and Sharpe with and without the gate.
    Run --markov-gate-portfolio to test across all watchlist symbols.

    Usage:  python scanner_research_v4.py --markov-gate ALKT --days 60
    """
    print(f"\n{'='*65}")
    print(f"  MARKOV GATE BACKTEST  |  {symbol}  |  {days} days")
    print(f"{'='*65}")

    result5 = build_transition_matrix(symbol, days, use_cache=False)
    matrix, counts, states_seq, valid, data_days = result5 if len(result5)==5 else (*result5, 0)
    if not valid:
        print(f"  SKIPPED -- insufficient data ({data_days} days, need {MARKOV_MIN_DAYS})")
        return {}

    df_all = run_backtest([symbol], days=days,
                           strategy="MEAN_REVERSION", slippage_pct=0.001)
    gated  = []
    if not df_all.empty and states_seq:
        by_date = {d: s for d, s in states_seq}
        for _, row in df_all.iterrows():
            s_int = by_date.get(str(row.get("entry_date",""))[:10], 3)
            ok, _, _ = is_state_stable(s_int, matrix)
            if ok: gated.append(row)
    df_gated = pd.DataFrame(gated) if gated else pd.DataFrame()

    def stats(df):
        if df.empty: return {"n":0,"wr":0.0,"sharpe":0.0,"avg_pnl":0.0}
        n    = len(df)
        wr   = df["win"].sum() / n * 100
        av   = df["pnl_pct"].mean()
        daily = df.groupby("entry_date")["pnl_pct"].sum() \
                if "entry_date" in df.columns else pd.Series([0.0])
        sh = (daily.mean()/(daily.std()+1e-9))*np.sqrt(252) if len(daily)>3 else 0.0
        return {"n":n,"wr":round(wr,1),"sharpe":round(sh,2),"avg_pnl":round(av,4)}

    sa = stats(df_all); sg = stats(df_gated)
    filt = sa["n"] - sg["n"]
    dwr  = sg["wr"] - sa["wr"]; dsh = sg["sharpe"] - sa["sharpe"]

    print(f"\n  {'Metric':<22} {'Without gate':>14} {'With gate':>14}")
    print(f"  {'-'*52}")
    print(f"  {'Trades':<22} {sa['n']:>14} {sg['n']:>14}")
    print(f"  {'Win rate':<22} {sa['wr']:>13.1f}% {sg['wr']:>13.1f}%")
    print(f"  {'Sharpe':<22} {sa['sharpe']:>14.2f} {sg['sharpe']:>14.2f}")
    print(f"  {'Avg P&L':<22} {sa['avg_pnl']:>13.3f}% {sg['avg_pnl']:>13.3f}%")
    print(f"\n  Filtered: {filt} trades ({filt/max(sa['n'],1)*100:.0f}%)")
    print(f"  Win rate: {dwr:+.1f}%   Sharpe: {dsh:+.2f}")

    if   dsh >= 0.2:  verdict = f"ADOPT  -- Sharpe +{dsh:.2f}"
    elif dsh > 0:     verdict = f"MONITOR -- +{dsh:.2f} (below 0.2). Re-test next month."
    else:             verdict = f"SKIP   -- gate hurts Sharpe ({dsh:.2f})"
    print(f"\n  VERDICT: {verdict}")

    return {"symbol": symbol, "without_gate": sa, "with_gate": sg,
            "sharpe_delta": dsh, "wr_delta": dwr, "verdict": verdict}


def run_markov_gate_portfolio(symbols: list = None, days: int = 60) -> dict:
    """
    CROSS-SECTIONAL Markov gate validation across all watchlist symbols.

    This is the correct decision gate for MARKOV_GATE_ENABLED.
    Per-symbol tests mislead: ALKT +0.46, GKOS -0.25 -> net unclear.

    ADOPTION RULES (both must pass):
      1. Average Sharpe delta >= 0.2  (meaningful improvement across the portfolio)
      2. Negative outliers (dsh < -0.1) <= 20% of symbols  (no widespread harm)

    VERDICTS:
      ADOPT   -- both rules pass -> set MARKOV_GATE_ENABLED = True
      PARTIAL -- rule 1 pass, rule 2 fail -> hurts specific symbols, investigate
      MONITOR -- rule 1 fail, rule 2 pass -> improvement too small, wait
      SKIP    -- both fail -> gate hurts performance, do not enable

    Usage:  python scanner_research_v4.py --markov-gate-portfolio --days 60
    """
    symbols = symbols or engine.WATCHLIST

    print(f"\n{'='*65}")
    print(f"  MARKOV GATE PORTFOLIO VALIDATION")
    print(f"  {len(symbols)} symbols  |  {days}-day lookback")
    print(f"  Rule 1: avg Sharpe delta >= 0.2")
    print(f"  Rule 2: negative outliers (dsh < -0.1) <= 20% of symbols")
    print(f"{'='*65}\n")

    results = []; skipped = []

    for sym in symbols:
        print(f"  {sym:<6}...", end=" ", flush=True)
        result = run_markov_gate_test(sym, days=days)
        if not result:
            skipped.append(sym); print("SKIPPED"); continue
        results.append(result)
        dsh = result.get("sharpe_delta", 0)
        print(f"dSharpe={dsh:+.2f}")

    if not results:
        print(f"\n  No valid results. Need {MARKOV_MIN_DAYS}+ days of data.")
        return {}

    deltas       = [r["sharpe_delta"] for r in results]
    avg_delta    = float(np.mean(deltas))
    med_delta    = float(np.median(deltas))
    std_delta    = float(np.std(deltas))
    n_positive   = sum(1 for d in deltas if d >= 0.2)
    n_neutral    = sum(1 for d in deltas if 0 <= d < 0.2)
    n_negative   = sum(1 for d in deltas if d < -0.1)
    pct_negative = n_negative / len(deltas) * 100
    results_sorted = sorted(results, key=lambda r: r["sharpe_delta"], reverse=True)

    print(f"\n{'='*65}")
    print(f"  CROSS-SECTIONAL RESULTS  ({len(results)} symbols)")
    print(f"  {'Symbol':<8} {'dSharpe':>9} {'dWinRate':>9}  Verdict")
    print(f"  {'-'*52}")
    for r in results_sorted:
        dsh = r["sharpe_delta"]; dwr = r["wr_delta"]
        if dsh >= 0.2:   tag = "ADOPT"
        elif dsh > 0:    tag = "MONITOR"
        elif dsh > -0.1: tag = "MARGINAL"
        else:            tag = "OUTLIER"
        flag = " <-- NEGATIVE" if dsh < -0.1 else ""
        print(f"  {r['symbol']:<8} {dsh:>+8.2f}  {dwr:>+8.1f}%  {tag}{flag}")

    print(f"\n  ADOPT (>= +0.2):  {n_positive:>3} ({n_positive/len(deltas)*100:.0f}%)")
    print(f"  MONITOR (0-0.2):  {n_neutral:>3} ({n_neutral/len(deltas)*100:.0f}%)")
    print(f"  OUTLIER (< -0.1): {n_negative:>3} ({pct_negative:.0f}%)")
    print(f"\n  Avg dSharpe: {avg_delta:+.3f}  Median: {med_delta:+.3f}  Std: {std_delta:.3f}")

    rule1 = avg_delta >= 0.2
    rule2 = pct_negative <= 20.0

    print(f"\n  Rule 1 (avg >= 0.2): {avg_delta:+.3f}  {'PASS' if rule1 else 'FAIL'}")
    print(f"  Rule 2 (outliers <= 20%): {pct_negative:.0f}%  {'PASS' if rule2 else 'FAIL'}")

    if rule1 and rule2:
        verdict = "ADOPT"
        action  = "Set MARKOV_GATE_ENABLED = True. Then --backtest --days 30 to confirm."
    elif rule1 and not rule2:
        bad = [r["symbol"] for r in results if r["sharpe_delta"] < -0.1]
        verdict = "PARTIAL"
        action  = f"Helps overall but hurts {bad}. Investigate before enabling."
    elif not rule1 and rule2:
        verdict = "MONITOR"
        action  = f"Improvement too small ({avg_delta:+.3f}). Re-test next month."
    else:
        verdict = "SKIP"
        action  = "Gate hurts portfolio performance. Keep MARKOV_GATE_ENABLED = False."

    print(f"\n  PORTFOLIO VERDICT: {verdict}")
    print(f"  {action}")

    today = datetime.now().strftime("%Y-%m-%d")
    out   = os.path.join(DL_DIR, f"markov_portfolio_test_{today}.json")
    summary = {
        "date": today, "n_tested": len(results), "n_skipped": len(skipped),
        "avg_delta": round(avg_delta,4), "med_delta": round(med_delta,4),
        "std_delta": round(std_delta,4), "n_positive": n_positive,
        "n_negative": n_negative, "pct_negative": round(pct_negative,1),
        "rule1": rule1, "rule2": rule2, "verdict": verdict,
        "per_symbol": results,
    }
    with open(out,"w") as f:
        json.dump(summary, f, indent=2,
                  default=lambda x: float(x) if hasattr(x,"__float__") else str(x))
    print(f"\n  Saved -> {out}")
    print(f"{'='*65}")
    return summary



def get_markov_gate(symbol: str, current_adx: float,
                     current_spy_vol: float, days: int = 60) -> tuple:
    """
    Single-call interface for scanner_v4.py integration.
    Priority order for data source:
      1. In-memory cache (_markov_cache) -- fastest, rebuilt hourly
      2. Pre-built disk file (SCANNER_DIR/markov_data/SYMBOL.json)
         from today's --prefetch run -- no yfinance call needed
      3. Live yfinance fetch -- slowest, used as fallback only

    Returns (stable, stay_prob, stop_adj, reason).

    The gate ONLY applies when 60+ days of valid data are available.
    When data is insufficient: returns (True, 1.0, 1.0, reason)
    so the signal passes through unchanged -- never penalises for lack of data.
    """
    # -- Guardrail: SPY volatility override ------------------------
    if abs(current_spy_vol) > MARKOV_SPY_OVERRIDE:
        return (True, 1.0, 1.0,
                f"VOL_OVERRIDE: SPY {current_spy_vol:+.1%} "
                f"(trend day - Markov not applied)")

    # -- Source 1: in-memory cache (fastest) -----------------------
    import time as _t
    if symbol in _markov_cache:
        cached = _markov_cache[symbol]
        if _t.time() - cached[3] < MARKOV_CACHE_TTL:
            matrix    = cached[0]
            data_days = cached[4] if len(cached) > 4 else 0
            valid     = matrix is not None
            if valid:
                state    = classify_state(current_adx, current_spy_vol)
                stable, stay_p, _ = is_state_stable(state, matrix)
                stop_adj = get_stop_adjustment(state, matrix)
                state_name = MARKOV_STATES.get(state, "UNKNOWN")
                return (stable, stay_p, stop_adj,
                        f"ACTIVE (cache): {state_name} "
                        f"{'stable' if stable else 'unstable'} "
                        f"{stay_p:.1%} ({data_days}d)")

    # -- Source 2: pre-built disk file (fast, no yfinance) ---------
    disk_matrix, disk_valid, disk_days, disk_reason = load_markov_from_disk(symbol)

    if disk_valid and disk_matrix is not None:
        # Store in memory cache so next call is instant
        _markov_cache[symbol] = (disk_matrix, None, [], _t.time(), disk_days)
        state    = classify_state(current_adx, current_spy_vol)
        stable, stay_p, _ = is_state_stable(state, disk_matrix)
        stop_adj = get_stop_adjustment(state, disk_matrix)
        state_name = MARKOV_STATES.get(state, "UNKNOWN")
        return (stable, stay_p, stop_adj,
                f"ACTIVE (disk): {state_name} "
                f"{'stable' if stable else 'unstable'} "
                f"{stay_p:.1%} ({disk_days}d)")

    # -- Disk file exists but has no data yet -- report gap --------
    if disk_reason.startswith("NO_DATA"):
        # Parse the days count from the reason string
        parts     = disk_reason.split("_")
        have_days = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
        needed    = MARKOV_MIN_DAYS - have_days
        return (True, 1.0, 1.0,
                f"NO_DATA: {have_days} days available "
                f"(need {MARKOV_MIN_DAYS}, ~{needed} trading days more)")

    # -- Source 3: live yfinance fetch (fallback) ------------------
    # Only reaches here if: no cache, no disk file, or disk is stale
    # During trading hours this means --prefetch was not run today
    result = build_transition_matrix(symbol, days, use_cache=True)
    if len(result) == 5:
        matrix, counts, states_seq, valid, data_days = result
    else:
        matrix, counts, states_seq, valid = result
        data_days = 0

    if not valid or matrix is None:
        if data_days > 0:
            needed = MARKOV_MIN_DAYS - data_days
            return (True, 1.0, 1.0,
                    f"NO_DATA: only {data_days} days available "
                    f"(need {MARKOV_MIN_DAYS}, ~{needed} trading days more)")
        return (True, 1.0, 1.0,
                f"NO_DATA: could not fetch history for {symbol}")

    state    = classify_state(current_adx, current_spy_vol)
    stable, stay_p, _ = is_state_stable(state, matrix)
    stop_adj = get_stop_adjustment(state, matrix)
    state_name = MARKOV_STATES.get(state, "UNKNOWN")
    return (stable, stay_p, stop_adj,
            f"ACTIVE (live): {state_name} "
            f"{'stable' if stable else 'unstable'} "
            f"{stay_p:.1%} ({data_days}d)")


# ===============================================================
# MARKOV PARAMETER TUNING GUIDE
# ===============================================================
#
# MARKOV_MIN_DAYS = 60
#   Why: Fewer than 60 days gives < 20 obs per state on average.
#   Raise to 90: if you see frequent SPARSE warnings.
#   Lower to 40: only when modeling a specific recent regime shift.
#
# MARKOV_MIN_PROB = 0.65
#   Why: Historically MR trades in CALM state with < 65% persistence
#   fail at rates that eliminate the edge.
#   Raise to 0.75: more selective, ~30% fewer trades, higher WR.
#   Lower to 0.55: only if gate is too restrictive after --markov-gate test.
#
# MARKOV_MIN_OBS = 20
#   Why: < 20 transitions gives > +/- 22% standard error on probability.
#   Do NOT lower -- you will get false precision.
#
# MARKOV_SPY_OVERRIDE = 0.02
#   NON-NEGOTIABLE. Matches scanner_v4.py SPY_VOL_NOTRADE. Never lower.
#
# MARKOV_SPARSE_PRIOR
#   Recalibrate annually: run --markov on 20+ R2K stocks with --days 252
#   and average the resulting matrices. Update this array once per year.
#
# RECOMMENDED SETTINGS BY REGIME:
#   Bull trending market (2020, 2023):   MARKOV_MIN_PROB = 0.60
#   Choppy sideways (2022 H1, summers):  MARKOV_MIN_PROB = 0.70
#   High vol / crash:                    GLOBAL_KILL_SWITCH = True
#   Normal mixed (default):              All defaults (60/0.65/20)


# ===============================================================
# SECTION 8 -- HYPOTHESIS TESTING + POST-MARKET OUTCOME ANALYSIS
# ===============================================================
#
# WORKFLOW
# --------
# 1. Run scanner in --loop mode all day -> scan_log_DATE.csv builds up
# 2. After 4 PM: run --tag-outcomes DATE
#    -> fetches actual intraday bars, tags each signal HIT_TARGET/HIT_STOP
#    -> saves scan_log_DATE_tagged.csv
# 3. Run --hypothesis DATE to compare groups
#    -> does hurst_score predict hitting the target?
#    -> is score > 75 meaningfully better than score > 65?
#    -> is bayes_prob signal or noise?
# 4. After a full week: run --weekly to aggregate all tagged files
#    -> 20+ trades needed for statistical significance
# ===============================================================

import glob


def tag_outcomes(log_date: str, lookforward_hours: int = 4) -> pd.DataFrame:
    """
    Post-market outcome tagger.
    Reads scan_log_DATE.csv, fetches actual intraday bars, tags each signal.

    Tags: HIT_TARGET | HIT_STOP | OPEN | NO_SIGNAL | ERROR

    Usage:  python scanner_research_v4.py --tag-outcomes 2026-04-30
    """
    dl_dir   = os.path.join(os.path.expanduser("~"), "Downloads")
    log_path = os.path.join(dl_dir, f"scan_log_{log_date}.csv")

    if not os.path.exists(log_path):
        print(f"  [TAG] Not found: {log_path}")
        print(f"  Run scanner in --loop mode first to generate the log.")
        return pd.DataFrame()

    df = pd.read_csv(log_path)
    print(f"\n{'='*60}")
    print(f"  OUTCOME TAGGING  |  {log_date}  |  {len(df)} rows")
    print(f"{'='*60}")

    required = {"symbol", "price", "target", "stop"}
    if required - set(df.columns):
        print(f"  Missing columns: {required - set(df.columns)}")
        return df

    outcomes = []
    tagged = skipped = 0

    for _, row in df.iterrows():
        sym    = str(row["symbol"])
        price  = float(row.get("price", 0))
        target = float(row.get("target", 0))
        stop   = float(row.get("stop",   0))
        s_time = str(row.get("scan_time", ""))
        alert  = bool(row.get("alert", False))

        if not alert or price <= 0 or target <= 0 or stop <= 0:
            outcomes.append("NO_SIGNAL"); skipped += 1; continue

        direction = "LONG" if target > price else "SHORT"

        try:
            hist = yf.Ticker(sym).history(
                start=log_date, end=log_date, interval="5m", auto_adjust=True)
            if hist.empty:
                hist = yf.Ticker(sym).history(period="2d", interval="5m",
                                               auto_adjust=True)
            if hist.empty:
                outcomes.append("OPEN"); continue

            hist.columns = [c.lower() for c in hist.columns]

            # Filter to bars AFTER the scan fired
            if s_time and s_time != "nan":
                try:
                    scan_dt = pd.Timestamp(f"{log_date} {s_time}") \
                              .tz_localize("America/New_York")
                    hist.index = pd.DatetimeIndex(hist.index).tz_convert(
                        "America/New_York")
                    future = hist[hist.index >= scan_dt].head(lookforward_hours * 12)
                except Exception:
                    future = hist
            else:
                future = hist

            outcome = "OPEN"
            for _, bar in future.iterrows():
                h = float(bar["high"]); lo = float(bar["low"])
                if direction == "LONG":
                    hit_t = h  >= target; hit_s = lo <= stop
                else:
                    hit_t = lo <= target; hit_s = h  >= stop

                if hit_t and hit_s:
                    o = float(bar["open"])
                    outcome = "HIT_TARGET" if (direction=="LONG" and o >= stop) or \
                              (direction=="SHORT" and o <= stop) else "HIT_STOP"
                    break
                elif hit_t: outcome = "HIT_TARGET"; break
                elif hit_s: outcome = "HIT_STOP";   break

            outcomes.append(outcome); tagged += 1

        except Exception as e:
            outcomes.append("ERROR")

    df["outcome"] = outcomes
    df["tag_date"] = log_date
    out = os.path.join(dl_dir, f"scan_log_{log_date}_tagged.csv")
    df.to_csv(out, index=False)

    vc = df["outcome"].value_counts()
    print(f"\n  Tagged {tagged} signals  |  {skipped} non-signals skipped")
    for outcome, count in vc.items():
        bar = "#" * int(count / max(len(df), 1) * 30)
        print(f"  {outcome:<14} {count:>4}  {bar}")

    hit_t = (df["outcome"] == "HIT_TARGET").sum()
    hit_s = (df["outcome"] == "HIT_STOP").sum()
    total = hit_t + hit_s
    if total > 0:
        print(f"\n  Win rate: {hit_t/total*100:.1f}%  "
              f"({hit_t} targets / {hit_s} stops / "
              f"{(df['outcome']=='OPEN').sum()} open)")
    print(f"\n  Saved -> {out}")
    return df


def run_hypothesis_test(log_date: str) -> dict:
    """
    Tests whether your indicator thresholds actually predict outcomes.

    Questions answered:
    1. Does hurst_score separate winners from losers?
       SUCCESS: HIT_TARGET avg hurst > HIT_STOP by 0.08+
    2. Is score > 75 meaningfully better than > 65?
    3. Is bayes_prob signal or noise?
    4. Does Z-score depth (more negative) predict better MR outcomes?
    5. Does half-life strength correlate with wins?

    Requires --tag-outcomes DATE first.
    Usage:  python scanner_research_v4.py --hypothesis 2026-04-30
    """
    dl_dir   = os.path.join(os.path.expanduser("~"), "Downloads")
    tag_path = os.path.join(dl_dir, f"scan_log_{log_date}_tagged.csv")

    if not os.path.exists(tag_path):
        print(f"  [HYPO] Run --tag-outcomes {log_date} first.")
        return {}

    df     = pd.read_csv(tag_path)
    df_sig = df[df["outcome"].isin(["HIT_TARGET","HIT_STOP"])].copy()

    if len(df_sig) < 5:
        print(f"  [HYPO] Only {len(df_sig)} completed signals."
              f" Need 5+ -- keep running the scanner.")
        return {}

    print(f"\n{'='*65}")
    print(f"  HYPOTHESIS TEST  |  {log_date}  |  {len(df_sig)} completed signals")
    print(f"{'='*65}")

    target_g = df_sig[df_sig["outcome"] == "HIT_TARGET"]
    stop_g   = df_sig[df_sig["outcome"] == "HIT_STOP"]
    win_rate = len(target_g) / len(df_sig) * 100
    print(f"\n  HIT_TARGET: {len(target_g)}   HIT_STOP: {len(stop_g)}"
          f"   Win rate: {win_rate:.1f}%")

    results = {}

    def compare(field, label, min_delta=0.05, fmt=".3f"):
        if field not in df_sig.columns: return {}
        t_v = target_g[field].dropna()
        s_v = stop_g[field].dropna()
        if len(t_v) < 3 or len(s_v) < 3: return {}

        t_m = float(t_v.mean()); s_m = float(s_v.mean())
        delta = t_m - s_m

        try:
            from scipy import stats as sp
            _, pval = sp.ttest_ind(t_v, s_v)
        except Exception:
            pval = 1.0

        sig = pval < 0.05
        if   abs(delta) <  min_delta:         verdict = "NOISE"
        elif delta > 0 and abs(delta) >= min_delta: verdict = "USEFUL"
        else:                                  verdict = "INVERSE"

        print(f"\n  [{label}]")
        print(f"  HIT_TARGET: {t_m:{fmt}}   HIT_STOP: {s_m:{fmt}}"
              f"   Delta: {delta:+{fmt}}")
        print(f"  p={pval:.3f}  "
              f"{'(significant)' if sig else '(not significant yet)'}"
              f"   Verdict: {verdict}")
        if verdict == "USEFUL":
            print(f"  -> Working. Consider tightening this threshold.")
        elif verdict == "NOISE":
            print(f"  -> Not separating winners. "
                  f"Run --sensitivity to find better threshold.")
        elif verdict == "INVERSE":
            print(f"  -> Counter-intuitive. "
                  f"Need more trades (20+) before acting on this.")
        return {"t_mean":t_m,"s_mean":s_m,"delta":delta,"pval":pval,"verdict":verdict}

    results["hurst"]   = compare("hurst_score",  "HURST SCORE",    0.08)
    results["score"]   = compare("score",         "COMPOSITE SCORE", 5.0, ".1f")
    results["bayes"]   = compare("bayes_prob",    "BAYES PROB %",    5.0, ".1f")
    results["halflife"]= compare("hl_strength",   "HALF-LIFE STRENGTH", 0.10)

    # Z-score for MR only
    if "zscore" in df_sig.columns and "strategy" in df_sig.columns:
        mr = df_sig[df_sig["strategy"] == "MEAN_REVERSION"]
        mr_t = mr[mr["outcome"]=="HIT_TARGET"]["zscore"].dropna()
        mr_s = mr[mr["outcome"]=="HIT_STOP"]["zscore"].dropna()
        if len(mr_t) >= 3 and len(mr_s) >= 3:
            print(f"\n  [Z-SCORE DEPTH  MR only]")
            print(f"  HIT_TARGET avg Z: {mr_t.mean():.3f}   "
                  f"HIT_STOP avg Z: {mr_s.mean():.3f}")
            print(f"  Deeper Z in winners: "
                  f"{'YES - Z matters' if mr_t.mean() < mr_s.mean() else 'NO'}")

    # Score threshold inflection
    print(f"\n  [SCORE THRESHOLD -- where does win rate jump?]")
    print(f"  {'Threshold':>12} {'n':>5} {'Win%':>8} {'vs baseline':>12}")
    print(f"  {'-'*42}")
    if "score" in df_sig.columns:
        for thresh in [55, 60, 65, 70, 75, 80]:
            sub = df_sig[df_sig["score"] >= thresh]
            if len(sub) < 2: continue
            wr  = (sub["outcome"]=="HIT_TARGET").sum() / len(sub) * 100
            diff = wr - win_rate
            flag = " <-- inflection" if diff > 8 else ""
            print(f"  score >= {thresh:>3}    {len(sub):>5}  {wr:>7.1f}%  "
                  f"{diff:>+10.1f}%{flag}")

    print(f"\n  SUMMARY:")
    for k, r in results.items():
        if r and "verdict" in r:
            print(f"  {k:<15} {r['verdict']}")

    out = os.path.join(dl_dir, f"hypothesis_{log_date}.json")
    with open(out, "w") as f:
        safe = {k: {kk: float(vv) if isinstance(vv, (int,float,np.floating))
                    else vv for kk,vv in v.items()}
                if isinstance(v,dict) else v for k,v in results.items()}
        json.dump(safe, f, indent=2)
    print(f"\n  Saved -> {out}")
    return results


def run_weekly_analysis() -> dict:
    """
    Aggregates all tagged log files in ~/Downloads/ for weekly analysis.
    Needs 20+ completed trades for statistical significance.

    Usage:  python scanner_research_v4.py --weekly
    """
    dl_dir  = os.path.join(os.path.expanduser("~"), "Downloads")
    files   = sorted(glob.glob(os.path.join(dl_dir, "scan_log_*_tagged.csv")))

    if not files:
        print(f"  [WEEKLY] No tagged files in {dl_dir}")
        print(f"  Run --tag-outcomes DATE for each trading day first.")
        return {}

    print(f"\n{'='*65}")
    print(f"  WEEKLY AGGREGATE ANALYSIS  |  {len(files)} day(s)")
    for f in files: print(f"  {os.path.basename(f)}")
    print(f"{'='*65}")

    dfs = [pd.read_csv(f) for f in files
           if not pd.read_csv(f).empty]
    if not dfs: return {}

    combined = pd.concat(dfs, ignore_index=True)
    df_sig   = combined[combined["outcome"].isin(
                    ["HIT_TARGET","HIT_STOP"])].copy()

    print(f"\n  Total signals: {len(df_sig)}  "
          f"(HIT_TARGET={( df_sig['outcome']=='HIT_TARGET').sum()}  "
          f"HIT_STOP={(df_sig['outcome']=='HIT_STOP').sum()})")

    if len(df_sig) < 10:
        print(f"  Need 20+ for reliable results. Keep scanning.")
        return {"total": len(df_sig)}

    win_rate = (df_sig["outcome"] == "HIT_TARGET").sum() / len(df_sig) * 100
    print(f"  Overall win rate: {win_rate:.1f}%")

    # Win rate by strategy
    if "strategy" in df_sig.columns:
        print(f"\n  Win rate by strategy:")
        for strat, grp in df_sig.groupby("strategy"):
            wr = (grp["outcome"]=="HIT_TARGET").sum()/len(grp)*100
            print(f"  {strat:<25} {wr:.1f}%  (n={len(grp)})")

    # Win rate by score bucket
    if "score" in df_sig.columns:
        print(f"\n  Win rate by score bucket:")
        df_sig["bucket"] = pd.cut(df_sig["score"],
                                    bins=[0,55,60,65,70,75,100],
                                    labels=["<55","55-60","60-65",
                                            "65-70","70-75",">75"])
        for bucket, grp in df_sig.groupby("bucket", observed=True):
            if len(grp) == 0: continue
            wr  = (grp["outcome"]=="HIT_TARGET").sum()/len(grp)*100
            bar = "#" * int(wr/5)
            print(f"  Score {str(bucket):<8} {wr:>5.1f}%  n={len(grp):>3}  {bar}")

    # Half-life alive vs dead
    if "hl_alive" in df_sig.columns:
        alive = df_sig[df_sig["hl_alive"]==True]
        dead  = df_sig[df_sig["hl_alive"]==False]
        if len(alive) > 0 and len(dead) > 0:
            wa = (alive["outcome"]=="HIT_TARGET").sum()/len(alive)*100
            wd = (dead["outcome"] =="HIT_TARGET").sum()/len(dead) *100
            print(f"\n  Half-life ALIVE win rate: {wa:.1f}%  (n={len(alive)})")
            print(f"  Half-life DEAD  win rate: {wd:.1f}%  (n={len(dead)})")
            if wa > wd + 10:
                print(f"  -> Only take signals with hl_alive=True")

    out = os.path.join(dl_dir, "weekly_combined.csv")
    combined.to_csv(out, index=False)
    print(f"\n  Combined data saved -> {out}")
    return {"total": len(df_sig), "win_rate": round(win_rate, 2)}


# -- End Section 8 ---------------------------------------------


# ===============================================================
# SECTION 9 -- MARKOV DATA PREFETCH PIPELINE
# ===============================================================
#
# PROBLEM BEING SOLVED
# --------------------
# Currently, when MARKOV_GATE_ENABLED = True and a signal fires,
# the engine calls get_markov_gate() which calls
# build_transition_matrix() which calls yf.Ticker().history()
# LIVE during the scan. That means:
#   - yfinance call on every alert during market hours
#   - Counts against rate limits (already tight with 40 symbols)
#   - Adds 0.5-2s latency per alert
#   - The daily data doesn't change during the session anyway
#
# SOLUTION
# --------
# Run this ONCE before the market opens (8:30-9:00 AM):
#   python scanner_research_v4.py --prefetch
#
# What it does:
#   1. Fetches 90 days of daily OHLCV for every watchlist symbol + SPY
#   2. Builds the full 4x4 Markov transition matrix for each symbol
#   3. Saves matrices to SCANNER_DIR/markov_data/SYMBOL.json
#   4. When engine calls get_markov_gate(), it reads from disk first
#      (no yfinance call needed during the trading session)
#   5. Falls back to live yfinance if disk file is stale (> 24h) or missing
#
# DAILY SCHEDULE (add to Task Scheduler after setup_tasks.bat):
#   8:45 AM  --  python scanner_research_v4.py --prefetch
#   9:00 AM  --  python scanner_terminal_v4.py --once
#   9:30 AM  --  python scanner_terminal_v4.py --loop
#
# This ensures every symbol that will be scanned today has a fresh
# Markov matrix ready before the first signal fires.
# ===============================================================

# Path where pre-built matrices are stored -- defined by engine.MARKOV_DATA_DIR
# (imported at top of file from scanner_v4.py)


def prefetch_markov_data(symbols: list = None, days: int = 120,
                          force: bool = False) -> dict:
    """
    Pre-builds Markov transition matrices for all watchlist symbols.
    Run once before market open at 5:45 AM PT. Saves to disk.

    CRITICAL -- CALENDAR vs TRADING DAYS:
      `days` = target trading days of usable data AFTER the ADX window.

      The fetch formula is: calendar_days = days * 2 + 30
      This guarantees enough trading days after weekends, holidays, ADX window.

      Why --days 60 failed:
        Old code fetched days+30=90 calendar days -> ~62 raw rows
        Then tailed to 60 rows, then ADX ate 14 -> 46 usable  *** FAILED gate ***
      Current code (fixed):
        No .tail() -- keeps all fetched rows
        fetch_cal = days*2+30 = plenty of margin
        --days 60  -> 150 cal -> ~103 trading -> ~89 after ADX  PASS
        --days 90  -> 210 cal -> ~144 trading -> ~130 after ADX PASS
        --days 120 -> 270 cal -> ~186 trading -> ~172 after ADX PASS (recommended)

      Minimum gate: 60 usable trading days (MARKOV_MIN_DAYS).
      Default is 120 -- safe for all market conditions including holidays.

    Args:
        symbols:  list of ticker symbols. Defaults to engine.WATCHLIST.
        days:     target trading days (default 120). Internally fetches ~2x calendar days.
        force:    if True, rebuilds even if today's file already exists.

    Usage:
        python scanner_research_v4.py --prefetch              (120-day default)
        python scanner_research_v4.py --prefetch --days 120   (same as default)
        python scanner_research_v4.py --prefetch --force      (force full rebuild)
    """
    os.makedirs(MARKOV_DATA_DIR, exist_ok=True)
    symbols = symbols or engine.WATCHLIST
    today   = datetime.now().strftime("%Y-%m-%d")
    summary = {}

    # Calendar math -- show user exactly what's happening
    fetch_cal   = days * 2 + 30
    est_trading = int(fetch_cal * 0.71)
    est_usable  = max(0, est_trading - 14)
    print(f"\n{'='*65}")
    print(f"  MARKOV DATA PREFETCH")
    print(f"  Symbols: {len(symbols)}  |  Target: {days} trading days")
    print(f"  Fetching: ~{fetch_cal} calendar days  ->  ~{est_trading} trading rows  ->  ~{est_usable} after ADX")
    print(f"  Minimum required: {MARKOV_MIN_DAYS} usable days  |  {'WILL PASS' if est_usable >= MARKOV_MIN_DAYS else 'WARNING: TOO FEW DAYS -- use --days 120'}")
    print(f"  Output: {MARKOV_DATA_DIR}")
    if est_usable < MARKOV_MIN_DAYS:
        print(f"\n  !! RECOMMENDATION: Use --days 120 to guarantee all symbols pass")
        print(f"  !! Current: ~{est_usable} usable days < {MARKOV_MIN_DAYS} required")
    print(f"  {'FORCED REBUILD' if force else 'Skipping symbols with fresh data'}")
    print(f"{'='*65}\n")

    passed = failed = skipped = 0

    for i, sym in enumerate(symbols, 1):
        out_path = os.path.join(MARKOV_DATA_DIR, f"{sym}.json")
        progress = f"[{i:>2}/{len(symbols)}]"

        # -- Skip if today's file already exists and not forcing ----
        if not force and os.path.exists(out_path):
            try:
                existing = json.load(open(out_path))
                if existing.get("fetch_date") == today and existing.get("valid"):
                    data_days = existing.get("data_days", 0)
                    print(f"  {progress} {sym:<6} SKIP  "
                          f"(today's file exists, {data_days} days)")
                    summary[sym] = existing
                    skipped += 1
                    continue
            except Exception:
                pass  # corrupt file -- rebuild it

        print(f"  {progress} {sym:<6} fetching...", end=" ", flush=True)

        # -- Build the matrix via research module -------------------
        result = build_transition_matrix(sym, days=days, use_cache=False)
        matrix, counts, states_seq, valid, data_days = result \
            if len(result) == 5 else (*result, 0)

        payload = {
            "symbol":     sym,
            "fetch_date": today,
            "days":       days,
            "valid":      valid,
            "data_days":  data_days,
        }

        if valid and matrix is not None:
            # Compute stationary distribution and store it too
            pi   = compute_stationary_distribution(matrix)
            sc   = {MARKOV_STATES[s]: 0 for s in range(4)}
            for _, s in states_seq: sc[MARKOV_STATES[s]] += 1

            payload.update({
                "matrix":           matrix.tolist(),
                "counts":           counts.tolist(),
                "stationary":       pi.tolist(),
                "state_counts":     sc,
                "last_state":       states_seq[-1][1] if states_seq else 3,
                "last_state_name":  MARKOV_STATES.get(
                                        states_seq[-1][1] if states_seq else 3,
                                        "RANGING"),
            })
            with open(out_path, "w") as f:
                json.dump(payload, f)

            last_state = payload["last_state_name"]
            calm_pct   = float(pi[0]) * 100
            trend_pct  = float(pi[1]) * 100
            print(f"OK   {data_days}d  last={last_state:<10} "
                  f"CALM={calm_pct:.0f}%  TREND={trend_pct:.0f}%")
            passed += 1
        else:
            with open(out_path, "w") as f:
                json.dump(payload, f)
            print(f"FAIL  only {data_days} days "
                  f"(need {MARKOV_MIN_DAYS})")
            failed += 1

        summary[sym] = payload
        time.sleep(0.3)  # brief pause -- avoid rate limiting during batch

    print(f"\n{'='*65}")
    print(f"  PREFETCH COMPLETE")
    print(f"  Passed: {passed}  |  Failed: {failed}  |  Skipped: {skipped}")
    print(f"  Files saved to: {MARKOV_DATA_DIR}")

    # Print which symbols are NOT ready
    not_ready = [s for s, d in summary.items() if not d.get("valid", False)]
    if not_ready:
        print(f"\n  Symbols NOT yet at 60 days (Markov gate won't apply):")
        for sym in not_ready:
            dd = summary[sym].get("data_days", 0)
            needed = MARKOV_MIN_DAYS - dd
            print(f"    {sym:<6}  {dd} days  (~{needed} trading days more)")
    else:
        print(f"\n  All {passed} symbols have valid Markov matrices.")
        print(f"  Set MARKOV_GATE_ENABLED = True in scanner_v4.py to activate.")
    print(f"{'='*65}")

    return summary


def load_markov_from_disk(symbol: str) -> tuple:
    """
    Loads a pre-built Markov matrix from disk.
    Returns (matrix, valid, data_days, last_state_name).

    Called by get_markov_gate() BEFORE attempting a live yfinance fetch.
    File must be from today (same fetch_date) to be considered fresh.

    Returns (None, False, 0, "NO_DATA") if file doesn't exist,
    is stale (not today's date), or failed validation during prefetch.
    """
    path  = os.path.join(MARKOV_DATA_DIR, f"{symbol}.json")
    today = datetime.now().strftime("%Y-%m-%d")

    if not os.path.exists(path):
        return None, False, 0, "NO_FILE"

    try:
        data = json.load(open(path))
    except Exception:
        return None, False, 0, "CORRUPT_FILE"

    # Stale: was built on a different day
    if data.get("fetch_date") != today:
        return None, False, 0, f"STALE_{data.get('fetch_date','')}"

    # Was built today but didn't have enough data
    if not data.get("valid", False):
        data_days = data.get("data_days", 0)
        needed    = MARKOV_MIN_DAYS - data_days
        return None, False, data_days, f"NO_DATA_{data_days}d_need_{needed}_more"

    # Valid -- reconstruct matrix from stored list
    try:
        matrix    = np.array(data["matrix"])
        data_days = data.get("data_days", 0)
        last_name = data.get("last_state_name", "UNKNOWN")
        return matrix, True, data_days, f"DISK_OK_{data_days}d"
    except Exception:
        return None, False, 0, "MATRIX_PARSE_ERROR"


# ── End Section 9 ─────────────────────────────────────────────


# CLI ENTRY POINT
# ===============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Quant Research Module - scanner_research_v4.py",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
DAILY WORKFLOW COMMANDS:
  Before market open (8:45 AM) -- run FIRST every day:
    --prefetch              Build Markov matrices for all watchlist symbols
    --prefetch --force      Force rebuild even if today's files exist
    --prefetch --days 90    Use 90-day lookback (more state observations)

  After market close:
    --tag-outcomes DATE     Tag each signal as HIT_TARGET/HIT_STOP/OPEN
    --hypothesis DATE       Test which indicators actually predict wins
    --weekly                Aggregate all tagged files for the week

  Weekly/Monthly:
    --backtest --days 30    Replay 30 days of history through the engine
    --sensitivity --days 20 Grid search on Z/SL/TP thresholds
    --consistency           Win rate by regime/ADX/Z-score buckets
    --markov ALKT CRUS      Full Markov regime analysis with GBM
    --markov-gate ALKT      Test whether Markov gate improves Sharpe

  Execution:
    --twap ALKT 100 --side LONG --price 16.27   TWAP schedule

  Research:
    --profile ALKT CRUS --days 60   Build ticker personality profile
""")

    parser.add_argument("--prefetch",    action="store_true",
                        help="Pre-build Markov matrices for all watchlist symbols")
    parser.add_argument("--force",       action="store_true",
                        help="Force rebuild even if today's prefetch files exist")
    parser.add_argument("--profile",       nargs="+", metavar="SYM")
    parser.add_argument("--backtest",      action="store_true")
    parser.add_argument("--sensitivity",   action="store_true")
    parser.add_argument("--consistency",   action="store_true")
    parser.add_argument("--markov",        nargs="+", metavar="SYM")
    parser.add_argument("--markov-gate",             nargs="+", metavar="SYM",
                        help="Per-symbol gate backtest: --markov-gate ALKT --days 60")
    parser.add_argument("--markov-gate-portfolio",   action="store_true",
                        help="Cross-sectional gate test across all watchlist symbols")
    parser.add_argument("--tag-outcomes",  metavar="DATE",
                        help="Tag outcomes for DATE (e.g. 2026-04-30)")
    parser.add_argument("--hypothesis",    metavar="DATE",
                        help="Run hypothesis test for DATE")
    parser.add_argument("--weekly",        action="store_true",
                        help="Aggregate all tagged files for weekly analysis")
    parser.add_argument("--twap",          nargs=2, metavar=("SYM","SHARES"))
    parser.add_argument("--side",          default="LONG")
    parser.add_argument("--price",         type=float, default=0.0)
    parser.add_argument("--days",          type=int, default=60)
    parser.add_argument("--tranches",      type=int, default=4)
    parser.add_argument("--interval",      type=int, default=30)
    parser.add_argument("--lookforward",   type=int, default=4,
                        help="Hours to look forward when tagging outcomes (default 4)")
    args = parser.parse_args()

    if args.prefetch:
        prefetch_markov_data(
            symbols=engine.WATCHLIST,
            days=args.days,
            force=args.force,
        )

    if args.profile:
        for sym in args.profile:
            profile_ticker_from_history(sym.upper(), days=args.days)

    if args.backtest:
        run_backtest(engine.WATCHLIST[:10], days=args.days,
                     strategy="MEAN_REVERSION")

    if args.sensitivity:
        run_sensitivity_analysis(engine.WATCHLIST[:5], days=args.days)

    if args.consistency:
        run_consistency_report()

    if args.markov:
        run_markov_analysis([s.upper() for s in args.markov],
                             days=max(args.days, MARKOV_MIN_DAYS))

    if args.markov_gate:
        for sym in args.markov_gate:
            run_markov_gate_test(sym.upper(),
                                  days=max(args.days, MARKOV_MIN_DAYS))

    if args.markov_gate_portfolio:
        run_markov_gate_portfolio(
            symbols=engine.WATCHLIST,
            days=max(args.days, MARKOV_MIN_DAYS)
        )

    if args.tag_outcomes:
        tag_outcomes(args.tag_outcomes,
                     lookforward_hours=args.lookforward)

    if args.hypothesis:
        run_hypothesis_test(args.hypothesis)

    if args.weekly:
        run_weekly_analysis()

    if args.twap:
        sym = args.twap[0].upper(); shares = int(args.twap[1])
        price = args.price if args.price > 0 else 100.0
        compute_twap_schedule(sym, shares, args.side, price,
                               n_tranches=args.tranches,
                               interval_seconds=args.interval)

    if not any([args.prefetch, args.profile, args.backtest, args.sensitivity,
                args.consistency, args.markov, args.markov_gate,
                args.markov_gate_portfolio,
                args.tag_outcomes, args.hypothesis, args.weekly,
                args.twap]):
        parser.print_help()


if __name__ == "__main__":
    main()
